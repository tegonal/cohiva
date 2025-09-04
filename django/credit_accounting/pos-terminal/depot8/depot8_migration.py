#!/usr/bin/env python3

import hashlib
import json
import os
import sys
from dataclasses import dataclass, field
from datetime import datetime, timedelta

import openpyxl as pxl
import pytz

NOTE_DEPOSIT = "Einzahlung Depot8"
NOTE_PURCHASE = "Einkauf Depot8"
DATE_FORMAT = "%y/%m/%d"
timezone = pytz.timezone("Europe/Zurich")
timezone_filenames = pytz.utc

workbook = pxl.load_workbook("/tmp/data.xlsx")
gutschriften_sheet = workbook["Gutschriften"]
einkaeufe_sheet = workbook["Einkaeufe"]

date_history_path = "/tmp/depot8-history/hourly"
date_history_file = "/tmp/depot8_date_history.json"
date_history_start = "data-2022-05-21-22.xlsx"


@dataclass
class Transaction:
    amount: float
    date: datetime
    id: str = ""


@dataclass
class Account:
    name: str = ""
    name_coordinate: str = ""
    pin: str = ""
    id: int = -1
    deposits: list[Transaction] = field(default_factory=list)
    purchases: list[Transaction] = field(default_factory=list)

    def print_info(self):
        deposits_total = sum([deposit.amount for deposit in self.deposits])
        purchases_total = sum([purchase.amount for purchase in self.purchases])
        print(f"Name:\t{self.name}")
        print(f"Pin:\t{self.pin}")
        print(f"Id:\t{self.id}")
        print(f"{len(self.deposits)} deposits with total of {deposits_total:.2f}")
        print(f"{len(self.purchases)} purchases with total of {purchases_total:.2f}")
        print(f"Saldo: {(deposits_total + purchases_total):.2f}")


def get_accounts() -> list:
    dates = {}
    accounts = []
    fix_invalid_dates = {
        "2306/20": "23/06/20",
        "23/0822": "23/08/22",
        "21/02/09": "24/02/09",
        "21/0216": "24/02/16",
        "24/0328": "24/03/28",
        "240516": "24/05/16",
    }
    for column in gutschriften_sheet.iter_cols():
        account = Account()
        for cell in column:
            if cell.column == 1:
                if cell.row > 4:
                    if cell.value:
                        date_str = str(cell.value)
                        if date_str in fix_invalid_dates:
                            date_str = fix_invalid_dates[date_str]
                        date = datetime.strptime(date_str, DATE_FORMAT).replace(tzinfo=timezone)
                    else:
                        date = datetime(2022, 5, 21, tzinfo=timezone)
                        print(
                            f"NO date found for deposit {cell.column}_{cell.row}_{cell.value}. "
                            f"Using fixed start date: {date}"
                        )
                    dates[cell.row] = date
            else:
                if cell.row == 1 and cell.value:
                    account.name = str(cell.value)
                    account.name_coordinate = f"Gutschriften!{cell.coordinate}"
                elif cell.row == 2:
                    account.pin = str(cell.value)
                elif cell.row == 3 or cell.row == 4:
                    continue
                else:
                    if cell.value:
                        account.deposits.append(
                            Transaction(
                                date=dates[cell.row],
                                amount=float(cell.value),
                                id=f"import_dep_{cell.column}_{cell.row}_{cell.value}",
                            )
                        )
        if len(account.name):
            accounts.append(account)
    return accounts


def add_purchases(accounts: list[Account], date_history: dict[str]) -> list[Account]:
    for column in einkaeufe_sheet.iter_cols(
        min_col=2, max_col=einkaeufe_sheet.max_column, min_row=1, max_row=einkaeufe_sheet.max_row
    ):
        date = None
        last_d = None
        last_d_count = 0
        for cell in column:
            if cell.row == 1:
                if cell.value:
                    account = [a for a in accounts if str(cell.value).endswith(a.name_coordinate)][
                        0
                    ]
            elif cell.row > 2:
                if account and cell.value:
                    hist_id = f"hist_{cell.column}_{cell.row}_{cell.value}"
                    if hist_id in date_history:
                        d = date_history[hist_id].split("-")
                        date = datetime(
                            int(d[0]), int(d[1]), int(d[2]), int(d[3]), tzinfo=timezone_filenames
                        )
                        if d == last_d:
                            last_d_count += 1
                            print(
                                f"Multiple purchases with same date {d} for "
                                f"{account.name}/{hist_id}. Adding {last_d_count} seconds."
                            )
                            date += timedelta(0, last_d_count)
                        else:
                            last_d_count = 0
                        last_d = d
                    elif date:
                        date += timedelta(0, 1)
                        print(
                            f"NO date in date_history found for {account.name}/{hist_id}. "
                            f"Using last date + 1second: {date}"
                        )
                    else:
                        date = datetime(2022, 5, 21, 0, tzinfo=timezone)
                        print(
                            f"NO date in date_history found for {account.name}/{hist_id}. "
                            f"Using FIXED start date: {date}"
                        )
                    # minutes = cell.row // 60
                    # seconds = cell.row % 60
                    account.purchases.append(
                        Transaction(
                            date=date,
                            amount=float(cell.value),
                            id=f"import_pur_{cell.column}_{cell.row}_{cell.value}",
                        )
                    )

    return accounts


def create_account(name: str, pin: str) -> int:
    account_id = -1
    print(f"  :: Create new account for {name}:{pin} with account id {account_id}")
    return account_id


def add_transaction(account: int, amount: float, date: datetime, note: str):
    print(f"  :: Add transaction to account {account}: {amount}, {date}, {note}")


def read_history_file(file: str, data: dict):
    hist_workbook = pxl.load_workbook(f"{date_history_path}/{file}")
    hist_einkaeufe_sheet = hist_workbook["Einkaeufe"]
    for column in hist_einkaeufe_sheet.iter_cols(
        min_col=2,
        max_col=hist_einkaeufe_sheet.max_column,
        min_row=1,
        max_row=hist_einkaeufe_sheet.max_row,
    ):
        for cell in column:
            if cell.row > 2:
                if cell.value:
                    cell_id = f"hist_{cell.column}_{cell.row}_{cell.value}"
                    if cell_id not in data:
                        data[cell_id] = file[11:-5]


def update_date_history():
    ## Load data and filenames
    try:
        with open(date_history_file) as input_file:
            date_history = json.load(input_file)
            last_known_file = date_history["last_known_file"]
    except FileNotFoundError:
        date_history = {}
        last_known_file = None
        print("No previous data found.")
    files = []
    for file in os.listdir(date_history_path):
        if file.startswith("abrechnung-") and file.endswith(".xlsx"):
            files.append(file)
        else:
            print(f"Ignoring unknown file {file}.")
    files.sort()

    ## Process files
    if last_known_file:
        last_known_found = False
    else:
        last_known_found = True
    start_found = False
    last_md5 = None
    count = 0
    for file in files:
        # "abrechnung-2024-03-04-05.xlsx"
        ## Skip files until start file
        if not start_found:
            if file == date_history_start:
                start_found = True
            else:
                # print(f"Skip before start: {file}.")
                continue
        ## Skip files that we have already imported
        if last_known_file and not last_known_found:
            if file == last_known_file:
                last_known_found = True
            # print(f"Skip known file: {file}.")
            continue
        ## Process new files
        try:
            md5 = hashlib.md5(open(f"{date_history_path}/{file}", "rb").read()).hexdigest()
            if not last_md5 or md5 != last_md5:
                print(f"Processing file {file}.")
                read_history_file(file, date_history)
                count += 1
        except Exception as e:
            print(f"Error while processing file {file}: {e}. Stopping after {count} files.")
            break
        # else:
        #    print(f"Skip identical md5: {file}.")
        last_known_file = file
        last_md5 = md5
        if count >= 500:
            print(f"Stopping after {count} files.")
            break

    ## Save data
    date_history["last_known_file"] = last_known_file
    with open(date_history_file, "w") as outfile:
        outfile.write(json.dumps(date_history))

    return date_history


def import_accounts():
    date_history = update_date_history()
    return add_purchases(get_accounts(), date_history)


def main():
    date_history = update_date_history()
    accounts = add_purchases(get_accounts(), date_history)
    for account in accounts:
        account.print_info()

    sys.exit()

    for account in accounts:
        # Migrate accounts
        account.id = create_account(name=account.name, pin=account.pin)

        # Migrate Einzahlungen
        for deposit in account.deposits:
            add_transaction(
                account=account.id, amount=deposit.amount, date=deposit.date, note=NOTE_DEPOSIT
            )

        # Migrate Einkaeufe
        for purchase in account.purchases:
            add_transaction(
                account=account.id, amount=purchase.amount, date=purchase.date, note=NOTE_PURCHASE
            )


if __name__ == "__main__":
    main()
