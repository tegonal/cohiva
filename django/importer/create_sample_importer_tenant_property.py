"""
Create a sample Excel file for Tenant and Property import.

This script generates a sample Excel file with the correct column structure
for importing tenant and property data into the system.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def create_sample_tenant_property_excel(filename="sample_tenant_property_import.xlsx"):
    """
    Create a sample Excel file for tenant and property import.

    Args:
        filename: Output filename for the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Tenant Property Import"

    # Define headers
    headers = [
        "Liegenschaft Bezeichnung",
        "Liegenschaft Nummer",
        "Liegenschaft Strasse",
        "Liegenschaft Postleitzahl",
        "Liegenschaft Ort",
        "Liegenschaft Gebaeudeideidg",
        "Objekttyp Bezeichnung",
        "Nummer",
        "Bezeichnung",
        "Etage Nummer",
        "Zimmerzahl Anzahl",
        "Nutzflaeche",
        "Akonti",
        "Pauschalen",
        "Brutto",
        "Nettomiete",
        "Position",
        "Wohnung Id Eidg",
        "Wohnung Nummer Administrativ",
        "Id",
        "Mieter Person Person Name Vorname",
        "Mieter Person Person Adresse",
        "Mieter Person Person Id",
        "Verhaeltnis Gueltigvon",
        "Verhaeltnis Gueltigbis",
        "Mieter Bezeichnung",
        "Mieter Nummer",
    ]

    # Style for headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Sample data rows
    sample_data = [
        [
            "Musterhaus A",  # Liegenschaft Bezeichnung
            "1001",  # Liegenschaft Nummer
            "Hauptstrasse 10",  # Liegenschaft Strasse
            "3011",  # Liegenschaft Postleitzahl
            "Bern",  # Liegenschaft Ort
            "123456",  # Liegenschaft Gebaeudeideidg
            "Wohnung",  # Objekttyp Bezeichnung
            "1.1",  # Nummer
            "Erdgeschoss links",  # Bezeichnung
            "1",  # Etage Nummer
            "3.5",  # Zimmerzahl Anzahl
            "85.5",  # Nutzflaeche
            "150.00",  # Akonti
            "",  # Pauschalen
            "",  # Brutto
            "1200.00",  # Nettomiete
            "links",  # Position
            "",  # Wohnung Id Eidg
            "",  # Wohnung Nummer Administrativ
            "",  # Id
            "Muster Max",  # Mieter Person Person Name Vorname
            "Testweg 5, 3011 Bern",  # Mieter Person Person Adresse
            "1001",  # Mieter Person Person Id
            "2024-01-01",  # Verhaeltnis Gueltigvon
            "",  # Verhaeltnis Gueltigbis
            "Hauptmieter",  # Mieter Bezeichnung
            "5001",  # Mieter Nummer
        ],
        [
            "Musterhaus A",  # Same building
            "1001",
            "Hauptstrasse 10",
            "3011",
            "Bern",
            "123456",
            "Wohnung",
            "1.2",  # Different unit
            "Erdgeschoss rechts",
            "1",
            "2.5",
            "65.0",
            "120.00",
            "",
            "",
            "950.00",
            "rechts",
            "",
            "",
            "",
            "",  # No tenant
            "",
            "",
            "",
            "",
            "",
            "",
        ],
        [
            "Musterhaus A",
            "1001",
            "Hauptstrasse 10",
            "3011",
            "Bern",
            "123456",
            "Parkplatz",
            "P.1",
            "Parkplatz Einfahrt",
            "",
            "",
            "",
            "",
            "",
            "",
            "80.00",
            "",
            "",
            "",
            "",
            "Anna Schmidt",
            "Bergstrasse 12, 3000 Bern",
            "1002",
            "2024-03-01",
            "",
            "",
            "5002",
        ],
        [
            "Gebäude B",  # Different building
            "2001",
            "Seestrasse 25",
            "8000",
            "Zürich",
            "",
            "Wohnung",
            "2.1",
            "Obergeschoss",
            "2",
            "4.5",
            "110.0",
            "200.00",
            "",
            "",
            "1800.00",
            "",
            "",
            "",
            "",
            "",  # Vacant (Leerstand)
            "",
            "",
            "",
            "",
            "",
            "9999",  # Special marker for Leerstand
        ],
    ]

    # Write sample data
    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-size columns
    from openpyxl.utils import get_column_letter

    for col_idx, header in enumerate(headers, start=1):
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = max(len(header) + 2, 12)  # Minimum width of 12

    # Save workbook
    wb.save(filename)
    print(f"Sample Excel file created: {filename}")
    print(f"  - {len(sample_data)} sample rows")
    print(f"  - {len(headers)} columns")
    print("\nColumn descriptions:")
    print("  - Liegenschaft: Building information")
    print("  - Objekttyp: Rental unit type (Wohnung, Parkplatz, Hobbyraum, etc.)")
    print("  - Nummer: Rental unit number/identifier")
    print("  - Mieter Nummer: 9999 = Leerstand (vacant)")
    print("  - Verhaeltnis Gueltigvon: Contract start date (YYYY-MM-DD)")


if __name__ == "__main__":
    create_sample_tenant_property_excel()
