"""
Business logic for processing Excel imports.
"""

import logging

import openpyxl
from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils.translation import gettext_lazy as _

from .models import ImportJob, ImportRecord

logger = logging.getLogger(__name__)


class ExcelImporter:
    """Service class for processing Excel file imports."""

    def __init__(self, import_job: ImportJob):
        """
        Initialize the importer with an ImportJob.

        Args:
            import_job: The ImportJob instance to process
        """
        self.import_job = import_job
        self.workbook = None
        self.worksheet = None

    def process(self) -> dict:
        """
        Process the import job.

        Returns:
            dictionary with processing results
        """
        try:
            self.import_job.status = "processing"
            self.import_job.save()

            # Load the Excel file
            self._load_workbook()

            # Extract data from the worksheet
            data = self._extract_data()

            # Process each row
            results = self._process_rows(data)

            # Update job status
            self.import_job.status = "completed"
            self.import_job.records_imported = results["success_count"]
            self.import_job.result_data = results
            self.import_job.save()

            logger.info(
                f"Import job {self.import_job.id} completed successfully. "
                f"Imported {results['success_count']} records."
            )

            return results

        except Exception as e:
            logger.exception(f"Error processing import job {self.import_job.id}")
            self.import_job.status = "failed"
            self.import_job.error_message = str(e)
            self.import_job.save()
            raise

    def _load_workbook(self):
        """Load the Excel workbook and select the first worksheet."""
        try:
            self.workbook = openpyxl.load_workbook(
                self.import_job.file.path, read_only=True, data_only=True
            )
            self.worksheet = self.workbook.active
        except Exception as e:
            raise ValidationError(f"Failed to load Excel file: {str(e)}")

    def _extract_data(self) -> list[dict]:
        """
        Extract data from the worksheet.

        Returns:
            List of dictionaries, each representing a row
        """
        data = []
        rows = list(self.worksheet.iter_rows(values_only=True))

        if not rows:
            raise ValidationError("The Excel file is empty")

        # Assume first row is header
        headers = rows[0]
        if not headers:
            raise ValidationError("The Excel file has no headers")

        # Process data rows
        for row_idx, row in enumerate(rows[1:], start=2):
            row_data = {}
            for header, value in zip(headers, row):
                if header:  # Skip columns without headers
                    row_data[str(header).strip()] = value
            if any(row_data.values()):  # Skip completely empty rows
                row_data["_row_number"] = row_idx
                data.append(row_data)

        return data

    def _process_rows(self, data: list[dict]) -> dict:
        """
        Process extracted rows.

        Args:
            data: List of row dictionaries to process

        Returns:
            dictionary with processing statistics
        """
        results = {
            "total_rows": len(data),
            "success_count": 0,
            "error_count": 0,
            "errors": [],
        }

        for row_data in data:
            row_number = row_data.pop("_row_number")
            try:
                with transaction.atomic():
                    # Process the row (customize this based on your needs)
                    if not self.import_job.override_existing:
                        # check for existing records
                        self._has_existing(row_data)

                    self._process_single_row(row_data)

                    # Import executed and record success
                    ImportRecord.objects.create(
                        job=self.import_job,
                        row_number=row_number,
                        data=row_data,
                        success=True,
                    )
                    results["success_count"] += 1

            except Exception as e:
                logger.warning(
                    f"Error processing row {row_number} in job {self.import_job.id}: {str(e)}"
                )
                # Record error
                ImportRecord.objects.create(
                    job=self.import_job,
                    row_number=row_number,
                    data=row_data,
                    error_message=str(e),
                    success=False,
                )
                results["error_count"] += 1
                results["errors"].append({"row": row_number, "error": str(e), "data": row_data})

        return results

    def _has_existing(self, row_data: dict):
        """
        Check if a record already exists based on the row data.

        Override this method in subclasses to implement specific existence checks.

        Args:
            row_data: dictionary containing the row data
        Raises:
            ValidationError: If a record already exists
        """
        # Default implementation: no existence check
        pass

    def _process_single_row(self, row_data: dict):
        """
        Process a single row of data.

        Override this method in subclasses to implement specific business logic.

        Args:
            row_data: dictionary containing the row data

        Raises:
            ValidationError: If the row data is invalid
        """
        # Default implementation: just validate that row has some data
        if not row_data:
            raise ValidationError(str(_("Zeile hat keine Daten")))

        # Add your custom processing logic here
        # For example:
        # - Create or update database records
        # - Validate data
        # - Call external APIs
        # etc.

        logger.debug(f"Processing row data: {row_data}")


def process_import_job(import_job_id: int, importer_class=None) -> dict:
    """
    Process an import job by ID.

    This function can be called from views or as a Celery task.

    Args:
        import_job_id: ID of the ImportJob to process
        importer_class: Optional importer class to use (defaults to auto-detect based on import_type)

    Returns:
        dictionary with processing results
    """
    import_job = ImportJob.objects.get(id=import_job_id)

    if importer_class is None:
        # Auto-detect importer based on import_type
        if import_job.import_type == "member_address":
            from .importer_member_address import ImporterMemberAddress

            importer_class = ImporterMemberAddress
        elif import_job.import_type == "tenant_property":
            from .importer_tenant_property import ImporterTenantProperty

            importer_class = ImporterTenantProperty
        else:
            # Default fallback
            importer_class = ExcelImporter

    importer = importer_class(import_job)
    return importer.process()

