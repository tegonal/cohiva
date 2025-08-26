import datetime
import numbers
import string
from tempfile import NamedTemporaryFile

import vdirsyncer
import vobject
from django.conf import settings
from django.http import HttpResponse
from django.template import loader
from django.utils import timezone
from openpyxl import load_workbook
from vdirsyncer.storage.dav import CardDAVStorage

from .gnucash import get_reference_nr
from .models import Address, Invoice, RentalUnit, get_active_contracts


def vcard_from_address(adr):
    vcard = vobject.vCard()
    vcard.add("n")
    if adr.name:
        vcard.n.value = vobject.vcard.Name(family=adr.name, given=adr.first_name)
    elif adr.organization:
        vcard.n.value = vobject.vcard.Name(family=adr.organization)
    vcard.add("fn")
    if adr.organization:
        vcard.fn.value = adr.organization
        vcard.add("org").value = (adr.organization,)
    else:
        vcard.fn.value = "%s %s" % (adr.first_name, adr.name)
    if adr.city:
        vcard.add("adr")
        vcard.adr.value = vobject.vcard.Address(
            street=adr.street, city=adr.city, extended=adr.extra
        )
    if adr.email:
        vcard.add("email").value = adr.email
    if adr.email2:
        vcard.add("email").value = adr.email2
    if adr.telephone:
        if adr.telephone[0:2] == "00":
            tel = "+%s" % adr.telephone[2:]
        elif adr.telephone[0] == "0":
            tel = "+41 %s" % adr.telephone[1:]
        else:
            tel = adr.telephone
        vcard.add("tel").value = tel
    if adr.mobile:
        if adr.mobile[0:2] == "00":
            tel = "+%s" % adr.mobile[2:]
        elif adr.mobile[0] == "0":
            tel = "+41 %s" % adr.mobile[1:]
        else:
            tel = adr.mobile
        vcard.add("tel").value = tel
    if adr.date_birth:
        vcard.add("bday").value = adr.date_birth.strftime("%Y%m%d")
    return vcard


## Export addresses to a CardDAV server
def export_addresses_carddav(delete_all_first=False):
    ret = []

    if not settings.GENO_CARDDAV_URI:
        return [{"info": "ERROR: GENO_CARDDAV_URI is not set."}]

    cd = CardDAVStorage(
        url=settings.GENO_CARDDAV_URI,
        username=settings.GENO_CARDDAV_USER,
        password=settings.GENO_CARDDAV_PASS,
    )

    etags = {}
    for href, etag in cd.list():
        etags[href] = etag

    if delete_all_first:
        for href in etags:
            # Force deletion
            etag = etags[href]
            cd.delete(href, etag)
            ret.append("DELETE: %s [etag: %s]" % (href, etag))
        for adr in Address.objects.all():
            if adr.carddav_href:
                adr.carddav_href = ""
                adr.carddav_etag = ""
                adr.carddav_syncts = None
                adr.save()

    # return [ {"info": "Synced addresses:", "objects": ret } ]

    for adr in Address.objects.filter(active=True):
        if len(adr.name) and adr.name[0] == "*":
            ret.append("IGNORE: %s (special address)" % (adr))
            continue
        if not adr.email and not adr.email2 and not adr.telephone and not adr.mobile:
            ret.append("IGNORE: %s (no email/phone)" % (adr))
            continue

        update = False
        new = False
        if adr.carddav_href:
            if adr.carddav_href in etags:
                if (
                    etags[adr.carddav_href] != adr.carddav_etag
                    or adr.carddav_syncts + datetime.timedelta(seconds=10) < adr.ts_modified
                ):
                    update = True
            else:
                new = True
        else:
            new = True

        if update or new:
            vcard = vcard_from_address(adr).serialize()
            try:
                vcard_item = vdirsyncer.vobject.Item(vcard)
                if update:
                    etag = cd.update(adr.carddav_href, vcard_item, etags[adr.carddav_href])
                    tag = "UPDATE"
                    etag_str = "%s -> " % (etags[adr.carddav_href])
                else:
                    href, etag = cd.upload(vcard_item)
                    adr.carddav_href = href
                    tag = "NEW"
                    etag_str = ""
                if etag is None:
                    item, etag = cd.get(adr.carddav_href)
                    etag_str = "%s%s" % (etag_str, etag)
                if etag is None:
                    etag = ""
                adr.carddav_etag = etag
                adr.carddav_syncts = timezone.now()
                adr.save()
                # ret.append("%s: %s %s" % (tag, adr, vcard))
                ret.append("%s: %s [etag: %s]" % (tag, adr, etag_str))
                # break
            except vdirsyncer.exceptions.PreconditionFailed:
                ret.append("SYNC FAILED: %s [%s]" % (adr, vcard))

    return [{"info": "Synced addresses:", "objects": ret}]


def export_data_to_xls(data, title="Data export", header=None, filename_suffix="data"):
    if header is None:
        header = {}
    fields = data[0]._fields
    if isinstance(header, list):
        ## Convert header array to dict first
        new_header = {}
        for i in range(min(len(fields), len(header))):
            new_header[fields[i]] = header[i]
        header = new_header
    ## Set missing headers to field name
    for field in fields:
        if field not in header:
            header[field] = field
    return export_to_xls_generic(data, fields, title, header, filename_suffix)


def export_to_xls_generic(data, fields, title, header, filename_suffix):
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title

    row_num = 0

    ## Fill header
    max_width = []
    col_num = 0
    for field in fields:
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = header[field]
        max_width.append(len(c.value))
        c.font = Font(bold=True)
        col_num += 1

    ## Fill data
    for obj in data:
        row_num += 1
        col_num = 0
        for field in fields:
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            value = getattr(obj, field, "")
            if isinstance(value, numbers.Number):
                c.value = value
            else:
                c.value = str(value)
            max_width[col_num] = max(max_width[col_num], len(str(c.value)))
            col_num += 1

    col_num = 0
    for _field in fields:
        if col_num < len(string.ascii_uppercase):
            width = int(0.95 * max_width[col_num]) + 4
            ws.column_dimensions[string.ascii_uppercase[col_num]].width = min(width, 60)
        col_num += 1

    with NamedTemporaryFile() as temporary_file:
        wb.save(temporary_file.name)
        temporary_file.seek(0)
        response = HttpResponse(
            temporary_file.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f"attachment; filename={settings.GENO_ID}_Export_{filename_suffix}.xlsx"
        )
        return response


class ExportXlsMixin:
    def export_as_xls(self, request, queryset):
        fields = []
        header = {}
        meta = self.model._meta
        for field in meta.fields:
            fields.append(field.name)
            header[field.name] = field.verbose_name
        return export_to_xls_generic(queryset, fields, str(meta), header, str(meta))

    export_as_xls.short_description = "Ausgewählte als XLS exportieren"


def export_adit_file():
    data = []
    data_fields = [
        "No",
        "Name",
        "Vorname",
        "Wohnungsnr",
        "Seriennr1",
        "Seriennr2",
        "Bemerkungen",
        "Fehler",
    ]

    names = {}
    no_counter = 0
    log = []
    for ru in RentalUnit.objects.filter(active=True):
        if ru.label:
            ru_name = "%s %s" % (ru.label, ru.name)
        else:
            ru_name = "%s %s" % (ru.rental_type, ru.name)
        if ru.adit_serial:
            ## Get names
            contracts = get_active_contracts().filter(rental_units__id__exact=ru.id)
            n_contracts = contracts.count()
            name_found = False
            if n_contracts > 0:
                if n_contracts > 1:
                    raise RuntimeError("Rental unit %s has %d contracts!" % (ru, n_contracts))
                contract = contracts.first()
                for adr in contract.contractors.filter(ignore_in_lists=False):
                    if adr.organization:
                        str_name_full = adr.organization
                        str_name = adr.organization
                        str_firstname = ""
                    else:
                        str_name_full = "%s %s" % (adr.name, adr.first_name)
                        str_name = adr.name
                        str_firstname = adr.first_name
                    if str_name_full in names:
                        log.append(
                            "Ignoring duplicate name: %s (%s/%s)"
                            % (str_name_full, ru_name, ru.adit_serial)
                        )
                    else:
                        no_counter += 1
                        obj = lambda: None
                        obj._fields = data_fields
                        obj.No = no_counter
                        obj.Name = str_name
                        obj.Vorname = str_firstname
                        obj.Wohnungsnr = ru.name
                        obj.Seriennr1 = int(ru.adit_serial)
                        obj.Bemerkungen = ru_name
                        obj.Fehler = ""
                        if len(obj.Name) + len(obj.Vorname) > 31:
                            obj.Fehler = "%s %s" % (obj.Fehler, "Name+Vorname MAX. 32 Zeichen!")
                        if len(obj.Wohnungsnr) > 8:
                            obj.Fehler = "%s %s" % (obj.Fehler, "Wohnungsnr. MAX. 8 Zeichen!")
                        data.append(obj)
                        names[str_name_full] = ru_name
                        name_found = True
                for child in contract.children.filter(name__ignore_in_lists=False):
                    str_name_full = "%s %s" % (child.name.name, child.name.first_name)
                    if str_name_full in names:
                        log.append(
                            "Ignoring duplicate name: %s (%s/%s)"
                            % (str_name_full, ru_name, ru.adit_serial)
                        )
                    else:
                        no_counter += 1
                        obj = lambda: None
                        obj._fields = data_fields
                        obj.No = no_counter
                        obj.Name = child.name.name
                        obj.Vorname = child.name.first_name
                        obj.Wohnungsnr = ru.name
                        obj.Seriennr1 = int(ru.adit_serial)
                        obj.Bemerkungen = ru_name
                        obj.Fehler = ""
                        if len(obj.Name) + len(obj.Vorname) > 31:
                            obj.Fehler = "%s %s" % (obj.Fehler, "Name+Vorname MAX. 32 Zeichen!")
                        if len(obj.Wohnungsnr) > 8:
                            obj.Fehler = "%s %s" % (obj.Fehler, "Wohnungsnr. MAX. 8 Zeichen!")
                        data.append(obj)
                        names[str_name_full] = ru_name
                        name_found = True
            if not name_found:
                ## No name --> add RentalUnit name as placeholder
                if ru_name in names:
                    log.append(
                        "Ignoring duplicate name: %s (%s/%s)" % (ru_name, ru_name, ru.adit_serial)
                    )
                else:
                    no_counter += 1
                    obj = lambda: None
                    obj._fields = data_fields
                    obj.No = no_counter
                    obj.Name = ru_name
                    obj.Vorname = ""
                    obj.Wohnungsnr = ru.name
                    obj.Seriennr1 = int(ru.adit_serial)
                    obj.Bemerkungen = ru_name
                    obj.Fehler = ""
                    if len(obj.Name) + len(obj.Vorname) > 31:
                        obj.Fehler = "%s %s" % (obj.Fehler, "Name+Vorname MAX. 32 Zeichen!")
                    if len(obj.Wohnungsnr) > 8:
                        obj.Fehler = "%s %s" % (obj.Fehler, "Wohnungsnr. MAX. 8 Zeichen!")
                    data.append(obj)
                    names[ru_name] = ru_name
        else:
            log.append("No adit_serial for %s" % ru)
            # print("No adit_serial for %s" % ru)

    ## Append log messages
    for msg in log:
        obj = lambda: None
        obj._fields = data_fields
        obj.No = "#"
        obj.Name = msg
        data.append(obj)

    return export_data_to_xls(
        data,
        title="Sonnerie ADIT Belegungsplan",
        header={},
        filename_suffix="Sonnerie_ADIT_Belegungsplan",
    )


def export_adit_file_xlsm():
    filename = "ADIT Belegungsplan"
    template_filename = "/tmp/%s.xlsm" % filename
    wb = load_workbook(template_filename)
    if "ADIT1000" not in wb.sheetnames:
        # print(wb.sheetnames)
        # return 1
        raise RuntimeError("Sheet ADIT1000 not found!")

    ws = wb["ADIT1000"]

    row_count = 12
    col_count = 3

    col_name = 2
    col_serial1 = 6
    col_notes = 10

    row_list_start = 17
    row_list_max = 1014

    row = row_list_start
    names = {}
    for ru in RentalUnit.objects.filter(active=True):
        if ru.label:
            ru_name = "%s %s" % (ru.label, ru.name)
        else:
            ru_name = "%s %s" % (ru.rental_type, ru.name)
        if ru.adit_serial:
            ## Get names
            contracts = get_active_contracts().filter(rental_units__id__exact=ru.id)
            n_contracts = contracts.count()
            name_found = False
            if n_contracts > 0:
                if n_contracts > 1:
                    raise RuntimeError("Rental unit %s has %d contracts!" % (ru, n_contracts))
                contract = contracts.first()
                for adr in contract.contractors.filter(ignore_in_lists=False):
                    if adr.organization:
                        str_name = adr.organization
                    else:
                        str_name = "%s %s" % (adr.name, adr.first_name)
                    if str_name in names:
                        print(
                            "Ignoring duplicate name: %s (%s/%s)"
                            % (str_name, ru_name, ru.adit_serial)
                        )
                    else:
                        ws.cell(row=row, column=col_name).value = str_name
                        ws.cell(row=row, column=col_serial1).value = int(ru.adit_serial)
                        ws.cell(row=row, column=col_notes).value = ru_name
                        row += 1
                        names[str_name] = ru_name
                        name_found = True
                for child in contract.children.filter(name__ignore_in_lists=False):
                    str_name = "%s %s" % (child.name.name, child.name.first_name)
                    if str_name in names:
                        print(
                            "Ignoring duplicate name: %s (%s/%s)"
                            % (str_name, ru_name, ru.adit_serial)
                        )
                    else:
                        ws.cell(row=row, column=col_name).value = str_name
                        ws.cell(row=row, column=col_serial1).value = int(ru.adit_serial)
                        ws.cell(row=row, column=col_notes).value = ru_name
                        row += 1
                        names[str_name] = ru_name
                        name_found = True
            if not name_found:
                ## No name --> add RentalUnit name as placeholder
                if ru_name in names:
                    print(
                        "Ignoring duplicate name: %s (%s/%s)" % (ru_name, ru_name, ru.adit_serial)
                    )
                else:
                    ws.cell(row=row, column=col_name).value = ru_name
                    ws.cell(row=row, column=col_serial1).value = int(ru.adit_serial)
                    ws.cell(row=row, column=col_notes).value = ru_name
                    row += 1
                    names[ru_name] = ru_name
        else:
            print("No adit_serial for %s" % ru)

    ws.cell(row=row_count, column=col_count).value = len(names)

    ## Delete rest
    stop = 0
    while row <= row_list_max and stop == 0:
        if ws.cell(row=row, column=col_serial1).value:
            ws.cell(row=row, column=col_name).value = "Leer %s" % row
            ws.cell(row=row, column=col_serial1).value = ""
            ws.cell(row=row, column=col_notes).value = ""
            row += 1
        else:
            stop = 1

    with NamedTemporaryFile() as temporary_file:
        wb.save(temporary_file.name)
        temporary_file.seek(0)
        response = HttpResponse(
            temporary_file.read(), content_type="application/vnd.ms-excel.sheet.macroEnabled.12"
        )
        response["Content-Disposition"] = 'attachment; filename="{}_generated_{}.xlsm"'.format(
            filename, datetime.datetime.now().strftime("%Y%m%d%H%M")
        )
        return response


def generate_demo_camt053_file():
    template = loader.get_template("geno/camt053_demo_data.xml")
    context = {"payments": []}
    date = datetime.date.today().strftime("%Y-%m-%d")
    ts = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    comment = "DEMO Einzahlung"
    payments = generate_demo_payments()
    include_next = True
    for refnr, info in payments.items():
        info["refnr"] = refnr.replace(" ", "")
        info["transaction_id"] = f"DEMO_{info['refnr']}_{ts}"
        info["date"] = date
        info["comment"] = comment
        info["amount"] = format(info["amount"], ".2f")
        if include_next:
            context["payments"].append(info)
            include_next = False
        else:
            include_next = True
    return template.render(context)


def generate_demo_payments():
    payments = {}
    for invoice in Invoice.objects.filter(consolidated=False, active=True, invoice_type="Invoice"):
        if invoice.invoice_category.linked_object_type == "Contract":
            refnr = get_reference_nr(invoice.invoice_category, invoice.contract.id)
            adr = invoice.contract.contractors.first()
        else:
            refnr = get_reference_nr(invoice.invoice_category, invoice.person.id, invoice.id)
            adr = invoice.person
        if refnr:
            if refnr not in payments:
                payments[refnr] = {"amount": 0, "debtor_name": str(adr)}
            payments[refnr]["amount"] += invoice.amount
    return payments
