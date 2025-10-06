import builtins
import contextlib
import csv
import datetime
import io
import re
import subprocess
import zipfile

from django.conf import settings
from django.db import IntegrityError, transaction
from django.db.models import Q
from django.http import HttpResponse
from openpyxl import load_workbook

## Installed from custom (modified) python-sepa subdirectory
from sepa import parser as sepa_parser

from .models import (
    Address,
    Child,
    Contract,
    Member,
    MemberAttribute,
    MemberAttributeType,
    RentalUnit,
    Share,
    ShareType,
    get_active_shares,
)
from .sepa_reader import SepaReaderException, read_camt
from .utils import decode_from_iso8859


def process_eigenmittel():
    import openpyxl
    from openpyxl.styles import Font

    # if not request.user.has_perm('geno.canview_share'):
    #    return unauthorized(request)

    response = []

    ## Import Eigenmittelliste -> members
    header = []
    members = {}
    count = 0
    with open("/tmp/Eigenmittel.csv", "rb") as csvfile:
        memberreader = csv.reader(csvfile, delimiter=b",", quotechar=b'"')
        for row in memberreader:
            fields = []
            if count < 3:
                header.append(row)
            else:
                ## Sanitze input
                for i in [0, 4, 7, 10, 13, 16, 20, 24, 28, 32, 34]:
                    row[i] = row[i].replace(" CHF", "")
                    row[i] = row[i].replace("'", "")
                    if row[i] == "":
                        row[i] = 0
                ## Assemble data
                record = {
                    "id": int(row[0]),
                    "as": int(row[4]) + int(row[7]),  ## Anteilscheine, vereinbart
                    "dz": int(row[10]) + int(row[13]),  ## Darlehen zinslos, vereinbart
                    "dv": int(row[16]) + int(row[20]),  ## Darlehen verzinst, vereinbart
                    "dk": int(row[24]) + int(row[28]),  ## Darlehenskasse, vereinabrt
                    "le": int(row[32]),  ## Legat, vereinbart
                    "total": int(row[34]),  ## Total vereinbart
                    "status": row[35],
                    "name": "NOT IN DB",
                    "as_eff": 0,
                    "dz_eff": 0,
                    "dv_eff": 0,
                    "dk_eff": 0,
                    "total_eff": 0,
                    "status_eff": "",
                }
                fields.append("Total %d" % (record["total"]))

                if record["id"] == 0:
                    if record["total"] > 0:
                        response.append(
                            {
                                "info": "WARN: Ignore member without ID but total %s"
                                % record["id"],
                                "objects": fields,
                            }
                        )
                else:
                    members[record["id"]] = record
            count += 1

    ## Add current shares to 'members'
    # today = datetime.datetime.today()
    share_type_map = {
        "Anteilschein": "as_eff",
        "Darlehen zinslos": "dz_eff",
        "Darlehen verzinst": "dv_eff",
        "Depositenkasse": "dk_eff",
    }
    for m in Address.objects.filter(active=True):
        if m.pk not in members:
            members[m.pk] = {
                "id": m.pk,
                "as": 0,
                "dz": 0,
                "dv": 0,
                "dk": 0,
                "le": 0,
                "total": 0,
                "status": "ONLY IN DB",
            }
        total = 0
        for share_type in ShareType.objects.all():
            summe = 0
            for s in get_active_shares().filter(name=m).filter(share_type=share_type):
                summe += s.quantity * float(s.value)
            members[m.pk][share_type_map[share_type.name]] = int(summe)
            total += summe
        members[m.pk]["total_eff"] = int(total)
        members[m.pk]["name"] = "%s %s, %s" % (m.name, m.first_name, m.city)
        try:
            member = Member.objects.get(name=s.name)
            if member.date_leave:
                members[m.pk]["status_eff"] = (
                    "Ausgetretenes Mitglied (per %s)" % member.date_leave.strftime("%d.%m.%Y")
                )
            else:
                members[m.pk]["status_eff"] = "Mitglied"
        except Member.DoesNotExist:
            members[m.pk]["status_eff"] = "Nichtmitglied"

    ## Export joined list
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=Eigenmittel.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Eigenmittel {settings.GENO_NAME}"

    row_num = 0

    ## Header
    width_amount = 10
    columns = [
        ("A", "ID", 7),
        ("B", "Name", 30),
        ("C", "ASver", width_amount),
        ("D", "ASeff", width_amount),
        ("E", "ASdiff", width_amount),
        ("F", "ASjoin", width_amount),
        ("G", "DZver", width_amount),
        ("H", "DZeff", width_amount),
        ("I", "DZdiff", width_amount),
        ("J", "DZjoin", width_amount),
        ("K", "DVver", width_amount),
        ("L", "DVeff", width_amount),
        ("M", "DVdiff", width_amount),
        ("N", "DVjoin", width_amount),
        ("O", "DKver", width_amount),
        ("P", "DKeff", width_amount),
        ("Q", "DKdiff", width_amount),
        ("R", "DKjoin", width_amount),
        ("S", "TOTver", width_amount),
        ("T", "TOTeff", width_amount),
        ("U", "TOTdiff", width_amount),
        ("V", "TOTjoin", width_amount),
        ("W", "Status", 200),
        ("X", "StatusDB", 200),
    ]
    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][1]
        c.font = Font(bold=True)
        # set column width
        ws.column_dimensions[columns[col_num][0]].width = columns[col_num][2]

    acc200 = 0
    for m in sorted(members):
        as_diff = members[m]["as_eff"] - members[m]["as"]
        dz_diff = members[m]["dz_eff"] - members[m]["dz"]
        dv_diff = members[m]["dv_eff"] - members[m]["dv"]
        dk_diff = members[m]["dk_eff"] - members[m]["dk"]
        total_diff = members[m]["total_eff"] - members[m]["total"]
        if members[m]["status"].find(b"Vereinbarung"):
            as_pot = max(members[m]["as"], members[m]["as_eff"])
            dz_pot = max(members[m]["dz"], members[m]["dz_eff"])
            dv_pot = max(members[m]["dv"], members[m]["dv_eff"])
            dk_pot = max(members[m]["dk"], members[m]["dk_eff"])
            total_pot = as_pot + dz_pot + dv_pot + dk_pot
        else:
            as_pot = members[m]["as_eff"]
            dz_pot = members[m]["dz_eff"]
            dv_pot = members[m]["dv_eff"]
            dk_pot = members[m]["dk_eff"]
            total_pot = members[m]["total_eff"]
        row = [
            members[m]["id"],
            members[m]["name"],
            members[m]["as"],
            members[m]["as_eff"],
            as_diff,
            as_pot,
            members[m]["dz"],
            members[m]["dz_eff"],
            dz_diff,
            dz_pot,
            members[m]["dv"],
            members[m]["dv_eff"],
            dv_diff,
            dv_pot,
            members[m]["dk"],
            members[m]["dk_eff"],
            dk_diff,
            dk_pot,
            members[m]["total"],
            members[m]["total_eff"],
            total_diff,
            total_pot,
            members[m]["status"],
            members[m]["status_eff"],
        ]
        if total_pot == -200:
            acc200 += 1
        elif total_pot != 0:
            row_num += 1
            for col_num in range(len(row)):
                c = ws.cell(row=row_num + 1, column=col_num + 1)
                c.value = row[col_num]

    row = [
        "",
        "%d Personen" % acc200,
        "",
        "",
        "",
        200 * acc200,
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        200 * acc200,
        "Akkumuliert",
    ]
    row_num += 1
    for col_num in range(len(row)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = row[col_num]

    ## TODO: Add legend
    # Legende:
    # ver =   Vereinbarte finanzielle Unterstützung
    # eff =   Effektiv eingezahlte Beträge
    # diff =  Differenz eingezahlt – vereinbart
    # join =  Maximum(vereinbart, eingezahlt)
    #
    # AS =    Anteilscheine
    # DZ =    Darlehen zinslos
    # DV =    Darlehen verzinst
    # DK =    Depositenkasse
    # TOT =   Total

    wb.save(response)
    return response


def import_members_from_file(empty_tables_first=False):
    response = []

    if empty_tables_first:
        Share.objects.all().delete()
        MemberAttribute.objects.all().delete()
        Member.objects.all().delete()
        Address.objects.all().delete()
        response.append(
            {"info": "Deleting all Address, Member, MemberAttribute, and Share objects!"}
        )

    count = 0
    filename = "Mitglieder.csv"
    with open("/tmp/%s" % filename, encoding="utf-8") as csvfile:
        memberreader = csv.reader(csvfile, delimiter=",", quotechar='"')
        for row in memberreader:
            if count == 0:
                header = row
            else:
                new_addr = Address()
                new_addr.active = True
                new_member = Member()
                fields = []
                att = {}
                att_as = {}
                plz = ""
                ## Set einzelmitglieder flag
                new_member.flag_02 = True
                new_member.notes = "Import von %s" % filename
                for i, val in enumerate(row):
                    val = str(val.strip())  # , 'utf-8')
                    if header[i] == "Name":
                        if val.find(",") == -1:
                            vorname, name = val.split(" ")
                        else:
                            name, vorname = val.split(",")
                        new_addr.name = name.strip()
                        new_addr.first_name = vorname.strip()
                    elif header[i] == "Vorname:":
                        new_addr.first_name = val
                    elif header[i] == "Name:":
                        new_addr.name = val
                    elif header[i] == "Organisation":
                        new_addr.organization = val
                    elif header[i] == "Strasse / Nr":
                        new_addr.street_name = val
                    elif header[i] == "Postleitzahl":
                        new_addr.city_zipcode = val
                    elif header[i] == "Ort":
                        new_addr.city = val
                    elif header[i] == "Anrede":
                        new_addr.title = val
                    elif header[i] == "Email" or header[i] == "E-Mail":
                        new_addr.email = val
                    elif header[i] == "Telefon":
                        new_addr.telephone = val
                    elif header[i] == "Mobile":
                        new_addr.mobile = val
                    elif header[i] == "Geburtsdatum":
                        if len(val) == 10:
                            new_addr.date_birth = datetime.datetime.strptime(
                                val, "%d.%m.%Y"
                            ).date()
                        elif len(val) == 8:
                            new_addr.date_birth = datetime.datetime.strptime(
                                val, "%d.%m.%y"
                            ).date()
                    elif header[i] == "Heimatort":
                        new_addr.hometown = val
                    elif header[i] == "Beruf/Ausbildung":
                        new_addr.occupation = val
                    elif header[i] == "Eintritt" or header[i] == "Datum des Eintrags":
                        if len(val) >= 10:
                            new_member.date_join = datetime.datetime.strptime(
                                val[0:10], "%d.%m.%Y"
                            ).date()
                        elif len(val) == 8:
                            new_member.date_join = datetime.datetime.strptime(
                                val, "%d.%m.%y"
                            ).date()
                        else:
                            raise RuntimeError("Can't parse date Eintritt")
                            # new_member.date_join = datetime.date.today()
                    elif header[i] == "Austritt":
                        if len(val) >= 10:
                            new_member.date_leave = datetime.datetime.strptime(
                                val[0:10], "%d.%m.%Y"
                            ).date()
                        elif len(val) == 8:
                            new_member.date_leave = datetime.datetime.strptime(
                                val, "%d.%m.%y"
                            ).date()
                        else:
                            raise RuntimeError("Can't parse date Eintritt")
                            # new_member.date_join = datetime.date.today()
                    elif header[i] == "Wohnen":
                        if val == "x":
                            new_member.flag_01 = True
                        elif val != "":
                            fields.append(
                                "WARNING: Ignoring unknown value for flag_wohnen: %s" % val
                            )
                    elif header[i] == "Gewerbe/ Arbeiten":
                        if val == "x":
                            new_member.flag_02 = True
                        elif val != "":
                            fields.append(
                                "WARNING: Ignoring unknown value for flag_gewerbe: %s" % val
                            )
                    elif header[i] == "Idee Umsetzen":
                        if val == "x":
                            new_member.flag_03 = True
                        elif val != "":
                            fields.append(
                                "WARNING: Ignoring unknown value for flag_ideen: %s" % val
                            )
                    elif (
                        header[i] == "Projekt unterstützen"
                        or header[i] == "Projekt unterst\xc3\xbczen"
                    ):
                        if val == "x":
                            new_member.flag_04 = True
                        elif val != "":
                            fields.append(
                                "WARNING: Ignoring unknown value for flag_support: %s" % val
                            )
                    elif header[i] == "Dranbleiben":
                        if val == "x":
                            new_member.flag_05 = True
                        elif val != "":
                            fields.append(
                                "WARNING: Ignoring unknown value for flag_dranbleiben: %s" % val
                            )
                    elif header[i] == "Bemerkungen":
                        new_member.notes = new_member.notes + val + "\n\n"
                    elif header[i] == "Bezahlt":
                        new_member.notes = new_member.notes + "[Bezahlt: %s]\n\n" % val
                    elif header[i] == "Jahresbeitrag 2022 bez.":
                        if val != "":
                            att["Mitgliederbeitrag 2022"] = {
                                "value": "Bezahlt",
                                "date": datetime.datetime.strptime(val, "%d.%m.%Y").date(),
                            }
                        # else:
                        #    att['Mitgliederbeitrag 2022'] = {'value': 'Gefordert', 'date': datetime.date.today()}
                    elif header[i].startswith("Ich beanspruche den reduzierten Jahresbeitrag"):
                        if val != "":
                            new_member.flag_03 = True
                    elif header[i] == "AS bezahlt":
                        if "AS1" not in att_as:
                            att_as["AS1"] = {"value": 1, "date": None}
                        if val != "":
                            att_as["AS1"]["date"] = datetime.datetime.strptime(
                                val, "%d.%m.%Y"
                            ).date()
                    elif header[i] == "zAS bezahlt":
                        if val != "":
                            if "AS2" not in att_as:
                                att_as["AS2"] = {"value": 1, "date": None}
                            att_as["AS2"]["date"] = datetime.datetime.strptime(
                                val, "%d.%m.%Y"
                            ).date()
                    elif header[i] == "Zusätzl. AS" or header[i] == "Zusätzliche Anteilscheine":
                        if val != "" and int(val) >= 1:
                            if "AS2" not in att_as:
                                att_as["AS2"] = {"value": 1, "date": None}
                            att_as["AS2"]["value"] = val
                    # elif header[i] == "AS2 Bez.":
                    #    if val != '':
                    #        att_as['AS2'] = [datetime.datetime.strptime(val, "%d.%m.%Y").date(),0]
                    # elif header[i] == "AS2 Anz.":
                    #    if val != '':
                    #        att_as['AS2'][1] = val
                    elif header[i] == "Zusicherung":
                        if val != "":
                            att_as["ZS"] = float(val)
                    else:
                        fields.append(
                            'WARNING: Ignoring field "%s" with value "%s".' % (header[i], val)
                        )

                ## Add plz
                if plz != "":
                    new_addr.city = "%s %s" % (plz, new_addr.city)

                try:
                    new_addr.save()
                except IntegrityError:
                    response.append(
                        {
                            "info": "WARNING: Skippting member %d: Name %s %s exists already!"
                            % (count, new_addr.first_name, new_addr.name)
                        }
                    )
                    continue
                fields.append(
                    "Added new address %s, %s %s, %s, %s"
                    % (
                        new_addr.organization,
                        new_addr.first_name,
                        new_addr.name,
                        new_addr.street,
                        new_addr.city,
                    )
                )
                new_member.name = new_addr
                new_member.save()
                fields.append("Added new membership %s" % (new_member.date_join))

                ## Add member attributes
                for a in att:
                    ma = MemberAttribute(
                        member=new_member,
                        attribute_type=MemberAttributeType.objects.get(name=a),
                        value=att[a]["value"],
                        date=att[a]["date"],
                    )
                    ma.save()
                    fields.append(
                        "Added new attribute: %s = %s, %s" % (a, att[a]["value"], att[a]["date"])
                    )
                for ass in att_as:
                    if ass == "ZS":
                        if att_as[ass] > 0:
                            sh = Share(
                                name=new_addr,
                                share_type=ShareType.objects.get(name="Entwicklungsbeitrag"),
                                date=datetime.datetime.today(),
                                quantity=1,
                                value=att_as[ass],
                                state="gefordert",
                            )
                            sh.save()
                            fields.append(
                                "Added new share: Entwicklungsbeitrag %s (gefordert)"
                                % (att_as[ass])
                            )
                    else:
                        if ass == "AS2":
                            share_type = "Anteilschein freiwillig"
                        else:
                            share_type = "Anteilschein Einzelmitglied"
                        if att_as[ass]["date"]:
                            pay_state = "bezahlt"
                            # else:
                            #    pay_state = 'gefordert'
                            #    att_as[ass]['date'] = datetime.date.today()
                            sh = Share(
                                name=new_addr,
                                share_type=ShareType.objects.get(name=share_type),
                                date=att_as[ass]["date"],
                                quantity=att_as[ass]["value"],
                                value=200,
                                state=pay_state,
                            )
                            sh.save()
                            fields.append(
                                "Added new share: %s %s (bezahlt %s)"
                                % (att_as[ass]["value"], share_type, att_as[ass]["date"])
                            )

                # fields = ( '%s: %s' % (header[i],val) for i,val in enumerate(row) )
                response.append({"info": "Importing member %d" % count, "objects": fields})
            count += 1
    return response


def update_address_from_file():
    ## Update/Check field from Eigenmittel spreadsheet
    anrede = {"M": "Herr", "F": "Frau", "D": "Divers", "P": "Paar", "O": "Org"}
    count = 0
    response = []
    with open("/tmp/Eigenmittelliste.csv", "rb") as csvfile:
        reader = csv.reader(csvfile, delimiter=",", quotechar='"')
        m = {}
        for row in reader:
            if count == 0:
                header = row
                for i, val in enumerate(header):
                    m[val] = i
            else:
                if row[m["Name"]].strip() == "" and row[m["Vorname"]].strip() == "":
                    continue
                du = row[m["Anrede"]][0]
                mf = row[m["Anrede"]][1]
                msg = []
                try:
                    adr = Address.objects.get(
                        name=row[m["Name"]].strip(), first_name=row[m["Vorname"]].strip()
                    )
                    if (
                        adr.street_name != row[m["Strasse"]].strip()
                        and adr.street_name != row[m["Strasse"]].strip().replace("str.", "strasse")
                        and adr.street_name.replace("str.", "strasse") != row[m["Strasse"]].strip()
                    ):
                        msg.append(
                            'DIFF: Strasse "%s" - IMPORT: "%s"'
                            % (adr.street_name.replace("str.", "strasse"), row[m["Strasse"]])
                        )
                    if adr.city_name != row[m["Wohnort"]]:
                        msg.append(
                            'DIFF: Wohnort "%s" - IMPORT: "%s"'
                            % (adr.city_name, row[m["Wohnort"]])
                        )
                    if adr.email != row[m["Email"]].strip():
                        msg.append(
                            'DIFF: Email "%s" - IMPORT: "%s"' % (adr.email, row[m["Email"]])
                        )
                    if adr.telephone.replace(" ", "") != row[m["Telefon"]].replace(
                        " ", ""
                    ) and adr.mobile.replace(" ", "") != row[m["Telefon"]].replace(" ", ""):
                        msg.append(
                            'DIFF: Telefon "%s/%s" - IMPORT: "%s"'
                            % (adr.telephone, adr.mobile, row[m["Telefon"]])
                        )
                    if adr.title != anrede[mf]:
                        msg.append('DIFF: Anrede "%s" - IMPORT: "%s"' % (adr.title, mf))

                    ## Update formal
                    if du == "S":
                        msg.append("UPDATE Du->Sie")
                        adr.formal = "Sie"
                        adr.save()
                    elif du != "D":
                        raise RuntimeError("Unknown Anrede (du): %s" % du)
                except Address.DoesNotExist:
                    msg.append("Address not found -> IMPORTING NEW!")
                    new_addr = Address()
                    new_addr.active = True
                    new_addr.name = row[m["Name"]].strip()
                    new_addr.first_name = row[m["Vorname"]].strip()
                    new_addr.title = anrede[mf]
                    if du == "S":
                        new_addr.formal = "Sie"
                    new_addr.street_name = row[m["Strasse"]].strip()
                    new_addr.city_name = row[m["Wohnort"]].strip()
                    new_addr.telephone = row[m["Telefon"]].strip()
                    new_addr.email = row[m["Email"]].strip()
                    if row[m["Geburtsdatum"]].strip() != "":
                        new_addr.date_birth = datetime.datetime.strptime(
                            row[m["Geburtsdatum"]].strip(), "%d.%m.%Y"
                        ).date()
                    new_addr.hometown = row[m["Heimatort"]].strip()
                    new_addr.occupation = row[m["Beruf/Ausbildung"]].strip()
                    new_addr.save()
                if msg:
                    response.append(
                        {
                            "info": "Person: %s, %s" % (row[m["Name"]], row[m["Vorname"]]),
                            "objects": msg,
                        }
                    )
            count += 1
    return response


def import_contracts_from_file(empty_tables_first=False):
    response = []

    if empty_tables_first:
        Contract.objects.all().delete()
        Address.objects.all().delete()
        response.append({"info": "Deleting all Address and Contract objects!"})

    count = 0
    header = []
    with open("/tmp/Raumliste_Adressen_Vertraege.csv", "rb") as csvfile:
        reader = csv.reader(csvfile, delimiter=b",", quotechar=b'"')
        for row in reader:
            if count == 0:
                header = row
            else:
                new_addr = Address()
                new_addr.active = True
                new_contract = Contract()
                fields = []
                for i, val in enumerate(row):
                    val = str(val.strip(), "utf-8")
                    if header[i] == "Name":
                        # print val
                        if val.find(",") != -1:
                            name, vorname = val.split(",", 1)
                        elif val.find(" ") != -1:
                            vorname, name = val.split(" ", 1)
                        else:
                            name = val
                            vorname = ""
                        new_addr.name = name.strip()
                        new_addr.first_name = vorname.strip()
                    elif header[i] == "Email":
                        new_addr.email = val
                    elif header[i] == "Stock":
                        new_contract.floor = val
                    elif header[i] == "Nutzflaeche (m2)":
                        if val:
                            # print("Area: %f" % float(val))
                            new_contract.area = float(val)
                    elif header[i] == "NK akonto":
                        new_contract.nk = float(val)
                    elif header[i] == "Miete total":
                        new_contract.rent_total = float(val)
                    elif header[i] == "Gebaeudeteil":
                        new_contract.building = val
                    elif header[i] == "Raumbez.":
                        new_contract.name = val
                    elif header[i] == "Mietbeginn":
                        new_contract.date = datetime.datetime.strptime(val, "%d.%m.%y").date()
                    elif header[i] == "Depot":
                        new_contract.depot = float(val)
                    # else:
                    #    fields.append('WARNING: Ignoring field "%s" with value "%s".' % (header[i],val))

                try:
                    new_addr.save()
                except IntegrityError:
                    new_addr = Address.objects.get(
                        name=new_addr.name, first_name=new_addr.first_name
                    )
                    # response.append({"info": "WARNING: Skippting member %d: Name %s %s exists already!" % (count,new_addr.first_name,new_addr.name)})
                    response.append(
                        {
                            "info": "INFO: Adding multiple contract %d to: Name %s %s."
                            % (count, new_addr.first_name, new_addr.name)
                        }
                    )
                    # continue
                fields.append(
                    "Added new address %s, %s %s, %s, %s"
                    % (
                        new_addr.organization,
                        new_addr.first_name,
                        new_addr.name,
                        new_addr.street,
                        new_addr.city,
                    )
                )
                new_contract.person = new_addr
                new_contract.save()
                fields.append("Added new contract %s" % (new_contract))

                # fields = ( '%s: %s' % (header[i],val) for i,val in enumerate(row) )
                response.append({"info": "Importing contract %d" % count, "objects": fields})
            count += 1
    return response


def import_codes_from_file(empty_tables_first=False):
    response = []

    try:
        atype = MemberAttributeType.objects.get(name="Ausschreibung Code")
    except MemberAttributeType.DoesNotExist:
        response.append({"info": "ERROR: Mitglieder Attribut Typ nicht gefunden!"})
        return response

    if empty_tables_first:
        MemberAttribute.objects.filter(attribute_type=atype).delete()
        response.append({"info": "Deleting all codes from MemberAttribute!"})

    ## Load existing codes
    existing_codes = {}
    members_with_codes = []
    for c in MemberAttribute.objects.filter(attribute_type=atype):
        existing_codes[c.value] = c.member
        members_with_codes.append(c.member)

    free_codes = []
    found_codes = []
    try:
        with open("/tmp/Zugangscodes.csv", "rb") as csvfile:
            reader = csv.reader(csvfile, delimiter=b",", quotechar=b'"')
            header = []
            column_code = None
            column_status = None
            count = 0
            for row in reader:
                for i, val in enumerate(row):
                    row[i] = str(val.strip(), "utf-8")
                if count == 0:
                    header = row
                    for i, val in enumerate(header):
                        if val == "Zugangscode":
                            column_code = i
                        elif val == "Status":
                            column_status = i
                else:
                    if row[column_status] == "automatisch vergeben" and row[
                        column_code
                    ].startswith("WBM"):
                        if row[column_code] in existing_codes:
                            found_codes.append(row[column_code])
                        else:
                            free_codes.append(row[column_code])
                count += 1
    except OSError as e:
        response.append({"info": "ERROR: Konnte CSV Datei nicht lesen: %s" % e})
        return response

    today = datetime.datetime.today()
    missing_codes = []
    new_code_count = 0
    with transaction.atomic():
        for m in Member.objects.filter(Q(date_leave=None) | Q(date_leave__gt=today)):
            if m not in members_with_codes:
                if free_codes:
                    code = free_codes.pop()
                    new = MemberAttribute(member=m, attribute_type=atype, date=today, value=code)
                    new.save()
                    response.append({"info": "Code %s neu vergeben an %s." % (code, m)})
                    new_code_count += 1
                else:
                    missing_codes.append("%s" % m)

    if new_code_count:
        response.append({"info": "Es wurden %d Codes neu vergeben." % new_code_count})

    if missing_codes:
        response.append(
            {
                "info": "WARNUNG: Es fehlen %d Codes für folgende Mitglieder:"
                % len(missing_codes),
                "objects": missing_codes,
            }
        )
    else:
        response.append({"info": "Es hat noch %d freie Codes übrig." % len(free_codes)})

    if found_codes:
        response.append(
            {
                "info": "%d codes waren bereits an Mitglieder vergeben:" % len(found_codes),
                "objects": found_codes,
            }
        )

    return response


def import_rentalunits_from_file(empty_tables_first=False):
    response = []

    if empty_tables_first:
        RentalUnit.objects.all().delete()
        response.append({"info": "Deleting all RentalUnit objects!"})

    count = 0
    header = []
    with open("/tmp/export_objekte.csv", "rb") as csvfile:
        reader = csv.reader(csvfile, delimiter=b",", quotechar=b'"')
        for row in reader:
            if count == 0:
                header = row
            else:
                new_unit = RentalUnit()
                new_unit.building = "Holligerhof 8"
                fields = []
                rent_net = None
                for i, val in enumerate(row):
                    val = str(val.strip(), "utf-8")
                    # "Details","Bezeichnung","Typ","Bruttomiete","Nettomiete","Zimmer","Fläche","Loggiafläche","Balkonfläche","Ausschreibungstext","Stockwerk","Mindestbelegung total","Mietbedingungen","Anteilskapital","Erstellungsdatum"
                    if header[i] == "Bezeichnung":
                        new_unit.name = val
                    elif header[i] == "Typ":
                        new_unit.rental_type = val
                    elif header[i] == "Bruttomiete":
                        new_unit.rent_total = float(val)
                    elif header[i] == "Nettomiete":
                        rent_net = float(val)
                    elif header[i] == "Zimmer":
                        new_unit.rooms = float(val)
                    elif str(header[i], "utf-8") == "Fläche":
                        new_unit.area = float(val)
                    elif str(header[i], "utf-8") == "Loggiafläche":
                        if val == "":
                            new_unit.area_add = None
                        else:
                            new_unit.area_add = float(val)
                    elif str(header[i], "utf-8") == "Balkonfläche":
                        if val == "":
                            new_unit.area_balcony = None
                        else:
                            new_unit.area_balcony = float(val)
                    elif header[i] == "Ausschreibungstext":
                        new_unit.height = val
                    elif header[i] == "Stockwerk":
                        new_unit.floor = val
                    elif header[i] == "Mietbedingungen":
                        new_unit.min_occupancy = float(val)
                    elif header[i] == "Anteilskapital":
                        if val == "":
                            new_unit.share = None
                        else:
                            new_unit.share = float(val)
                    elif header[i] == "Erstellungsdatum":
                        new_unit.comment = "Von eMonitor %s" % (val)
                if rent_net and new_unit.rent_total:
                    new_unit.nk = new_unit.rent_total - rent_net

                new_unit.save()
                fields.append("Added new RentalUnit %s" % (new_unit.name))

                # fields = ( '%s: %s' % (header[i],val) for i,val in enumerate(row) )
                response.append({"info": "Importing unit %d" % count, "objects": fields})
            count += 1
    return response


def import_emonitor_children_from_file(empty_tables_first=False):
    response = []

    # if empty_tables_first:
    #    RentalUnit.objects.all().delete()
    #    response.append({'info': 'Deleting all RentalUnit objects!'})

    count = 0
    header = []
    with open("/tmp/export_kinder.csv", encoding="utf8") as csvfile:
        reader = csv.reader(csvfile, delimiter=",", quotechar='"')
        for row in reader:
            if count == 0:
                header = row
            else:
                new_adr = Address()
                new_child = Child()
                fields = []
                contract_id = None
                for i, val in enumerate(row):
                    val = val.strip()
                    # Bewerbung,ID,Vorname,Nachname,Anwesenheit Kinder,Eltern(teil) des Kindes,Geburtsdatum,Erstellungsdatum
                    # if str(header[i], 'utf-8') == "Bewerbung":
                    if header[i] == "Bewerbung":
                        contract_id = int(val)
                    elif header[i] == "ID":
                        new_child.import_id = int(val)
                    elif header[i] == "Vorname":
                        new_adr.first_name = val
                    elif header[i] == "Nachname":
                        new_adr.name = val
                    elif header[i] == "Geburtsdatum":
                        if val:
                            # new_adr.date_birth = datetime.datetime.strptime(val, '%Y-%m-%d').date()
                            new_adr.date_birth = datetime.datetime.strptime(val, "%d.%m.%y").date()
                    elif header[i] == "Anwesenheit Kinder":
                        new_child.presence = float(val)
                    elif header[i] == "Eltern(teil) des Kindes":
                        new_child.parents = val
                    elif header[i] == "Erstellungsdatum":
                        new_adr.comment = "Von eMonitor %s" % (val)
                        new_child.comment = "Von eMonitor %s" % (val)

                contract = Contract.objects.get(import_id=contract_id)

                if not new_adr.name and not new_adr.first_name:
                    fields.append("Ignoring empty name!")
                else:
                    new_adr.save()
                    new_child.name = new_adr
                    new_child.save()
                    fields.append("Added Child %s" % new_child)

                    contract.children.add(new_child)
                    contract.save()
                    fields.append("Added Child to contract %s" % (contract))

                # fields = ( '%s: %s' % (header[i],val) for i,val in enumerate(row) )
                response.append({"info": "Importing child %d" % count, "objects": fields})
            count += 1
    return response


def import_emonitor_addresses_from_file(empty_tables_first=False):
    response = []

    # if empty_tables_first:
    #    RentalUnit.objects.all().delete()
    #    response.append({'info': 'Deleting all RentalUnit objects!'})

    p_tel_swiss = re.compile(
        r"^\s*(?:0041 ?|\+?\+?41 ?|0|041 0)?(\d\d) ?(\d\d\d) ?(\d\d) ?(\d\d)[^0-9]*$"
    )

    count = 0
    header = []
    with open("/tmp/export_vertrag.csv", "rb") as csvfile:
        reader = csv.reader(csvfile, delimiter=b",", quotechar=b'"')
        for row in reader:
            if count == 0:
                header = row
            else:
                new_adr = Address()
                fields = []
                plz = None
                contract_id = None
                for i, val in enumerate(row):
                    val = str(val.strip(), "utf-8")
                    # Bewerbungs-ID,Erwachsener ID,Anrede,Vorname,Name,Nationalität,Aufenthaltsstatus,Zivilstand,Geburtsdatum,Beschäftigungsstatus,Beruf,Straße,PLZ,Ort,E-Mail,Telefonnummer,Arbeitgeber Telefon,Wohnung,Zusätzliche Objekte,Vertragsstatus,Anzahl Kinder
                    if str(header[i], "utf-8") == "Bewerbungs-ID":
                        contract_id = int(val)
                    elif str(header[i], "utf-8") == "Erwachsener ID":
                        new_adr.import_id = int(val)
                    elif str(header[i], "utf-8") == "Anrede":
                        new_adr.title = val
                    elif str(header[i], "utf-8") == "Vorname":
                        new_adr.first_name = val
                    elif str(header[i], "utf-8") == "Name":
                        new_adr.name = val
                    elif str(header[i], "utf-8") == "Geburtsdatum":
                        if val:
                            new_adr.date_birth = datetime.datetime.strptime(val, "%Y-%m-%d").date()
                    elif str(header[i], "utf-8") == "Straße":
                        ## TODO: Split Nr?
                        new_adr.street_name = val
                    elif str(header[i], "utf-8") == "PLZ":
                        new_adr.city_zipcode = val
                    elif str(header[i], "utf-8") == "Ort":
                        new_adr.city_name = val
                    elif str(header[i], "utf-8") == "E-Mail":
                        new_adr.email = val
                    elif str(header[i], "utf-8") == "Telefonnummer":
                        if val == "":
                            continue
                        if len(val) > 20:
                            parts = val.split("  ")
                            val = parts[0]
                        m = p_tel_swiss.match(val)
                        if m:
                            new_adr.telephone = "0%s %s %s %s" % (
                                m.group(1),
                                m.group(2),
                                m.group(3),
                                m.group(4),
                            )
                        else:
                            raise Exception('Couldnt parse telephone: "%s"' % val)
                if plz:
                    new_adr.city_name = "%s %s" % (plz, new_adr.city)

                contract = Contract.objects.get(import_id=contract_id)

                if not new_adr.name and not new_adr.first_name:
                    fields.append("Ignoring empty name!")
                else:
                    ## Try to find match
                    save = False
                    existing = Address.objects.filter(
                        name=new_adr.name, first_name=new_adr.first_name, organization=""
                    )
                    if not existing:
                        ## Try email+anrede
                        existing = Address.objects.filter(email=new_adr.email, title=new_adr.title)
                        fields.append(
                            "WARNING: Searching existing Address by email+title %s+%s"
                            % (new_adr.email, new_adr.title)
                        )
                    if not existing:
                        ## Try first name and tel
                        existing = Address.objects.filter(
                            first_name=new_adr.first_name, telephone=new_adr.telephone
                        )
                        fields.append(
                            "WARNING: Searching existing Address by first_name+telephone %s+%s"
                            % (new_adr.first_name, new_adr.telephone)
                        )
                    if not existing:
                        ## Try last name and tel
                        existing = Address.objects.filter(
                            name=new_adr.name, telephone=new_adr.telephone
                        )
                        fields.append(
                            "WARNING: Searching existing Address by name+telephone %s+%s"
                            % (new_adr.name, new_adr.telephone)
                        )
                    if len(existing) > 0:
                        if len(existing) != 1:
                            raise Exception("Too many matches found")
                        fields.append(
                            "Found existing Address %s, %s" % (new_adr.name, new_adr.first_name)
                        )
                        old_adr = existing[0]
                        for a in (
                            "import_id",
                            "title",
                            "date_birth",
                            "street",
                            "city",
                            "email",
                            "telephone",
                        ):
                            att_old = getattr(old_adr, a)
                            att_new = getattr(new_adr, a)
                            if att_new and att_old != att_new:
                                if a == "import_id" and not att_old:
                                    setattr(old_adr, a, att_new)
                                    fields.append("Setting %s: %s -> %s" % (a, att_old, att_new))
                                    save = True
                                elif a == "email" and not old_adr.email2:
                                    old_adr.email2 = att_new
                                    fields.append(
                                        "Adding 2nd email address: %s + %s" % (att_old, att_new)
                                    )
                                    save = True
                                elif a == "email" and att_new == old_adr.email2:
                                    ## Match with email2 is OK
                                    continue
                                elif a == "telephone" and not old_adr.mobile:
                                    old_adr.mobile = att_new
                                    fields.append(
                                        "Adding 2nd telephone: %s + %s" % (att_old, att_new)
                                    )
                                    save = True
                                elif a == "telephone" and att_new == old_adr.mobile:
                                    ## Match with phone2 is OK
                                    continue
                                else:
                                    fields.append(
                                        "NOT MATCHING %s: %s != %s" % (a, att_old, att_new)
                                    )
                        adr = old_adr
                    else:
                        fields.append(
                            "Adding new Address %s, %s" % (new_adr.name, new_adr.first_name)
                        )
                        save = True
                        adr = new_adr
                    if save:
                        adr.save()
                    contract.contractors.add(adr)
                    contract.save()
                    fields.append("Added Address to contract %s" % (contract))

                # fields = ( '%s: %s' % (header[i],val) for i,val in enumerate(row) )
                response.append({"info": "Importing address %d" % count, "objects": fields})
            count += 1
    return response


def import_keller_from_file(empty_tables_first=False):
    response = []

    if empty_tables_first:
        RentalUnit.objects.filter(rental_type="Kellerabteil").delete()
        response.append({"info": "Deleting all RentalUnit objects with type Kellerabteil!"})

    count = 0
    header = []
    with open("/tmp/Mieterkeller_Wohnungen.csv", "rb") as csvfile:
        reader = csv.reader(csvfile, delimiter=b",", quotechar=b'"')
        for row in reader:
            if count == 0:
                header = row
            else:
                new_unit = RentalUnit()
                new_unit.building = "Holligerhof 8"
                new_unit.rent_netto = 0.0
                new_unit.floor = "Keller"
                fields = []
                linked_unit = None
                desc = None
                for i, val in enumerate(row):
                    val = str(val.strip(), "utf-8")
                    # "Raumnr.","Bezeichnung","Zugewiesene Whg.","Bodenfläche (m2)","Raumhöhe (m)"
                    if header[i] == "Raumnr.":
                        new_unit.name = val
                    elif header[i] == "Bezeichnung":
                        desc = val
                        if val.startswith("Mieterkeller"):
                            new_unit.rental_type = "Kellerabteil"
                    elif header[i] == "Zugewiesene Whg.":
                        linked_unit = val
                    elif str(header[i], "utf-8") == "Bodenfläche (m2)":
                        vals = val.split("/")
                        if val == "":
                            new_unit.area = None
                            new_unit.area_add = None
                        elif len(vals) == 1:
                            new_unit.area = float(vals[0])
                            new_unit.area_add = None
                        elif len(vals) == 2:
                            new_unit.area = float(vals[0])
                            new_unit.area_add = float(vals[1])
                        else:
                            raise ValueError("More than two area values found!")
                    elif str(header[i], "utf-8") == "Raumhöhe (m)":
                        new_unit.height = val

                if desc:
                    new_unit.note = desc
                    if linked_unit:
                        new_unit.note = "%s %s" % (new_unit.note, linked_unit)

                if new_unit.rental_type:
                    new_unit.save()
                    fields.append("Added new RentalUnit %s" % (new_unit.name))
                    if linked_unit:
                        try:
                            other_unit = RentalUnit.objects.get(name=linked_unit)
                            contracts = Contract.objects.filter(rental_units__pk=other_unit.pk)
                            if contracts.count() == 1:
                                contract = contracts.first()
                                contract.rental_units.add(new_unit)
                                contract.save()
                                fields.append("Linked to contract %s" % (contract))
                            else:
                                fields.append(
                                    "WARNING: Not linking to contract (None or too many contracts found for %s)"
                                    % (other_unit.name)
                                )
                        except RentalUnit.DoesNotExist:
                            fields.append(
                                "WARNING: Not linking to contract (linked unit %s not found)"
                                % (linked_unit)
                            )

                # fields = ( '%s: %s' % (header[i],val) for i,val in enumerate(row) )
                response.append({"info": "Importing unit %d" % count, "objects": fields})
            count += 1
    return response


def import_emonitor_contracts_from_file(empty_tables_first=False):
    response = []

    if empty_tables_first:
        Contract.objects.all().delete()
        response.append({"info": "Deleting all Contract objects!"})

    count = 0
    header = []
    with open("/tmp/export_zuweisung.csv", "rb") as csvfile:
        reader = csv.reader(csvfile, delimiter=b",", quotechar=b'"')
        for row in reader:
            if count == 0:
                header = row
            else:
                new = Contract()
                new.state = "angeboten"
                new.date = datetime.date(2021, 11, 0o1)
                fields = []
                rentals = []
                for i, val in enumerate(row):
                    val = str(val.strip(), "utf-8")
                    # "ID","Erwachsene","Kinder","Erstellungsdatum","Zuweisung Wohnung","Parkplatz gewünscht","Zuweisung Parkplatz","Zuweisung Nebenräume","Telefonnummer","Status"
                    if str(header[i], "utf-8") == "ID":
                        new.import_id = int(val)  ## Bewerbungs-ID
                    elif str(header[i], "utf-8") == "Kinder":
                        new.children = val
                    elif str(header[i], "utf-8") == "Zuweisung Wohnung":
                        rental_ids = val.split("\n")
                        for r in rental_ids:
                            rentals.append(
                                RentalUnit.objects.get(name=r, building="Holligerhof 8")
                            )
                    elif str(header[i], "utf-8") == "Parkplatz gewünscht":
                        if val:
                            prep = ""
                            if new.note:
                                prep = "%s / " % new.note
                            new.note = prep + "Parkplatz: %s" % (val)
                    elif str(header[i], "utf-8") == "Status":
                        if val:
                            prep = ""
                            if new.note:
                                prep = "%s / " % new.note
                            new.note = prep + "eMonitor-Status: %s" % (val)

                if new.children:
                    fields.append("Children: %s" % new.children)
                if new.note:
                    fields.append("Note: %s" % new.note)

                ## Check if we have a contract already
                try:
                    contract = Contract.objects.get(import_id=new.import_id)
                    fields.append(
                        "===> FOUND EXISTING contract with import_id ID %s" % (contract.emonitor_id)
                    )
                except Contract.DoesNotExist:
                    contract = new
                    fields.append(
                        "Creating new contract with import_id ID %s" % (contract.emonitor_id)
                    )

                ## Add rental unit(s)
                if rentals:
                    contract.save()
                    for rental in rentals:
                        contract.rental_units.add(rental)
                        fields.append("Adding rental_unit %s, %s" % (rental.name, rental.building))
                    contract.save()
                else:
                    fields.append("No rental unit found. IGNORING!")

                response.append({"info": "Importing contract %d" % count, "objects": fields})
            count += 1
    return response


def run_cmd(response, info, command, **kwargs):
    messages = []
    # messages.append("Cmd: %s" % command)
    try:
        output = subprocess.check_output(command, **kwargs)
    except (subprocess.CalledProcessError, OSError) as e:
        if e.output:
            messages.append("Out: %s" % e.output)
        messages.append("ERROR: Failed with return code %s." % e.returncode)
        response.append({"info": info, "objects": messages})
        raise
    if output:
        messages.append("Out: %s" % output)
    messages.append("Done.")
    response.append({"info": info, "objects": messages})
    return output


def parse_transaction_file_camt(xmlfile):
    xml = xmlfile.read()
    data = {"log": []}

    try:
        camt_data = sepa_parser.parse_string(None, xml)
    except Exception as e:
        return {"type": "camt", "data": data, "error": f"SEPA parser error: {e}"}
    error = None
    doctype = camt_data["document_type"]
    try:
        data = read_camt(camt_data)
    except SepaReaderException as e:
        error = "Fehler beim Lesen der %s Datei: %s" % (doctype, e)
    return {"type": doctype, "data": data, "error": error}


def process_transaction_file(uploadfile, max_filesize=50 * 1024 * 1024, allowed_ext=None):
    if allowed_ext is None:
        allowed_ext = ["zip", "xml"]
    if uploadfile.size > max_filesize:
        return {"type": None, "data": None, "error": "Datei ist zu gross!"}
    ext = uploadfile.name[-3:]
    if ext not in allowed_ext:
        return {"type": None, "data": None, "error": f"Dateityp {ext} ist nicht erlaubt!"}
    if ext == "zip":
        ## Unpack and process zip files
        ret = None
        with zipfile.ZipFile(io.BytesIO(uploadfile.read())) as thezip:
            for zipinfo in thezip.infolist():
                with thezip.open(zipinfo) as thefile:
                    filedata = parse_transaction_file_camt(thefile)
                    if filedata["error"]:
                        return {"type": None, "data": None, "error": filedata["error"]}
                    filedata["data"]["log"].append(
                        {
                            "info": f"Imported {filedata['type']} file {zipinfo.filename} from ZIP-File {uploadfile.name}.",
                            "objects": [],
                        }
                    )
                    if not ret:
                        ret = filedata
                    else:
                        ## Merge data from files
                        if ret["type"] != filedata["type"]:
                            return {
                                "type": None,
                                "data": None,
                                "error": f"Incompatible document types in ZIP file: {ret['type']} vs. {filedata['type']}",
                            }
                        ret["data"]["log"].extend(filedata["data"]["log"])
                        ret["data"]["transactions"].extend(filedata["data"]["transactions"])
        return ret
    if ext == "xml":
        ret = parse_transaction_file_camt(uploadfile)
        ret["data"]["log"].append(
            {"info": f"Imported {ret['type']} file {uploadfile.name}.", "objects": []}
        )
        return ret
    if ext == "csv":
        return parse_transaction_file_csv(uploadfile)
    return {"type": None, "data": None, "error": f"Dateityp {ext} ist nicht implementiert!"}


def parse_transaction_file_csv(csvfile):
    csvreader = csv.reader(decode_from_iso8859(csvfile), delimiter=";", quotechar='"')
    header = True
    filetype = None
    data = []

    patterns = []
    ## New format
    patterns.append(
        {
            "type": "Post",
            "re": re.compile(
                "^GUTSCHRIFT (?P<account>CH[A-Z0-9]{19}) ABSENDER: (?P<person>.*) MITTEILUNGEN: (?P<note>.*)"
            ),
        }
    )
    patterns.append(
        {
            "type": "Post",
            "re": re.compile("^GUTSCHRIFT (?P<account>CH[A-Z0-9]{19}) ABSENDER: (?P<person>.*)"),
        }
    )
    patterns.append(
        {
            "type": "Bank",
            "re": re.compile(
                "^GUTSCHRIFT AUFTRAGGEBER: (?P<person>.*) MITTEILUNGEN: (?P<note>.*) SPESENBETRAG"
            ),
        }
    )
    patterns.append(
        {
            "type": "Bank",
            "re": re.compile(
                "^GUTSCHRIFT AUFTRAGGEBER: (?P<person>.*) MITTEILUNGEN: (?P<note>.*) REFERENZEN:"
            ),
        }
    )
    patterns.append(
        {"type": "Bank", "re": re.compile("^GUTSCHRIFT AUFTRAGGEBER: (?P<person>.*) SPESENBETRAG")}
    )
    patterns.append(
        {"type": "Bank", "re": re.compile("^GUTSCHRIFT AUFTRAGGEBER: (?P<person>.*) REFERENZEN:")}
    )
    ## Old format(?)
    patterns.append(
        {
            "type": "Bank",
            "re": re.compile(
                "^GUTSCHRIFT VON FREMDBANK.* AUFTRAGGEBER: (?P<person>.*) MITTEILUNGEN: (?P<note>.*) REFERENZEN:"
            ),
        }
    )  # 0
    patterns.append(
        {
            "type": "Bank",
            "re": re.compile(
                "^GUTSCHRIFT VON FREMDBANK.* AUFTRAGGEBER: (?P<person>.*) REFERENZEN:"
            ),
        }
    )  # 1
    patterns.append(
        {
            "type": "Post",
            "re": re.compile(
                "^GIRO AUS KONTO (?P<account>.*) ABSENDER: (?P<person>.*) MITTEILUNGEN: (?P<note>.*)"
            ),
        }
    )  # 2
    patterns.append(
        {
            "type": "Post",
            "re": re.compile("^GIRO AUS KONTO (?P<account>.*) ABSENDER: (?P<person>.*)"),
        }
    )  # 3
    patterns.append(
        {
            "type": "Post",
            "re": re.compile(
                "^GIRO AUS KONTO POSTFINANCE MOBILE (?P<account>[A-Z0-9]*) (?P<person>.*) MITTEILUNGEN: (?P<note>.*)"
            ),
        }
    )  # 4
    patterns.append(
        {
            "type": "Post",
            "re": re.compile(
                "^GIRO AUS KONTO POSTFINANCE MOBILE (?P<account>[A-Z0-9]*) (?P<person>.*)"
            ),
        }
    )  # 5
    patterns.append(
        {
            "type": "EZS",
            "re": re.compile(
                "^EINZAHLUNGSSCHEIN/QR-ZAHLTEIL ABSENDER: (?P<person>.*) REFERENZEN:"
            ),
        }
    )  # 6
    patterns.append({"type": "EZS", "re": re.compile("^EINZAHLUNGSSCHEIN")})  # 7

    for row in csvreader:
        if header:
            if row[0:5] == [
                "Buchungsdatum",
                "Avisierungstext",
                "Gutschrift in CHF",
                "Lastschrift in CHF",
                "Valuta",
            ]:
                filetype = "Postfinance"
                header = False
        elif len(row) > 2 and len(row[0]) == 10 and len(row[1]) and len(row[2]):
            ## Normalize date format
            date_str = row[0]
            date_patterns = ["%Y-%m-%d", "%d.%m.%Y"]
            for pattern in date_patterns:
                with contextlib.suppress(builtins.BaseException):
                    date_str = datetime.date.strftime(
                        datetime.datetime.strptime(row[0], pattern), "%d.%m.%Y"
                    )
            d = {
                "date": date_str,
                "person": None,
                "note": None,
                "account": None,
                "type": None,
                "received": False,
                "amount": 0,
            }
            text = row[1].strip()
            for p in patterns:
                m = p["re"].match(text)
                if m:
                    d["type"] = p["type"]
                    # print('Match found: %s' % (text))
                    for g in list(m.groupdict().items()):
                        # print(' %s -> %s' % g)
                        d[g[0]] = g[1]
                    # print('-------------')
                    d["type"] = p["type"]
                    d["received"] = True
                    d["amount"] = float(row[2])
                    break
            if d["type"] == "EZS":
                d["note"] = "Einzahlungsschein"

            data.append(d)

    if not filetype:
        return {"type": "csv", "data": None, "error": "Dateityp unbekannt"}
    elif data:
        return {"type": "csv", "data": data, "error": None}
    else:
        return {"type": "csv", "data": None, "error": "Keine Daten gefunden"}


def import_adit_serial():
    response = []

    filename = "/tmp/ADIT Belegungsplan.xlsm"
    wb = load_workbook(filename)
    if "ADIT1000" not in wb.sheetnames:
        response.append(
            {"info": "ERROR: Sheet ADIT1000 not found! (%s)" % str(wb.sheetname), "objects": []}
        )
        return response

    ws = wb["ADIT1000"]

    col_serial1 = 6
    col_notes = 10

    row_list_start = 17
    row_list_max = 1014

    serials = {}  ## serial -> room
    rooms = {}  ## room -> serial
    count_serials = 0
    count_rooms = 0

    room_pattern = re.compile(r"^(Wohnung|Joker) (\d{3})\b")
    for row in range(row_list_start, row_list_max):
        serial1 = ws.cell(row=row, column=col_serial1).value
        notes = ws.cell(row=row, column=col_notes).value
        # print("No=%s, Name=%s, Serial=%s, Notes=%s" % (ws.cell(row=row, column=col_nr).value, ws.cell(row=row, column=col_name).value, serial1, notes))
        if not serial1:
            # print("No serial1, assuming end of list!")
            break
        ## Parse room number
        match = room_pattern.match(notes)
        if match:
            # print("/%s/%s/" % (match.group(1),match.group(2)))
            room = match.group(2)
            if serial1 in serials:
                if serials[serial1] != room:
                    response.append(
                        {
                            "info": "ERROR: Multiple rooms for serial %s! (parsed room=%s)"
                            % (serial1, room),
                            "objects": [],
                        }
                    )
                    return response
                    # raise RuntimeError("Multiple rooms for serial %s! (parsed room=%s)" % (serial1, room))
                    # return
            else:
                serials[serial1] = room
                count_serials += 1
            if room in rooms:
                if rooms[room] != serial1:
                    response.append(
                        {
                            "info": "ERROR: Multiple serials for room %s! (parsed serial1=%s)"
                            % (room, serial1),
                            "objects": [],
                        }
                    )
                    return response
                    # raise RuntimeError("Multiple serials for room %s! (parsed serial1=%s)" % (room, serial1))
                    # return
            else:
                rooms[room] = serial1
                count_rooms += 1

        else:
            response.append({"info": "ERROR: No match! notes=%s" % notes, "objects": []})
            return response
            # raise RuntimeError("No match! notes=%s" % notes)
            # return 2

    response.append(
        {"info": "Read %d serials and %d rooms." % (count_serials, count_rooms), "objects": []}
    )
    # print("Imported %d serials and %d rooms." % (count_serials, count_rooms))
    # print(serials)

    update_obj = []
    added_obj = []

    for room in rooms:
        ru = RentalUnit.objects.get(name=room)
        if ru.adit_serial:
            if str(ru.adit_serial) != str(rooms[room]):
                update_obj.append(
                    "Update serial %s: %s -> %s" % (room, ru.adit_serial, rooms[room])
                )
                ru.adit_serial = rooms[room]
                ru.save()
        else:
            added_obj.append("Add serial %s: %s" % (room, rooms[room]))
            ru.adit_serial = rooms[room]
            ru.save()

    if update_obj:
        response.append(
            {"info": "Updated %d rental units:" % len(update_obj), "objects": update_obj}
        )
    if added_obj:
        response.append(
            {"info": "Added serial to %d rental units:" % len(added_obj), "objects": added_obj}
        )

    return response
