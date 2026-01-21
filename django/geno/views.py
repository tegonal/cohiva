import datetime
import io
import json
import os
import re
import subprocess
import tempfile
import zipfile
from collections import OrderedDict

from dateutil.relativedelta import relativedelta
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.contenttypes.models import ContentType
from django.core.exceptions import ImproperlyConfigured, PermissionDenied
from django.db.models import Q
from django.forms import formset_factory
from django.http import FileResponse, Http404, HttpResponse, HttpResponseRedirect
from django.shortcuts import redirect, render
from django.template import Context
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.utils.encoding import smart_str
from django.utils.http import url_has_allowed_host_and_scheme
from django.utils.translation import gettext_lazy as _
from django.views.generic import FormView, TemplateView
from django_tables2 import RequestConfig
from oauthlib.oauth2 import TokenExpiredError

## For OAuth client
from requests_oauthlib import OAuth2Session
from stdnum.ch import esr

from finance.accounting import Account, AccountingManager, AccountKey

if hasattr(settings, "SHARE_PLOT") and settings.SHARE_PLOT:
    ## For Plotting
    # from matplotlib import pyplot as plt
    import io

    from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
    from matplotlib.figure import Figure

if hasattr(settings, "CREATESEND_API_KEY") and settings.CREATESEND_API_KEY:
    import createsend

if hasattr(settings, "MAILMAN_API") and settings.MAILMAN_API["password"]:
    import mailmanclient

import geno.settings as geno_settings
from cohiva.views.admin import CohivaAdminViewMixin, ResponseVariant

from .billing import (
    add_invoice,
    consolidate_invoices,
    create_invoices,
    create_qrbill,
    get_duplicate_invoices,
    get_reference_nr,
    guess_transaction,
    invoice_detail,
    invoice_overview,
    pay_invoice,
    process_sepa_transactions,
    process_transaction,
)
from .documents import context_format, create_documents, get_context_data, send_member_mail_process
from .exporter import (
    export_addresses_carddav,
    export_adit_file,
    export_data_to_xls,
    generate_demo_camt053_file,
)
from .forms import (
    InvoiceFilterForm,
    ManualInvoiceForm,
    ManualInvoiceLineForm,
    MemberMailActionForm,
    MemberMailForm,
    MemberMailSelectForm,
    Odt2PdfForm,
    SendInvoicesForm,
    ShareOverviewFilterForm,
    ShareStatementForm,
    TransactionForm,
    TransactionFormInvoice,
    TransactionUploadForm,
    TransactionUploadProcessForm,
    WebstampForm,
    process_registration_forms,
)
from .importer import (
    import_adit_serial,
    import_codes_from_file,
    import_contracts_from_file,
    import_emonitor_addresses_from_file,
    import_emonitor_children_from_file,
    import_emonitor_contracts_from_file,
    import_keller_from_file,
    import_members_from_file,
    import_rentalunits_from_file,
    process_eigenmittel,
    process_transaction_file,
    update_address_from_file,
)
from .models import (
    Address,
    Building,
    Contract,
    Document,
    DocumentType,
    GenericAttribute,
    Invoice,
    InvoiceCategory,
    Member,
    MemberAttribute,
    MemberAttributeType,
    RentalUnit,
    Share,
    ShareType,
    get_active_contracts,
    get_active_shares,
)
from .shares import (
    check_rental_shares_report,
    create_interest_transactions,
    get_share_statement_data,
    share_interest_calc,
)
from .tables import (
    MemberTable,
    MemberTableAdmin,
)
from .utils import fill_template_pod, is_member, nformat, odt2pdf

# from .decorators import login_required

if "portal" in settings.INSTALLED_APPS:
    import portal.auth as portal_auth

import logging

logger = logging.getLogger("geno")


def unauthorized(request):
    c = {
        "response": [{"info": "Sie haben keine Berechtigung für diese Aktion."}],
        "title": "Keine Berechtigung",
    }
    return render(request, "geno/default.html", c)


@login_required
def import_generic(request, what):
    if not request.user.has_perm("geno.admin_import"):
        return unauthorized(request)
    if not settings.ALLOW_IMPORT:
        title = "Error"
        ret = [
            {"info": "Import is disabled in settings."},
        ]
    else:
        if what == "members":
            title = "Mitgliederliste importieren"
            ret = import_members_from_file(empty_tables_first=False)
        elif what == "address":
            title = "Adressen aktualisieren"
            ret = update_address_from_file()
        elif what == "contracts":
            title = "Verträge importieren"
            ret = import_contracts_from_file(empty_tables_first=True)
        elif what == "rentalunits":
            title = "Mietobjekte aus eMonitor importieren"
            ret = import_rentalunits_from_file(empty_tables_first=True)
        elif what == "emonitorcontracts":
            title = "Zuweisungen aus eMonitor importieren"
            ret = import_emonitor_contracts_from_file(empty_tables_first=True)
        elif what == "emonitoraddresses":
            title = "Adressen/Vertragsparteien aus eMonitor importieren"
            ret = import_emonitor_addresses_from_file(empty_tables_first=True)
        elif what == "emonitorchildren":
            title = "Kinder aus eMonitor importieren"
            ret = import_emonitor_children_from_file()
        elif what == "keller":
            title = "Mieterkeller importieren"
            ret = import_keller_from_file(empty_tables_first=True)
        elif what == "eigenmittel":
            title = "Eigenmittelliste abgleichen"
            # ret = process_eigenmittel()
            return process_eigenmittel()
        elif what == "codes":
            title = "Codes importieren"
            ret = import_codes_from_file(empty_tables_first=False)
        elif what == "adit_serial":
            title = "ADIT Seriennummern importieren"
            ret = import_adit_serial()
        else:
            raise Http404("Ungültige Adresse")
    c = {"response": ret, "title": title}
    return render(request, "geno/default.html", c)


def export_generic(request, what):
    if what == "addresses_carddav":
        title = "Adressen nach CardDAV exportieren"
        ret = export_addresses_carddav(request.GET.get("delete", "") == "yes")
    elif what == "adit":
        title = "ADIT Gegensprechanlage exportieren"
        return export_adit_file()
    c = {"response": ret, "title": title}
    return render(request, "geno/default.html", c)


@login_required
def documents(request, doctype, obj_id, action):
    try:
        doctype_obj = DocumentType.objects.get(name=doctype)
        if action == "create":
            ## Create new document and attach to content_object
            if not request.user.has_perm("geno.add_document"):
                return unauthorized(request)
            data = get_context_data(doctype, obj_id, {})
        elif action == "download":
            ## Regenerate document
            if not request.user.has_perm("geno.regenerate_document"):
                return unauthorized(request)
            doc = Document.objects.get(pk=obj_id)
            data = {"visible_filename": doc.name, "context": json.loads(doc.data)}
        else:
            raise RuntimeError("Action '%s' is not implemented." % action)
        filename = fill_template_pod(
            doctype_obj.template.file.path, context_format(data["context"]), output_format="odt"
        )
        if not filename:
            raise RuntimeError("Could not fill template")
        if "content_object" in data:
            ## Attach document data to object
            d = Document(
                name=data["visible_filename"],
                doctype=doctype_obj,
                data=json.dumps(data["context"]),
                content_object=data["content_object"],
            )
            d.save()
        resp = FileResponse(
            open(filename, "rb"), as_attachment=True, filename=smart_str(data["visible_filename"])
        )  # content_type = "application/pdf")
        os.remove(filename)
        return resp
    except (RuntimeError, DocumentType.DoesNotExist) as e:
        error = "Fehler beim Erzeugen des Dokuments: %s" % e
    ret = [
        {"info": "ERROR: %s" % error},
    ]
    return render(request, "geno/default.html", {"response": ret, "title": "Document error"})


class ShareOverviewView(CohivaAdminViewMixin, TemplateView):
    title = "Übersicht Beteiligungen"
    permission_required = "geno.canview_share_overview"
    template_name = "geno/share_overview.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Parse date from URL parameter
        reference_date = None
        if len(self.request.GET.get("date", "")) == 10:
            try:
                reference_date = datetime.datetime.strptime(
                    self.request.GET.get("date"), "%Y-%m-%d"
                ).date()
                context["reference_date_str"] = reference_date.strftime("%Y-%m-%d")
            except ValueError:
                pass

        # Initialize the filter form - always show today's date as default
        form_initial = {"date": reference_date if reference_date else datetime.date.today()}

        form = ShareOverviewFilterForm(initial=form_initial)
        context["form"] = form
        context["reference_date"] = reference_date

        # Get share type statistics
        share_stats = []
        total_value = 0

        try:
            stype_AS = ShareType.objects.get(name="Anteilschein")
        except ShareType.DoesNotExist:
            stype_AS = None

        for share_type in ShareType.objects.all():
            stat = {"quantity": 0, "value": 0, "last_date": None}
            for s in get_active_shares(date=reference_date).filter(share_type=share_type):
                stat["quantity"] += s.quantity
                stat["value"] += s.quantity * s.value
                if stat["last_date"] is None or s.date > stat["last_date"]:
                    stat["last_date"] = s.date

            share_stats.append(
                {
                    "type": str(share_type),
                    "quantity": nformat(stat["quantity"], 0),
                    "value": nformat(stat["value"]),
                    "last_date": stat["last_date"],
                }
            )
            total_value += stat["value"]

        context["share_stats"] = share_stats
        context["total_value"] = nformat(total_value)

        # Check for non-members with shares (warning)
        non_members = []
        if not reference_date and self.request.user.has_perm("geno.canview_share") and stype_AS:
            for s in get_active_shares(date=reference_date).filter(share_type=stype_AS):
                try:
                    m = Member.objects.get(name=s.name)
                    if m.date_leave:
                        non_members.append(
                            {
                                "name": s.name,
                                "quantity": s.quantity,
                                "date_leave": m.date_leave,
                            }
                        )
                except Member.DoesNotExist:
                    non_members.append(
                        {
                            "name": s.name,
                            "quantity": s.quantity,
                            "date_leave": None,
                        }
                    )

        context["non_members"] = non_members

        # Show plot if enabled and viewing current data (today or no date specified)
        is_today = not reference_date or reference_date == datetime.date.today()
        context["show_plot"] = is_today and hasattr(settings, "SHARE_PLOT") and settings.SHARE_PLOT

        return context

    def post(self, request, *args, **kwargs):
        form = ShareOverviewFilterForm(request.POST)
        if form.is_valid():
            date = form.cleaned_data.get("date")
            if date:
                # Redirect with date parameter
                target = f"{request.path}?date={date.strftime('%Y-%m-%d')}"
                if url_has_allowed_host_and_scheme(target, allowed_hosts={request.get_host()}):
                    return redirect(target)
                else:
                    return redirect("/")
            else:
                # Redirect without date parameter (show current date)
                target = request.path
                if url_has_allowed_host_and_scheme(target, allowed_hosts={request.get_host()}):
                    return redirect(target)
                else:
                    return redirect("/")
        return self.get(request, *args, **kwargs)


def member_overview_plot(request):
    if not request.user.has_perm("geno.canview_member_overview"):
        return unauthorized(request)

    if not (hasattr(settings, "SHARE_PLOT") and settings.SHARE_PLOT):
        raise Http404("Plot not found")

    # age_limits = (20,30,40,50,60,70,80,1000)

    changes = []
    for m in Member.objects.all():
        ## Age stat
        # if m.name.date_birth:
        #    born = m.name.date_birth
        #    today = datetime.date.today()
        #    age = today.year - born.year - ((today.month, today.day) < (born.month, born.day))
        #    for i,limit in enumerate(age_limits):
        #        if age < limit:
        #            age_stat['u%d' % limit] += 1
        #            break
        # else:
        #    age_stat['Unbekannt'] += 1

        changes.append({"date": m.date_join, "value": 1})
        if m.date_leave:
            changes.append({"date": m.date_leave, "value": -1})

    changes_sorted = sorted(changes, key=lambda x: x["date"])
    x = []
    y = []
    agg = 0
    for data in changes_sorted:
        if data["date"] > datetime.date.today():
            break
        x.append(data["date"])
        agg += data["value"]
        y.append(agg)

    fig = Figure(figsize=(10, 7))
    ax = fig.add_subplot(111)
    ax.plot(x, y, "b-")
    ax.set_xlabel("Jahr")
    ax.set_ylabel("Anzahl Mitglieder")
    ax.grid(axis="x", linestyle=":")
    ax.grid(axis="y", linestyle=":")
    canvas = FigureCanvas(fig)
    output = io.BytesIO()
    canvas.print_png(output)

    response = HttpResponse(output.getvalue(), content_type="image/png")
    output.close()
    return response


@login_required
def share_overview_boxplot(request):
    if not request.user.has_perm("geno.canview_share_overview"):
        return unauthorized(request)

    if not (hasattr(settings, "SHARE_PLOT") and settings.SHARE_PLOT):
        raise Http404("Plot not found")

    today = datetime.datetime.today()
    try:
        stype_DarlehenSpezial = ShareType.objects.get(name="Darlehen spezial")
    except ShareType.DoesNotExist:
        stype_DarlehenSpezial = None
    try:
        stype_Hypothek = ShareType.objects.get(name="Hypothek")
    except ShareType.DoesNotExist:
        stype_Hypothek = None

    ## Statisik: Beteiligung pro Typ und Mitglieder
    stat = []
    labels = []
    for share_type in ShareType.objects.exclude(name="Darlehen spezial").exclude(name="Hypothek"):
        stat_share = []
        for m in Member.objects.filter(Q(date_leave=None) | Q(date_leave__gt=today)):
            total = 0
            for s in get_active_shares().filter(name=m.name).filter(share_type=share_type):
                total += s.quantity * float(s.value)
            if total >= 1000:
                stat_share.append(total / 1000.0)
        labels.append("%s\nn=%d" % (share_type.name, len(stat_share)))
        stat.append(stat_share)

    ## Statistik: Beteiligung Total pro Mitglied
    count_below_threshold = 0
    stat_share = []
    for m in Member.objects.filter(Q(date_leave=None) | Q(date_leave__gt=today)):
        total = 0
        for s in (
            get_active_shares()
            .filter(name=m.name)
            .exclude(share_type=stype_DarlehenSpezial)
            .exclude(share_type=stype_Hypothek)
        ):
            total += s.quantity * float(s.value)
        if total >= 1000:
            stat_share.append(total / 1000.0)
        else:
            count_below_threshold += 1
    stat.append(stat_share)
    labels.append("%s\nn=%d" % ("Gesamt", len(stat_share)))

    ## Statistik: Beteiligung Nichtmitglieder
    stat_share = []
    for a in Address.objects.filter(active=True):
        try:
            m = Member.objects.get(name=a.id)
        except Member.DoesNotExist:
            total = 0
            for s in (
                get_active_shares()
                .filter(name=a.id)
                .exclude(share_type=stype_DarlehenSpezial)
                .exclude(share_type=stype_Hypothek)
            ):
                total += s.quantity * float(s.value)
            if total >= 1000:
                stat_share.append(total / 1000.0)
    stat.append(stat_share)
    labels.append("%s\nn=%d" % ("Personen ohne Mitgliedschaft", len(stat_share)))

    fig = Figure(figsize=(10, 7))
    ax = fig.add_subplot(111)
    ax.boxplot(stat, tick_labels=labels)
    ax.set_ylabel("Betrag (kFr.)")
    ax.axvline(5.5)
    ax.set_title(
        "Beteiligungen ab Fr. 1000.- von Mitgliedern und Personen ohne Mitgliedschaft "
        "(ohne Darlehen spez./Hypothek)\n"
        "(%d Mitglieder sind mit weniger als Fr. 1000.- beteiligt)\n"
        "Rot = Median, Box = Quartile (50%% der Daten), Linien = Min-Max ohne Ausreisser (1.5-IQR)"
        % count_below_threshold
    )
    ax.grid(axis="y", linestyle=":")
    canvas = FigureCanvas(fig)
    output = io.BytesIO()
    canvas.print_png(output)

    response = HttpResponse(output.getvalue(), content_type="image/png")
    output.close()
    return response


class MemberOverviewView(CohivaAdminViewMixin, TemplateView):
    title = "Mitgliederspiegel"
    permission_required = "geno.canview_member_overview"
    template_name = "geno/member_overview.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        stats = self.calculate_member_statistics()
        context["stats"] = stats
        context["show_plot"] = hasattr(settings, "SHARE_PLOT") and settings.SHARE_PLOT
        return context

    def calculate_member_statistics(self):
        """Calculate member statistics for gender and age distribution."""
        gender_stat = OrderedDict(
            [
                ("Total", 0),
                ("Frauen", 0),
                ("Männer", 0),
                ("Divers", 0),
                ("Organisationen", 0),
                ("Andere/Unbekannt", 0),
            ]
        )

        age_limits = (30, 45, 60, 1000)
        age_stat = {}
        for limit in age_limits:
            age_stat[f"u{limit}"] = 0
        age_stat["Unbekannt"] = 0

        # Determine reference date
        date_mode = "strict"
        today = datetime.date.today()
        reference_date = today
        if self.request.GET.get("date", "") == "last_year":
            date_mode = "last_year"
            reference_date = datetime.date(today.year - 1, 12, 31)

        # Collect statistics
        for m in Member.objects.all():
            if is_member(m.name, date_mode=date_mode):
                gender_stat["Total"] += 1

                # Gender statistics
                if m.name.title == "Org" or m.name.organization:
                    gender_stat["Organisationen"] += 1
                elif m.name.title == "Herr":
                    gender_stat["Männer"] += 1
                elif m.name.title == "Divers":
                    gender_stat["Divers"] += 1
                elif m.name.title == "Frau":
                    gender_stat["Frauen"] += 1
                else:
                    gender_stat["Andere/Unbekannt"] += 1

                # Age statistics
                if m.name.date_birth:
                    born = m.name.date_birth
                    age = (
                        reference_date.year
                        - born.year
                        - ((reference_date.month, reference_date.day) < (born.month, born.day))
                    )
                    for limit in age_limits:
                        if age < limit:
                            age_stat[f"u{limit}"] += 1
                            break
                else:
                    age_stat["Unbekannt"] += 1

        # Format data for template
        total = gender_stat["Total"]

        if total == 0:
            return None

        # Format gender statistics
        gender_data = []
        for key, count in gender_stat.items():
            percentage = round(float(count) / float(total) * 100.0)
            gender_data.append(
                {
                    "label": key,
                    "count": count,
                    "percentage": percentage,
                }
            )

        # Format age statistics
        age_data = []
        last = 0
        for limit in age_limits:
            count = age_stat[f"u{limit}"]
            percentage = round(float(count) / float(total) * 100.0)
            if limit == 1000:
                label = f"Über {last}"
            else:
                label = f"{last} - {limit}"
            age_data.append(
                {
                    "label": label,
                    "count": count,
                    "percentage": percentage,
                }
            )
            last = limit

        if age_stat["Unbekannt"]:
            count = age_stat["Unbekannt"]
            percentage = round(float(count) / float(total) * 100.0)
            age_data.append(
                {
                    "label": "Unbekannt",
                    "count": count,
                    "percentage": percentage,
                }
            )

        return {
            "gender": gender_data,
            "age": age_data,
        }


@login_required
def member_list(request):
    if not request.user.has_perm("geno.canview_member"):
        return unauthorized(request)
    if "q" in request.GET:
        members = Member.objects.exclude(date_leave__isnull=False).filter(
            Q(name__organization__icontains=request.GET.get("q"))
            | Q(name__name__icontains=request.GET.get("q"))
            | Q(name__first_name__icontains=request.GET.get("q"))
            | Q(name__street_name__icontains=request.GET.get("q"))
            | Q(name__city_name__icontains=request.GET.get("q"))
            | Q(name__email__icontains=request.GET.get("q"))
            | Q(name__telephone__icontains=request.GET.get("q"))
            | Q(name__mobile__icontains=request.GET.get("q"))
            | Q(name__occupation__icontains=request.GET.get("q"))
        )
    else:
        members = Member.objects.exclude(date_leave__isnull=False)
    table = MemberTable(members)
    RequestConfig(request, paginate={"per_page": 50}).configure(table)
    return render(request, "geno/table.html", {"table": table})


@login_required
def member_list_admin(request):
    if not request.user.has_perm("geno.canview_billing"):
        return unauthorized(request)
    if "q" in request.GET:
        members = Member.objects.exclude(date_leave__isnull=False).filter(
            Q(name__organization__icontains=request.GET.get("q"))
            | Q(name__name__icontains=request.GET.get("q"))
            | Q(name__first_name__icontains=request.GET.get("q"))
            | Q(name__street_name__icontains=request.GET.get("q"))
            | Q(name__city_name__icontains=request.GET.get("q"))
            | Q(name__email__icontains=request.GET.get("q"))
            | Q(name__telephone__icontains=request.GET.get("q"))
            | Q(name__mobile__icontains=request.GET.get("q"))
            | Q(name__occupation__icontains=request.GET.get("q"))
            | Q(name__bankaccount__icontains=request.GET.get("q"))
        )
    else:
        members = Member.objects.exclude(date_leave__isnull=False)
    table = MemberTableAdmin(members)
    RequestConfig(request, paginate={"per_page": 50}).configure(table)
    return render(request, "geno/table.html", {"table": table})


@login_required
def address_export(request, show_wohnung=True):
    import openpyxl
    from openpyxl.styles import Font

    if not request.user.has_perm("geno.canview_member"):
        return unauthorized(request)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        "attachment; filename=%s_Adressen.xlsx" % settings.GENO_FILENAME_STR
    )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Adressliste %s" % settings.GENO_FILENAME_STR

    row_num = 0

    ## Header
    columns = [
        ("A", "ID", 7),
        ("B", "Organisation", 18),
        ("C", "Name", 18),
        ("D", "Vorname", 15),
        ("E", "Anrede", 8),
        ("F", "Du/Sie", 7),
        ("G", "Adresszusatz", 22),
        ("H", "Adresse", 22),
        ("I", "PLZ Ort", 22),
        ("J", "Telefon", 15),
        ("K", "Mobile", 15),
        ("L", "Email", 25),
        ("M", "Geb.dat.", 15),
        ("N", "Heimatort", 20),
        ("O", "Beruf/Ausbildung", 40),
        ("P", "Mitglied seit", 15),
        ("Q", "Austritt", 15),
        ("R", geno_settings.MEMBER_FLAGS[1], 10),
        ("S", geno_settings.MEMBER_FLAGS[2], 10),
        ("T", geno_settings.MEMBER_FLAGS[3], 10),
        ("U", geno_settings.MEMBER_FLAGS[4], 10),
        ("V", geno_settings.MEMBER_FLAGS[5], 10),
        ("W", "Status Mitgliedschaft", 22),
        ("X", "Anteilschein bezahlt", 15),
    ]
    if show_wohnung:
        columns.append(("Y", "Wohnung", 15))
        columns.append(("Z", "Kinder", 25))
    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][1]
        c.font = Font(bold=True)
        # set column width
        ws.column_dimensions[columns[col_num][0]].width = columns[col_num][2]

    ## Data
    try:
        matt_type = MemberAttributeType.objects.get(name="Status Mitgliedschaft")
    except:
        matt_type = None
    try:
        stype01 = ShareType.objects.get(name="Anteilschein Einzelmitglied")
    except:
        stype01 = None
    for a in Address.objects.filter(active=True):
        row = []
        row.append(a.pk)
        row.append(a.organization)
        row.append(a.name)
        row.append(a.first_name)
        row.append(a.title)
        row.append(a.formal)
        row.append(a.extra)
        row.append(a.street)
        row.append(a.city)
        row.append(a.telephone)
        row.append(a.mobile)
        row.append(a.email)
        row.append(a.date_birth)
        row.append(a.hometown)
        row.append(a.occupation)
        date_join = ""
        date_leave = ""
        flag_01 = ""
        flag_02 = ""
        flag_03 = ""
        flag_04 = ""
        flag_05 = ""
        status = ""
        share01_paid = ""
        wohnung = ""
        kinder = ""
        try:
            m = Member.objects.get(name=a)
            if m.date_leave:
                date_leave = m.date_leave
                # date_join = m.date_join
            else:
                date_join = m.date_join
            if m.flag_01:
                flag_01 = "X"
            else:
                flag_01 = "--"
            if m.flag_02:
                flag_02 = "X"
            else:
                flag_02 = "--"
            if m.flag_03:
                flag_03 = "X"
            else:
                flag_03 = "--"
            if m.flag_04:
                flag_04 = "X"
            else:
                flag_04 = "--"
            if m.flag_05:
                flag_05 = "X"
            else:
                flag_05 = "--"
            if matt_type:
                matt = MemberAttribute.objects.filter(member=m, attribute_type=matt_type).first()
                if matt:
                    status = matt.value
        except Member.DoesNotExist:
            pass

        if stype01:
            share = get_active_shares().filter(share_type=stype01).filter(name=a).first()
            if share:
                share01_paid = share.date.strftime("%d.%m.%Y")

        if show_wohnung:
            for c in get_active_contracts().filter(contractors__pk=a.pk):
                for child in c.children.all():
                    if kinder == "":
                        kinder = "%s %s (%s)" % (
                            child.name.first_name,
                            child.name.name,
                            child.age(precision=0),
                        )
                    else:
                        kinder = "%s, %s %s (%s)" % (
                            kinder,
                            child.name.first_name,
                            child.name.name,
                            child.age(precision=0),
                        )
                for w in c.rental_units.exclude(rental_type="Kellerabteil"):
                    if wohnung == "":
                        wohnung = w.name
                    else:
                        wohnung = "%s/%s" % (wohnung, w.name)
        row.append(date_join)
        row.append(date_leave)
        row.append(flag_01)
        row.append(flag_02)
        row.append(flag_03)
        row.append(flag_04)
        row.append(flag_05)
        row.append(status)
        row.append(share01_paid)
        if show_wohnung:
            row.append(wohnung)
            row.append(kinder)

        row_num += 1
        for col_num in range(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]

    wb.save(response)
    return response


@login_required
def share_export(request):
    import openpyxl
    from openpyxl.styles import Font

    if not request.user.has_perm("geno.canview_share"):
        return unauthorized(request)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        "attachment; filename=%s_Beteiligungen.xlsx" % settings.GENO_FILENAME_STR
    )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Beteiligungen %s" % settings.GENO_FILENAME_STR

    row_num = 0

    ## Header
    if request.GET.get("aggregate", "") == "yes":
        columns = [
            ("A", "Name", 30),
            ("B", "Typ", 20),
            ("C", "Anzahl", 8),
            ("D", "Summe CHF", 12),
            ("E", "Fälligkeit", 10),
            ("F", "Laufzeit", 8),
            ("G", "Mitgliedschaft", 30),
        ]
    else:
        columns = [
            ("A", "ID", 7),
            ("B", "Name", 30),
            ("C", "Adress-ID", 10),
            ("D", "Typ", 20),
            ("E", "Datum", 12),
            ("F", "Anzahl", 8),
            ("G", "Betrag CHF", 12),
            ("H", "Summe CHF", 12),
            ("I", "Zinssatz-Modus", 10),
            ("J", "Zinssatz", 10),
            ("K", "Zins?", 8),
            ("L", "Zusatzinfo", 30),
            ("M", "Erstellt", 22),
            ("N", "Geändert", 22),
        ]
    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][1]
        c.font = Font(bold=True)
        # set column width
        ws.column_dimensions[columns[col_num][0]].width = columns[col_num][2]

    ## Data
    today = datetime.datetime.today()
    if request.GET.get("jahresende", "") == "yes":
        valuta_date = datetime.date(today.year - 1, 12, 31)
    else:
        valuta_date = today
    if request.GET.get("anonymous", "") == "yes":
        anonymous = True
    else:
        anonymous = False
    person_count = 0
    person_nr = {}
    if request.GET.get("aggregate", "") == "yes":
        last = None
        row = []
        for a in get_active_shares(date=valuta_date).order_by("share_type", "name", "date"):
            if a.date_due:
                duedate = a.date_due
            elif a.duration:
                duedate = a.date + relativedelta(years=a.duration)
            else:
                duedate = ""

            if last == str(a.name) + str(a.share_type) + str(duedate):
                ## Update row
                row[2] += a.quantity
                row[3] += a.quantity * a.value
            else:
                ## Insert row
                if last is not None:
                    row_num += 1
                    for col_num in range(len(row)):
                        c = ws.cell(row=row_num + 1, column=col_num + 1)
                        c.value = row[col_num]
                row = []
                # row.append(a.pk)
                if anonymous:
                    if a.name.pk not in person_nr:
                        person_count += 1
                        person_nr[a.name.pk] = "Person %s" % person_count
                    person_name = person_nr[a.name.pk]
                else:
                    person_name = str(a.name)
                row.append(person_name)
                row.append(str(a.share_type))
                row.append(a.quantity)
                row.append(a.quantity * a.value)
                row.append(duedate)
                row.append(a.duration)
                if is_member(a.name):
                    row.append("Ja")
                else:
                    row.append("Nein")
            last = str(a.name) + str(a.share_type) + str(duedate)
        if last is not None:
            row_num += 1
            for col_num in range(len(row)):
                c = ws.cell(row=row_num + 1, column=col_num + 1)
                c.value = row[col_num]
    else:
        for a in get_active_shares(date=valuta_date).order_by("-date"):
            row = []
            row.append(a.pk)
            row.append(str(a.name))
            row.append(a.name.pk)
            row.append(str(a.share_type))
            row.append(a.date)
            row.append(a.quantity)
            row.append(a.value)
            row.append(a.quantity * a.value)
            row.append(a.interest_mode)
            row.append(a.interest())
            if a.is_interest_credit:
                row.append(1)
            else:
                row.append(0)
            row.append(a.note)
            row.append(timezone.localtime(a.ts_created).strftime("%Y-%m-%d %H:%M"))
            row.append(timezone.localtime(a.ts_modified).strftime("%Y-%m-%d %H:%M"))
            row_num += 1
            for col_num in range(len(row)):
                c = ws.cell(row=row_num + 1, column=col_num + 1)
                c.value = row[col_num]

    wb.save(response)
    return response


class DebtorView(CohivaAdminViewMixin, TemplateView):
    title = "Debitoren"
    permission_required = ("geno.canview_billing", "geno.transaction", "geno.transaction_invoice")
    template_name = "geno/debtor.html"
    action = "overview"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = self.get_form()
        if self.action == "overview":
            context.update(self.debtor_list())
        elif self.action == "detail":
            context.update(self.debtor_detail(self.kwargs["key_type"], self.kwargs["key"]))
        return context

    def get(self, request, *args, **kwargs):
        if "invoice_filter" not in request.session or request.GET.get("reset_filter", "0") == "1":
            request.session["invoice_filter"] = {"category_filter": "_all"}
            if request.GET.get("reset_filter", "0") == "1" and url_has_allowed_host_and_scheme(
                request.path, allowed_hosts=None
            ):
                ## Reload to get rid of get request argument
                return HttpResponseRedirect(request.path)
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        return self.get(request, *args, **kwargs)

    def get_form(self):
        initial = self.request.session["invoice_filter"].copy()
        if self.action == "detail":
            if "search_detail" in initial:
                initial["search"] = initial["search_detail"]
            else:
                initial["search"] = ""
        if self.request.POST:
            form = InvoiceFilterForm(self.request.POST)
            if form.is_valid():
                if self.action == "detail":
                    self.request.session["invoice_filter"]["search_detail"] = form.cleaned_data[
                        "search"
                    ]
                else:
                    self.request.session["invoice_filter"]["search"] = form.cleaned_data["search"]
                self.request.session["invoice_filter"]["category_filter"] = form.cleaned_data[
                    "category_filter"
                ]
                self.request.session["invoice_filter"]["show_consolidated"] = form.cleaned_data[
                    "show_consolidated"
                ]
                if form.cleaned_data["date_from"]:
                    date_from = form.cleaned_data["date_from"].strftime("%d.%m.%Y")
                else:
                    date_from = ""
                if form.cleaned_data["date_to"]:
                    date_to = form.cleaned_data["date_to"].strftime("%d.%m.%Y")
                else:
                    date_to = ""
                self.request.session["invoice_filter"]["date_from"] = date_from
                self.request.session["invoice_filter"]["date_to"] = date_to
        else:
            form = InvoiceFilterForm(initial=initial)
        return form

    def debtor_list(self):
        from django_tables2 import RequestConfig

        from .tables import InvoiceOverviewTable

        if "consolidate" in self.request.POST:
            consolidate_invoices()
        data = invoice_overview(self.request.session["invoice_filter"])

        # Create django_tables2 table with sorting and pagination
        table = InvoiceOverviewTable(data)
        RequestConfig(self.request, paginate={"per_page": 50}).configure(table)

        return {"title": "Debitoren", "table": table, "invoice_table": True}

    def debtor_detail(self, key_type, key):
        from django_tables2 import RequestConfig

        from .tables import InvoiceDetailTable

        if key_type == "c":
            obj = Contract.objects.get(pk=key)
        else:
            obj = Address.objects.get(pk=key)
        if "consolidate" in self.request.POST:
            consolidate_invoices(obj)
        data = invoice_detail(obj, self.request.session["invoice_filter"])

        # Create django_tables2 table with sorting and pagination
        table = InvoiceDetailTable(data)
        RequestConfig(self.request, paginate={"per_page": 50}).configure(table)

        return {
            "title": "Detailansicht: %s" % (obj),
            "table": table,
            "invoice_table": True,
            "breadcrumbs": [{"name": "Debitoren", "href": "/geno/debtor/"}],
        }


## TODO: Refactor to ClassBased view
@login_required
def check_payments(request):
    if not request.user.has_perm("geno.canview_billing"):
        return unauthorized(request)

    ret = []

    now = datetime.datetime.now()
    members = Member.objects.exclude(date_leave__isnull=False)
    for member in members:
        warn = []
        ## Check if member has at least one share:
        if (
            get_active_shares()
            .filter(name=member.name)
            .filter(share_type=ShareType.objects.get(name="Anteilschein"))
            .count()
            < 1
        ):
            warn.append("Mitglied hat keine Anteilscheine (Beitritt: %s)." % member.date_join)
        ## Check if entry fee is paid:
        if member.date_join > datetime.date(2015, 1, 1):
            if (
                MemberAttribute.objects.filter(member=member)
                .filter(Q(value="Bezahlt (mit Eintritt)") | Q(value="Erlass (mit Eintritt)"))
                .count()
                < 1
            ):
                warn.append(
                    "Mitglied hat evtl. Eintrittsgebühr noch nicht bezahlt (Beitritt: %s)."
                    % member.date_join
                )

        ## Check membership fees
        for year in range(geno_settings.CHECK_MEMBERFEE_STARTYEAR, now.year):
            try:
                attribute = MemberAttributeType.objects.get(name="Mitgliederbeitrag %d" % year)
            except MemberAttributeType.DoesNotExist:
                continue
            if member.date_join < datetime.date(year, 1, 1):
                paid = (
                    MemberAttribute.objects.filter(member=member)
                    .filter(attribute_type=attribute)
                    .count()
                )
                if paid < 1:
                    warn.append("Mitgliedergebühr %d noch offen." % year)
                elif paid > 1:
                    warn.append("Mitgliedergebühr %d %d-fach bezahlt!" % (year, paid))
        if warn:
            doc = (
                Document.objects.filter(object_id=member.pk)
                .filter(content_type=ContentType.objects.get(app_label="geno", model="member"))
                .order_by("-ts_created")[0:1]
            )
            if doc:
                warn.append("Neustes Dokument: %s" % doc[0])
            ret.append({"info": str(member), "objects": warn})

    return render(request, "geno/default.html", {"response": ret, "title": "Check Zahlungen"})


class DocumentGeneratorView(CohivaAdminViewMixin, TemplateView):
    doctype = None

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.error_message = ""
        self.result = []

    def should_generate(self):
        """Override to control when documents should be generated. Default: always generate."""
        return True

    def get_objects(self):
        return []

    def get_options(self):
        return {"beschreibung": "Dokumente"}

    def get(self, request, *args, **kwargs):
        if not self.doctype:
            self.error_message = "Dokumententyp fehlt!"
        elif self.should_generate():
            options = self.get_options()
            options["makezip"] = request.GET.get("makezip", "") == "yes"
            if not options.get("link_url", None) and url_has_allowed_host_and_scheme(
                request.path, allowed_hosts=None
            ):
                options["link_url"] = request.path
            self.result = create_documents(self.doctype, self.get_objects(), options)
            if isinstance(self.result, HttpResponse):
                return self.result
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.error_message:
            context["response"] = [{"info": self.error_message}]
        else:
            context["response"] = self.result
        context["title"] = "Dokumente erzeugen - %s" % self.get_options().get("beschreibung", "")
        return context


class DryRunActionView(CohivaAdminViewMixin, TemplateView):
    """
    Generic base class for views that follow the dry-run/execute pattern.

    Workflow:
    1. Initial load or dry_run=True: Show preview/results without executing
    2. User clicks execute: dry_run=False, action is performed

    Subclasses should override:
    - process_action(dry_run): Perform the action and return results list
    - get_action_params(): Return dict of parameters for building execute URL
    - get_item_count(): Return count of items for display in footer (optional)
    """

    template_name = "geno/dry_run_action.html"  # Default template

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.result = None

    def is_dry_run(self):
        """Check if this is a dry run (preview mode)."""
        return self.request.GET.get("dry_run") != "False"

    def process_action(self, dry_run):
        """
        Override this to perform the action.
        Return a list of result dicts with 'info' and optional 'objects' keys.
        """
        return []

    def get_action_params(self):
        """
        Override this to return URL parameters for the execute button.
        Should return a dict of parameter names and values.
        """
        return {}

    def get_item_count(self):
        """
        Override this to return the number of items that will be processed.
        Used for display in the footer (e.g., "X Rechnungen werden erstellt").
        Returns None by default (no count displayed).
        """
        return None

    def get_item_label(self):
        """
        Override this to return the label for items being processed.
        Used in footer message: "Es werden X {item_label} verarbeitet."
        Returns None by default (generic message used).
        """
        return None

    def build_execute_url(self):
        """Build the URL for executing the action (dry_run=False)."""
        params = self.get_action_params()
        params["dry_run"] = "False"

        query_parts = []
        for key, value in params.items():
            if isinstance(value, list):
                for item in value:
                    query_parts.append(f"{key}={item}")
            else:
                query_parts.append(f"{key}={value}")

        return "?" + "&".join(query_parts) if query_parts else ""

    def get(self, request, *args, **kwargs):
        dry_run = self.is_dry_run()
        self.result = self.process_action(dry_run)

        if isinstance(self.result, (FileResponse, HttpResponse)):
            return self.result

        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["response"] = self.result

        if self.is_dry_run():
            context["execute_url"] = self.build_execute_url()

        item_count = self.get_item_count()
        if item_count is not None:
            context["item_count"] = int(item_count)

        item_label = self.get_item_label()
        if item_label is not None:
            context["item_label"] = item_label

        return context


class MemberLetterView(DocumentGeneratorView):
    permission_required = ("geno.add_document", "geno.send_newmembers")
    template_name = "geno/member_letter.html"

    def get_objects(self):
        objects = []
        ## Find members with missing documents (after cutoff date 2020-01-01)
        for m in Member.objects.filter(
            date_join__gt=settings.GENO_MEMBER_LETTER_CUTOFF_DATE
        ).exclude(date_leave__isnull=False):
            doc = (
                Document.objects.filter(object_id=m.pk)
                .filter(content_type=ContentType.objects.get(app_label="geno", model="member"))
                .filter(doctype__name=self.doctype)
            )
            if doc.count() == 0:
                objects.append(
                    {
                        "obj": m,
                        "doctype": self.doctype,
                        "info": "%s, Beitritt %s" % (m, m.date_join.strftime("%d.%m.%Y")),
                    }
                )
        return objects

    def get_options(self):
        return {
            "beschreibung": "Mitgliederbriefe",
            "link_url": "/geno/member/confirm/%s" % self.doctype,
        }

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Build table data from response
        if self.result and not self.error_message:
            # Filter out the "Dokumente:" item
            member_items = [item for item in self.result if item.get("info") != "Dokumente:"]

            if member_items:
                headers = [_("Mitglied"), _("Beitrittsdatum")]
                rows = []

                for item in member_items:
                    # Extract member name and join date from info string
                    # Format is: "Name, Beitritt DD.MM.YYYY"
                    details = (
                        ", ".join(str(o) for o in item.get("objects", []))
                        if item.get("objects")
                        else "—"
                    )
                    rows.append(
                        [
                            str(item.get("info", "")),
                            details,
                        ]
                    )

                context["table_data"] = {
                    "headers": headers,
                    "rows": rows,
                }
                context["member_count"] = len(member_items)

        return context


class ShareConfirmationLetterView(DocumentGeneratorView):
    permission_required = (
        "geno.add_document",
        "geno.confirm_share",
        "geno.canview_share",
        "geno.canview_billing",
    )
    doctype = "shareconfirm_req"
    template_name = "geno/share_confirm.html"

    def get_objects(self):
        # Find shares without documents (ignore single AS)
        stype_share = ShareType.objects.get(name="Anteilschein")
        try:
            stype_hypo = ShareType.objects.get(name="Hypothek")
        except ShareType.DoesNotExist:
            stype_hypo = None
        objects = []
        for s in (
            get_active_shares(interest=False)
            .filter(date__gt=settings.GENO_SHARE_LETTER_CUTOFF_DATE)
            .exclude(share_type=stype_hypo)
            .order_by("-date")
        ):
            obj_data = {"obj": s, "info": "%s %dx %s" % (s.date, s.quantity, s.value)}
            if s.share_type == stype_share:
                obj_data["doctype"] = "shareconfirm"
                obj_data["info"] = "%s [Best. Anteilschein]" % obj_data["info"]
            doc = (
                Document.objects.filter(object_id=s.pk)
                .filter(content_type=ContentType.objects.get(app_label="geno", model="share"))
                .filter(doctype__name__startswith="shareconfirm")
            )
            if doc.count() == 0:
                objects.append(obj_data)
        return objects

    def get_options(self):
        return {
            "beschreibung": "Bestätigungen Einzahlung Beteiligung",
            "link_url": "/geno/share/confirm",
        }

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Build table data from response
        if self.result and not self.error_message:
            # Filter out the "Dokumente:" item
            share_items = [item for item in self.result if item.get("info") != "Dokumente:"]

            if share_items:
                headers = [_("Name/Adresse"), _("Details")]
                rows = []

                for item in share_items:
                    details = (
                        ", ".join(str(o) for o in item.get("objects", []))
                        if item.get("objects")
                        else "—"
                    )
                    rows.append(
                        [
                            str(item.get("info", "")),
                            details,
                        ]
                    )

                context["table_data"] = {
                    "headers": headers,
                    "rows": rows,
                }
                context["share_count"] = len(share_items)

        return context


class ShareReminderLetterView(DocumentGeneratorView):
    permission_required = (
        "geno.add_document",
        "geno.confirm_share",
        "geno.canview_share",
        "geno.canview_billing",
    )
    doctype = "loanreminder"
    template_name = "geno/share_duedate_reminder.html"

    def get_objects(self):
        cutoff_date = timezone.now() + relativedelta(months=16)
        today = datetime.date.today()
        next_year = today.year + 1

        # ret = []
        objects = []
        for adr in Address.objects.filter(active=True):
            ## Check if we have a recent reminder document already
            try:
                last_reminder = (
                    Document.objects.filter(object_id=adr.pk)
                    .filter(
                        content_type=ContentType.objects.get(app_label="geno", model="address")
                    )
                    .filter(doctype__name=self.doctype)
                    .latest("ts_created")
                )
                last_reminder_cutoff_date = last_reminder.ts_created + relativedelta(months=16)
            except Document.DoesNotExist:
                last_reminder = None
                last_reminder_cutoff_date = None

            share_contexts = []
            info = []
            ## Get active loans that have no end date
            for share in (
                get_active_shares()
                .filter(name=adr)
                .filter(share_type__name__startswith="Darlehen")
                .filter(date_end=None)
                .filter(is_interest_credit=False)
            ):
                startdate = share.date
                if share.date_due:
                    duedate = share.date_due
                    if share.duration:
                        startdate = share.date_due - relativedelta(years=share.duration)
                elif share.duration:
                    duedate = share.date + relativedelta(years=share.duration)
                else:
                    duedate = None
                    info.append("WARNUNG: %s hat KEIN FÄLLIGKEITSDATUM!" % (share))
                if duedate and duedate < cutoff_date.date():
                    if (
                        not last_reminder_cutoff_date
                        or duedate >= last_reminder_cutoff_date.date()
                    ):
                        info.append(
                            "%s[%s]: NEEDS REMINDER"
                            % (nformat(share.quantity * share.value), duedate)
                        )
                        share_context = {
                            "zaehler": "",
                            "betrag": nformat(share.quantity * share.value),
                        }
                        if share.duration:
                            share_context["laufzeit"] = "%s Jahre - " % share.duration
                        else:
                            share_context["laufzeit"] = ""
                        share_context["laufzeit"] += "%s – %s" % (
                            startdate.strftime("%d.%m.%Y"),
                            duedate.strftime("%d.%m.%Y"),
                        )
                        share_context["zinssatz"] = nformat(share.interest())
                        share_context["plus5jahre"] = (duedate + relativedelta(years=5)).strftime(
                            "%d.%m.%Y"
                        )
                        share_context["plus10jahre"] = (
                            duedate + relativedelta(years=10)
                        ).strftime("%d.%m.%Y")
                        share_context["datum_zins_neu"] = "01.01.%s" % next_year
                        share_contexts.append(share_context)
                    else:
                        info.append(
                            "%s[%s]: Already reminded (%s)"
                            % (
                                nformat(share.quantity * share.value),
                                duedate,
                                last_reminder.ts_created,
                            )
                        )
                else:
                    info.append(
                        "%s[%s]: Not due" % (nformat(share.quantity * share.value), duedate)
                    )

            if len(share_contexts) > 1:
                counter = 1
                for sc in share_contexts:
                    sc["zaehler"] = "(Darlehen %s von %s)" % (counter, len(share_contexts))
                    counter += 1

            if share_contexts:
                # ret.append({'info': '%s' % (adr), 'objects': objects})
                objects.append(
                    {
                        "obj": adr,
                        "info": "%s Darlehen: %s" % (len(share_contexts), " / ".join(info)),
                        "extra_context": {"darlehen": share_contexts},
                    }
                )
        return objects

    def get_options(self):
        return {
            "beschreibung": "Brief Erinnerung Darlehen",
            "link_url": "/geno/share/duedate_reminder",
        }

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Build table data from response
        if self.result and not self.error_message:
            # Filter out the "Dokumente:" item
            reminder_items = [item for item in self.result if item.get("info") != "Dokumente:"]

            if reminder_items:
                headers = [_("Name/Adresse"), _("Darlehen Details")]
                rows = []

                for item in reminder_items:
                    details = (
                        ", ".join(str(o) for o in item.get("objects", []))
                        if item.get("objects")
                        else "—"
                    )
                    rows.append(
                        [
                            str(item.get("info", "")),
                            details,
                        ]
                    )

                context["table_data"] = {
                    "headers": headers,
                    "rows": rows,
                }
                context["reminder_count"] = len(reminder_items)

        return context


class ShareInterestView(CohivaAdminViewMixin, TemplateView):
    title = _("Zinsabrechung")
    permission_required = (
        "geno.canview_share",
        "geno.canview_billing",
        "geno.share_interest_statements",
    )
    template_name = "geno/share_interest.html"
    actions = []  # No toolbar actions - booking in content, download in footer

    def get(self, request, *args, **kwargs):
        if self.action == "download":
            return self.download()
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["action"] = self.action
        context["can_create_transactions"] = self.request.user.has_perm("geno.transaction")
        if self.action == "create_transactions" and self.request.user.has_perm("geno.transaction"):
            response = create_interest_transactions()
            # Add variants based on message content
            for item in response:
                info = item.get("info", "")
                if info.startswith("FEHLER") or " FEHLER" in info:
                    item["variant"] = ResponseVariant.ERROR.value
                elif info.startswith("WARNUNG"):
                    item["variant"] = ResponseVariant.WARNING.value
                elif "GESPEICHERT" in info:
                    item["variant"] = ResponseVariant.SUCCESS.value
                else:
                    item["variant"] = ResponseVariant.DEFAULT.value
            context["response"] = response
        return context

    def download(self):
        year_current = datetime.datetime.now().year
        year = year_current - 1

        if self.request.GET.get("darlehen", "") == "yes":
            opt_darlehen = True
            output_tag = "Darlehenszins"
        else:
            opt_darlehen = False
            output_tag = "Zinsabrechnung"

        ## Spreadsheet
        import openpyxl
        from openpyxl.styles import Font

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=%s_%s_%d.xlsx" % (
            settings.GENO_FILENAME_STR,
            output_tag,
            year,
        )
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "%s %s %s" % (settings.GENO_FILENAME_STR, output_tag, year)
        ## Header
        columns = [
            ("A", "Name", 30),
            ("B", "Typ", 20),
            ("C", "Datum", 25),
            ("D", "Saldo CHF", 12),
            ("E", "Tage", 8),
            ("F", "Satz", 8),
            ("G", "Bruttozins", 12),
            ("H", "Zins Total", 12),
            ("I", "Steuerfrei", 12),
            ("J", "VSt. 35%", 12),
            ("K", "Gutschrift", 12),
        ]
        row_num = 0
        for col_num in range(len(columns)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = columns[col_num][1]
            c.font = Font(bold=True)
            # set column width
            ws.column_dimensions[columns[col_num][0]].width = columns[col_num][2]
        row_num += 2

        sum_interest = 0
        sum_interest_notax = 0
        sum_interest_tax = 0
        sum_interest_pay = 0
        count = 0
        for adr in Address.objects.filter(active=True).order_by("name"):
            try:
                interest = share_interest_calc(adr, year)
            except Exception as e:
                messages.error(self.request, "FEHLER bei der Zinsberechnung: %s" % str(e))
                return HttpResponseRedirect(reverse("geno:share-interest"))
            darlehen_spezial = ""
            if opt_darlehen:
                ## Only sum of Darlehen
                total = interest["total"][2] + interest["total"][4]
                tax = interest["tax"][2] + interest["tax"][4]
                pay = interest["pay"][2] + interest["pay"][4]
                if interest["total"][4] > 0:
                    darlehen_spezial = "/SPEZIAL"
            else:
                ## All
                total = interest["total_alltypes"]
                tax = interest["tax_alltypes"]
                pay = interest["pay_alltypes"]

            if total > 0:
                if total == pay:
                    notax = pay
                else:
                    notax = 0
                row = [
                    str(adr),
                    "TOTAL%s" % darlehen_spezial,
                    "",
                    "",
                    "",
                    "",
                    "",
                    total,
                    notax,
                    tax,
                    pay,
                ]
                for col_num in range(len(row)):
                    c = ws.cell(row=row_num + 1, column=col_num + 1)
                    c.value = row[col_num]
                    if not opt_darlehen:
                        c.font = Font(bold=True)
                row_num += 1
                sum_interest += total
                sum_interest_notax += notax
                sum_interest_tax += tax
                sum_interest_pay += pay
                count += 1

            if total > 0 and not opt_darlehen:
                for date_list in interest["dates"]:
                    for d in date_list:
                        row = [
                            "",
                            str(d["type"]),
                            "%s - %s"
                            % (d["start"].strftime("%d.%m.%Y"), d["end"].strftime("%d.%m.%Y")),
                            d["amount"],
                            d["days"],
                            d["interest_rate"],
                            d["interest"],
                        ]
                        for col_num in range(len(row)):
                            c = ws.cell(row=row_num + 1, column=col_num + 1)
                            c.value = row[col_num]
                            # c.font = Font(bold = True)
                        row_num += 1
                row_num += 1

        ## Sum
        row_num += 1
        row = [
            "Anzahl/Summe:",
            count,
            "",
            "",
            "",
            "",
            "",
            sum_interest,
            sum_interest_notax,
            sum_interest_tax,
            sum_interest_pay,
        ]
        for col_num in range(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            c.font = Font(bold=True)
        row_num += 1

        wb.save(response)
        return response


@login_required
def share_mailing(request):
    if not request.user.has_perm("geno.canview_share") or not request.user.has_perm(
        "geno.canview_billing"
    ):
        return unauthorized(request)

    ret = []
    objects = []
    for adr in Address.objects.filter(active=True).order_by("name"):
        if not is_member(adr, date_mode="end_date"):
            ## Nichtmitglied
            # ret.append({'info': str(adr), 'objects': ['Ohne Mitgliedschaft']})
            continue

        if not adr.city:
            ret.append({"info": str(adr), "objects": ["KEINE ADRESSE!"]})
            continue

        ## Flags
        share_mini = False
        loan = False
        deposit = False

        obj = []
        try:
            inter = share_interest_calc(adr, datetime.datetime.now().year)
        except Exception as e:
            messages.error(request, "FEHLER bei der Zinsberechnung: %s" % str(e))
            return HttpResponseRedirect("/admin/")
        deposit_betrag = ""
        if inter["end_amount"][0] > 1000:
            ## Anteilscheine > 5
            obj.append("AS: %d" % inter["end_amount"][0])
        else:
            ## Anteilscheine < 6
            obj.append("AS: %d [mini]" % inter["end_amount"][0])
            share_mini = True
        if inter["end_amount"][1]:
            ## Zinslose Darlehen
            obj.append("Darlehen zinslos: %d" % inter["end_amount"][1])
            loan = True
        if inter["end_amount"][2]:
            ## Verzinste Darlehen
            obj.append("Darlehen verzinst: %d" % inter["end_amount"][2])
            loan = True
        if inter["end_amount"][3]:
            ## Depositenkasse
            obj.append("Depositenkasse: %d" % inter["end_amount"][3])
            deposit_betrag = nformat(inter["end_amount"][3])
            deposit = True
        if inter["end_amount"][4]:
            ## Darlehen spezial
            obj.append("Darlehen spezial: %d" % inter["end_amount"][4])
            loan = True

        loan_info = []
        # loan_values = []
        # loan_due = []
        # loan_duenew = []
        if loan:
            letter = "mitDarlehen"
            for d in (
                get_active_shares()
                .filter(name=adr)
                .filter(
                    Q(share_type=ShareType.objects.get(name="Darlehen zinslos"))
                    | Q(share_type=ShareType.objects.get(name="Darlehen verzinst"))
                )
                .filter(is_interest_credit=False)
            ):
                # loan_values.append(nformat(d.value))
                duedate = d.date + relativedelta(years=d.duration)
                duedate_new = d.date + relativedelta(years=(d.duration + 5))
                # loan_due.append(duedate.strftime("%d.%m.%Y"))
                # loan_duenew.append(duedate_new.strftime("%d.%m.%Y"))
                loan_info.append(
                    {
                        "betrag": nformat(d.value),
                        "due": duedate.strftime("%d.%m.%Y"),
                        "duenew": duedate_new.strftime("%d.%m.%Y"),
                    }
                )
                obj.append(
                    "Darlehen-INFO: CHF %s,  %s  -> %s"
                    % (
                        nformat(d.value),
                        duedate.strftime("%d.%m.%Y"),
                        duedate_new.strftime("%d.%m.%Y"),
                    )
                )
        elif deposit:
            letter = "mitDeposito"
        elif share_mini:
            letter = "ohneBeitrag"
        else:
            letter = "mitAnteilsscheinen"
        obj.append("-> %s" % letter)

        ret.append({"info": str(adr), "objects": obj})
        objects.append(
            {
                "obj": adr,
                "info": "%s" % (letter),
                "extra_context": {
                    "loan_info": loan_info,
                    "deposit_betrag": deposit_betrag,
                    "filename_tag": letter,
                },
            }
        )

    options = {
        "beschreibung": "Mailings",
        "link_url": "/geno/share/mailing",
    }
    return create_documents_deprecated(request, "mailing", objects, options)


class ShareStatementView(DocumentGeneratorView):
    template_name = "geno/share_statement.html"
    permission_required = (
        "geno.canview_share",
        "geno.canview_billing",
        "geno.share_interest_statements",
    )
    doctype = "statement"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.enddate = None
        self.address_id = None
        self.extra_description_info = ""

    def setup(self, request, *args, **kwargs):
        super().setup(request, *args, **kwargs)
        self.enddate = self.get_enddate(self.kwargs.get("date", "previous_year"))
        self.address_id = self.kwargs.get("address", None)

    def get_options(self):
        options = {
            "link_url": "/geno/share/statement/%s" % (self.enddate.strftime("%Y-%m-%d")),
            "beschreibung": f"Kontoauszüge {self.enddate.year}{self.extra_description_info}",
        }
        if self.address_id:
            options["link_url"] = "%s/%s" % (options["link_url"], self.address_id)
        return options

    @staticmethod
    def get_enddate(date):
        year_current = datetime.datetime.now().year
        if date == "previous_year":
            return datetime.date(year_current - 1, 12, 31)
        if date == "current_year":
            return datetime.date(year_current, 12, 31)
        try:
            return datetime.datetime.strptime(date, "%Y-%m-%d").date()
        except ValueError:
            return None

    def get_objects(self, date="previous_year"):
        objects = []
        checksum_interest_pay = 0
        checksum_interest_tax = 0
        checksum_count = 0
        skip_count = 0
        if self.address_id:
            addresses = Address.objects.filter(pk=self.address_id)
        else:
            addresses = Address.objects
        for adr in addresses.filter(active=True).order_by("name"):
            try:
                statement_data = get_share_statement_data(adr, self.enddate.year, self.enddate)
            except Exception as e:
                self.error_message = "FEHLER beim Erstellen des Kontoauszugs: %s" % str(e)
                return objects

            if statement_data["sect_interest"]:
                checksum_interest_tax += statement_data["s_tax"]
                checksum_interest_pay += statement_data["s_pay"]
            statement_data["s_tax"] = nformat(statement_data["s_tax"])
            statement_data["s_pay"] = nformat(statement_data["s_pay"])

            info = (
                "%s/%s, Darlehen(zinsl.): %s, Darlehen: %s, Depositen: %s/%s, "
                "VSt: %s, ZinsAuszahlung: %s, Zeilen: %d, %s + %s"
                % (
                    statement_data["n_shares"],
                    statement_data["s_shares"],
                    statement_data["s_loan_no"],
                    statement_data["s_loan"],
                    statement_data["dep_start"],
                    statement_data["dep_end"],
                    statement_data["s_tax"],
                    statement_data["s_pay"],
                    statement_data["line_count"],
                    statement_data["loan_no_duedates"],
                    statement_data["loan_duedates"],
                )
            )
            if (
                statement_data["sect_shares"]
                or statement_data["sect_loan"]
                or statement_data["sect_deposit"]
            ):
                if not statement_data["thankyou"]:
                    ## Skip if max. GENO_SMALL_NUMBER_OF_SHARES_CUTOFF shares and no loan etc.
                    skip_count += 1
                else:
                    checksum_count += 1
                    objects.append(
                        {
                            "obj": adr,
                            "info": "%d: %s" % (checksum_count, info),
                            "extra_context": statement_data,
                        }
                    )
            self.extra_description_info = (
                " [Anzahl=%d, VSt=%s, ZinsAuszahlung=%s, Anzahl ignoriert=%d]"
                % (
                    checksum_count,
                    nformat(checksum_interest_tax),
                    nformat(checksum_interest_pay),
                    skip_count,
                )
            )
        return objects

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Clean up the title - remove technical details
        context["title"] = f"Kontoauszüge {self.enddate.year}"

        # Get the raw objects with extra_context before they're processed by create_documents
        objects = self.get_objects()

        if objects and not self.error_message:
            headers = [
                _("Mitglied"),
                _("Anteilscheine"),
                _("Darlehen"),
                _("Depositen"),
                _("Verrechnungssteuer"),
                _("Zinszahlung"),
            ]

            rows = []
            total_tax = 0
            total_interest = 0

            for item in objects:
                if "extra_context" in item:
                    statement_data = item["extra_context"]
                    obj = item["obj"]

                    # Parse the formatted values back to numbers for totals
                    try:
                        tax_value = float(
                            statement_data["s_tax"].replace("'", "").replace(",", ".")
                        )
                        total_tax += tax_value
                    except (ValueError, AttributeError):
                        tax_value = 0

                    try:
                        interest_value = float(
                            statement_data["s_pay"].replace("'", "").replace(",", ".")
                        )
                        total_interest += interest_value
                    except (ValueError, AttributeError):
                        interest_value = 0

                    rows.append(
                        [
                            str(obj),
                            f"{statement_data['n_shares']} ({statement_data['s_shares']})",
                            statement_data["s_loan"],
                            f"{statement_data['dep_start']}/{statement_data['dep_end']}",
                            statement_data["s_tax"],
                            statement_data["s_pay"],
                        ]
                    )

            context["table_data"] = {
                "headers": headers,
                "rows": rows,
            }
            context["statement_count"] = len(rows)
            context["total_tax"] = nformat(total_tax)
            context["total_interest"] = nformat(total_interest)

        return context


class ShareStatementFormView(CohivaAdminViewMixin, FormView):
    form_class = ShareStatementForm
    template_name = "geno/share_statement_form.html"
    title = "Kontoauszüge"
    permission_required = (
        "geno.canview_share",
        "geno.canview_billing",
        "geno.share_interest_statements",
    )

    def get_initial(self):
        initial = super().get_initial()
        initial["date"] = datetime.date(datetime.datetime.now().year - 1, 12, 31)
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["attrs_button"] = {"form": "statement-form"}
        if getattr(settings, "GENO_SMALL_NUMBER_OF_SHARES_CUTOFF", 0) > 0:
            context["help_text"] = (
                "Es werden nur Kontoauszüge erstellt, falls mehr als "
                f"{settings.GENO_SMALL_NUMBER_OF_SHARES_CUTOFF} Anteilsscheine "
                "oder andere Beteiligungen wie Darlehen/Depositen vorhanden sind."
            )
        return context

    def form_valid(self, form):
        self.success_url = reverse(
            "geno:share-statement", args=[form.cleaned_data.get("date").strftime("%Y-%m-%d")]
        )
        return super().form_valid(form)


@login_required
def contract_report(request, year="previous", address=None):
    if not request.user.has_perm("geno.canview_share") or not request.user.has_perm(
        "geno.canview_billing"
    ):
        return unauthorized(request)

    report = check_rental_shares_report()
    header = [
        "Mietobjekt/Name",
        "AS Soll",
        "AS Reduktion",
        "AS benötigt",
        "AS bezahlt",
        "AS ausstehend",
        "AS Leerstand",
        "Min. Beleg.",
        "Beleg.",
        "Diff.",
        "Kinder",
        "Details",
    ]
    return export_data_to_xls(report, header=header)


class ContractCheckFormsView(DocumentGeneratorView):
    title = "Formulare «Überprüfung Belegung/Fahrzeuge»"
    permission_required = ("geno.canview_share", "geno.rental_contracts")
    template_name = "geno/contract_check_forms.html"
    doctype = "contract_check"

    def should_generate(self):
        """Only generate documents when explicitly requested."""
        return self.request.GET.get("generate", "") == "yes"

    def get_objects(self):
        """Build the list of contracts that need check forms."""
        objects = []
        for c in get_active_contracts():
            for ru in c.rental_units.all():
                if ru.rental_type not in ("Gewerbe", "Lager", "Hobby", "Parkplatz"):
                    objects.append({"obj": c})
                    break  # Only add contract once
        return objects

    def get_options(self):
        """Return document generation options."""
        options = {"beschreibung": "Formulare «Überprüfung Belegung/Fahrzeuge»"}
        if url_has_allowed_host_and_scheme(self.request.path, allowed_hosts=None):
            options["link_url"] = self.request.path
        return options

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Always calculate the list of contracts for count
        contracts = []
        for c in get_active_contracts():
            for ru in c.rental_units.all():
                if ru.rental_type not in ("Gewerbe", "Lager", "Hobby", "Parkplatz"):
                    contracts.append(c)
                    break  # Only add contract once

        context["contract_count"] = len(contracts)

        # If not generating, show the preview table with contract overview
        if not self.should_generate():
            # Prepare table data
            headers = ["Vertrag", "Mieter", "Mietobjekte"]
            rows = []

            for contract in contracts:
                # Get contractors list
                contractors = []
                for c in contract.contractors.all():
                    if c == contract.main_contact:
                        contractors.insert(0, str(c))
                    else:
                        contractors.append(str(c))
                contractors_str = " / ".join(contractors) if contractors else "-"

                rental_units = ", ".join(
                    str(ru)
                    for ru in contract.rental_units.all()
                    if ru.rental_type not in ("Gewerbe", "Lager", "Hobby", "Parkplatz")
                )
                rows.append(
                    [
                        str(contract.get_contract_label() or f"Vertrag {contract.pk}"),
                        contractors_str,
                        rental_units,
                    ]
                )

            context["table_data"] = {
                "headers": headers,
                "rows": rows,
            }

        return context


## TODO: Refactor to ClassBased view
@login_required
def create_contracts(request, letter=False):
    if not request.user.has_perm("geno.canview_share") or not request.user.has_perm(
        "geno.canview_billing"
    ):
        return unauthorized(request)

    stock = {}
    stock["EG"] = "Erdgeschoss"
    stock["1.OG"] = "1. Obergeschoss"
    stock["2.OG"] = "2. Obergeschoss"
    stock["3.OG"] = "3. Obergeschoss"
    stock["4.OG"] = "4. Obergeschoss"
    stock["5.OG"] = "5. Obergeschoss"
    stock["6.OG"] = "6. Obergeschoss"
    stock["7.OG"] = "7. Obergeschoss"
    stock["8.OG"] = "8. Obergeschoss"
    stock["9.OG"] = "9. Obergeschoss"
    stock["1.UG"] = "1. Untergeschoss"
    stock["2.UG"] = "2. Untergeschoss"
    stock["3.UG"] = "3. Untergeschoss"
    stock["4.UG"] = "4. Untergeschoss"
    stock["5.UG"] = "5. Untergeschoss"
    stock["6.UG"] = "6. Untergeschoss"
    stock["7.UG"] = "7. Untergeschoss"
    stock["8.UG"] = "8. Untergeschoss"
    stock["9.UG"] = "9. Untergeschoss"

    objects = []
    for contract in Contract.objects.all():
        data = {"mietobjekt": ""}
        for rental_unit in contract.rental_units.filter(rental_type="Wohnung"):
            if data["mietobjekt"] != "":
                raise Exception("Currently only one rental unit is supported!")
            floor_str = stock.get(rental_unit.floor, rental_unit.floor)
            if floor_str:
                data["mietobjekt"] = "%s-Zimmer-Wohnung Nr. %s im %s" % (
                    rental_unit.rooms,
                    rental_unit.name,
                    floor_str,
                )
            else:
                data["mietobjekt"] = "%s-Zimmer-Wohnung Nr. %s" % (
                    rental_unit.rooms,
                    rental_unit.name,
                )
            data["filename_tag"] = "Wohnung_%s" % rental_unit.name
            data["miete_netto"] = nformat(rental_unit.rent_netto, 0)
            data["miete_brutto"] = nformat(
                (rental_unit.rent_netto if rental_unit.rent_netto else 0.0)
                + (rental_unit.nk if rental_unit.nk else 0.0)
                + (rental_unit.nk_flat if rental_unit.nk_flat else 0.0)
                + (rental_unit.nk_electricity if rental_unit.nk_electricity else 0.0),
                0,
            )
            data["nk_akonto"] = nformat(rental_unit.nk, 0)
            data["nk_pauschal"] = nformat(rental_unit.nk_flat, 0)
            data["nk_strom"] = nformat(rental_unit.nk_electricity, 0)
            data["mindestbelegung"] = nformat(rental_unit.min_occupancy, 0)
            data["pflichtanteil"] = nformat(rental_unit.share, 0)
            data["rate1"] = nformat(rental_unit.share / 2, 0)
            data["rate2"] = nformat(rental_unit.share / 2, 0)
        if data["mietobjekt"] == "":
            continue
        for rental_unit in contract.rental_units.filter(rental_type="Kellerabteil"):
            data["mietobjekt"] = (
                "%s, Kellerabteil Nr. %s (ungeheizt, kein kontrolliertes Raumklima)"
                % (data["mietobjekt"], rental_unit.name)
            )
        info = "%s" % (data["mietobjekt"])
        mieter = []
        mieter_txt = []
        for m in contract.contractors.all():
            mieter.append(
                {
                    "name": "%s %s" % (m.first_name, m.name),
                    "address": "%s, %s" % (m.street, m.city),
                }
            )
            mieter_txt.append(
                "%s %s\tBisherige Adresse: %s, %s" % (m.first_name, m.name, m.street, m.city)
            )
        data["mieter_txt"] = "\n".join(mieter_txt)
        data["mieter"] = mieter
        if len(mieter) == 1:
            data["MieterInnen"] = "Mieter*in"
        else:
            data["MieterInnen"] = "Mieter*innen"

        objects.append({"obj": contract, "info": "%s" % (info), "extra_context": data})

    if letter:
        options = {
            "beschreibung": "Mietverträge Begleitbrief",
            "link_url": "/geno/contract/create_letter",
        }
        return create_documents_deprecated(request, "contract_letter", objects, options)
    else:
        options = {
            "beschreibung": "Mietverträge",
            "link_url": "/geno/contract/create",
        }
        return create_documents_deprecated(request, "contract", objects, options)


## Remove this after all views use DocumentGeneratorBaseView
@login_required
def create_documents_deprecated(request, default_doctype, objects=None, options=None):
    if request.GET.get("makezip", "") == "yes":
        makezip = True
    else:
        makezip = False

    ret = []
    zipcount = 0
    zipfile_content = []

    try:
        default_doctype_obj = DocumentType.objects.get(name=default_doctype)
    except DocumentType.DoesNotExist:
        return render(
            request,
            "geno/default.html",
            {
                "response": [
                    {
                        "info": 'FEHLER: Dokumenttyp "%s" existiert nicht. Bitte zuerst erstellen.'
                        % default_doctype
                    }
                ],
                "title": "Dokumente erzeugen",
            },
        )

    if not options:
        options = {"beschreibung": default_doctype_obj.description}

    if not objects:
        if default_doctype == "contract_check":
            objects = []
            for c in get_active_contracts():
                for ru in c.rental_units.all():
                    if ru.rental_type not in ("Gewerbe", "Lager", "Hobby", "Parkplatz"):
                        objects.append({"obj": c})
                        break
        else:
            return render(
                request,
                "geno/default.html",
                {
                    "response": [
                        {"info": 'Keine Objekte gefunden (Dokumenttyp "%s").' % default_doctype}
                    ],
                    "title": "Dokumente erzeugen - %s" % options["beschreibung"],
                },
            )

    filenames = []
    for o in objects:
        zipcount += 1
        objects = []
        if "info" in o:
            objects.append(o["info"])
        ret.append({"info": str(o["obj"]), "objects": objects})
        if makezip:
            if "doctype" in o:
                doctype = o["doctype"]
                doctype_obj = DocumentType.objects.get(name=doctype)
            else:
                doctype = default_doctype
                doctype_obj = default_doctype_obj
            ## Create document
            if "extra_context" in o:
                data = get_context_data(doctype, o["obj"].pk, o["extra_context"])
            else:
                data = get_context_data(doctype, o["obj"].pk, {})
            filename = fill_template_pod(
                doctype_obj.template.file.path,
                context_format(data["context"]),
                output_format="odt",
            )
            if not filename:
                raise RuntimeError("Could not fill template")
            output_filename = data["visible_filename"]
            counter = 1
            while output_filename in filenames:
                counter += 1
                output_filename = "%s_%s%s" % (
                    data["visible_filename"][0:-4],
                    counter,
                    data["visible_filename"][-4:],
                )
            zipfile_content.append({"file": filename, "filename": output_filename})
            filenames.append(output_filename)
            if "content_object" in data:
                ## Attach document data to object
                d = Document(
                    name=output_filename,
                    doctype=doctype_obj,
                    data=json.dumps(data["context"]),
                    content_object=data["content_object"],
                )
                d.save()

    if makezip:
        ## Build ZIP-file from list of files
        file_like_object = io.BytesIO()
        zipfile_ob = zipfile.ZipFile(file_like_object, "w")
        for f in zipfile_content:
            # print "%s -> %s_%s" % (f['file'], f['member'].name.name,f['member'].name.first_name)
            zipfile_ob.write(f["file"], f["filename"])
        zipfile_ob.close()
        resp = HttpResponse(
            file_like_object.getvalue(), content_type="application/x-zip-compressed"
        )
        resp["Content-Disposition"] = "attachment; filename=%s" % "output.zip"
        return resp

    if zipcount > 0:
        link_text = "%s erstellen und herunterladen" % options["beschreibung"]
        if url_has_allowed_host_and_scheme(request.path, allowed_hosts=None):
            link_url = options.get("link_url", request.path)
        else:
            link_url = options.get("link_url", "")
        ret.append(
            {
                "info": "Dokumente:",
                "objects": [
                    '<a href="%s?makezip=yes">%s (%d LibreOffice Dokumente in ZIP)</a>'
                    % (link_url, link_text, zipcount)
                ],
            }
        )
    else:
        ret.append({"info": "Keine zu erstellenden Dokumente gefunden."})

    return render(
        request,
        "geno/default.html",
        {"response": ret, "title": "Dokumente erzeugen - %s" % options["beschreibung"]},
    )


class CheckMailinglistsView(CohivaAdminViewMixin, TemplateView):
    title = "Mailverteiler überprüfen"
    permission_required = "geno.canview_member_mailinglists"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["response"] = self.check_mailinglists()
        return context

    def check_mailinglists(self):
        if not hasattr(settings, "MAILMAN_API") or not settings.MAILMAN_API.get("password", None):
            if not getattr(settings, "DEMO", False):
                return [
                    {
                        "info": "FEHLER: Mailman-API ist nicht konfiguriert.",
                        "variant": ResponseVariant.ERROR.value,
                    }
                ]

        use_demo = getattr(settings, "DEMO", False)

        if use_demo:
            ml_members = {
                "genossenschaft": ["alma@example.com", "berta@example.com", "hugo@example.com"],
                "bewohnende": ["zora@example.com"],
                "gewerbemietende": [],
                "wohnpost": ["alma@example.com"],
            }
            ml_warnings = []
            demo_members = [
                {"name": "Hugo Demo", "email": "hugo@example.com"},
                {"name": "Dora Demo", "email": "dora@example.com"},
                {"name": "Ohnemail, Peter", "email": ""},
                {"name": "Verein WG Kunterbunt, Hans Muster", "email": "hans.muster@example.com"},
                {
                    "name": "Verein WG Kunterbunt, Hans Muster (Duplikat)",
                    "email": "hans.muster@example.com",
                },
            ]
            demo_bewohnende = [
                {"name": "Zora Demo", "email": "zora@example.com"},
                {"name": "Analog, Heidi", "email": ""},
            ]
            cs_active = ["hugo@example.com"]
            cs_unsubscribed = []
            cs_bounced = []
            cs_deleted = []
        else:
            mailman_client = mailmanclient.Client(
                settings.MAILMAN_API["url"],
                settings.MAILMAN_API["user"],
                settings.MAILMAN_API["password"],
            )
            ml_warnings = []
            ml_members = {}
            for ml in ("genossenschaft", "bewohnende", "gewerbemietende", "wohnpost"):
                mlist = mailman_client.get_list(f"{ml}@{settings.MAILMAN_API['lists_domain']}")
                ml_members[ml] = []
                for member in mlist.members:
                    ml_members[ml].append(member.email)
                    if member.bounce_score:
                        ml_warnings.append(
                            f"[{ml}] {member.email} has bounce_score = {member.bounce_score}"
                        )
                    if member.delivery_mode != "regular":
                        ml_warnings.append(
                            f"[{ml}] {member.email} has delivery_mode = {member.delivery_mode}"
                        )
                    if member.role != "member":
                        ml_warnings.append(f"[{ml}] {member.email} has role = {member.role}")
                    if member.subscription_mode != "as_address":
                        ml_warnings.append(
                            f"[{ml}] {member.email} has subscription_mode = {member.subscription_mode}"
                        )

        ignore_emails = settings.GENO_CHECK_MAILINGLISTS["ignore_emails"]

        ret = []

        ## Process bewohnende
        bewohnende = []
        bewohnende_email = []
        bewohnende_missing = []
        bewohnende_no_email = []
        bewohnende_duplicate = []

        if use_demo:
            for person in demo_bewohnende:
                if not person["email"]:
                    bewohnende_no_email.append(person["name"])
                elif person["email"] in bewohnende_email:
                    bewohnende_duplicate.append(f"{person['name']} ({person['email']})")
                else:
                    bewohnende_email.append(person["email"])
                    if person["email"] in ml_members["bewohnende"]:
                        ml_members["bewohnende"].remove(person["email"])
                    else:
                        bewohnende_missing.append(person["email"])
        else:
            for c in get_active_contracts():
                include = False
                for ru in c.rental_units.all():
                    if ru.rental_type not in ("Gewerbe", "Lager", "Hobby", "Parkplatz"):
                        include = True
                        break
                if include:
                    for adr in c.contractors.all():
                        if adr not in bewohnende:
                            bewohnende.append(adr)
                    ## Add children that have an email address
                    for child in c.children.exclude(name__email__exact=""):
                        if child.name not in bewohnende:
                            bewohnende.append(child.name)
            for adr in bewohnende:
                if not adr.email:
                    bewohnende_no_email.append(str(adr))
                elif adr.email in bewohnende_email:
                    bewohnende_duplicate.append(f"{adr} ({adr.email})")
                else:
                    bewohnende_email.append(adr.email)
                    if adr.email in ml_members["bewohnende"]:
                        ml_members["bewohnende"].remove(adr.email)
                    else:
                        bewohnende_missing.append(adr.email)

        ## Process members (genossenschaft)
        members_email = []
        genossenschaft_missing = []
        genossenschaft_no_email = []
        genossenschaft_duplicate = []
        wohnpost_missing = []

        if use_demo:
            for person in demo_members:
                if not person["email"]:
                    genossenschaft_no_email.append(person["name"])
                elif person["email"] in members_email:
                    genossenschaft_duplicate.append(f"{person['name']} ({person['email']})")
                else:
                    members_email.append(person["email"])
                    if person["email"] in ml_members["genossenschaft"]:
                        ml_members["genossenschaft"].remove(person["email"])
                    else:
                        genossenschaft_missing.append(person["email"])
                    if (
                        person["email"] not in ml_members["wohnpost"]
                        and person["email"] not in ignore_emails
                    ):
                        wohnpost_missing.append(person["email"])
        else:
            for member in Member.objects.all():
                if not is_member(member.name):
                    continue
                if not member.name.email:
                    genossenschaft_no_email.append(str(member.name))
                elif member.name.email in members_email:
                    genossenschaft_duplicate.append(f"{member.name} ({member.name.email})")
                else:
                    members_email.append(member.name.email)
                    if member.name.email in ml_members["genossenschaft"]:
                        ml_members["genossenschaft"].remove(member.name.email)
                    else:
                        genossenschaft_missing.append(member.name.email)
                    if (
                        member.name.email not in ml_members["wohnpost"]
                        and member.name.email not in ignore_emails
                    ):
                        wohnpost_missing.append(member.name.email)

        ## Bewohnende / Gewerbemietende not in genossenschaft-ML or wohnpost
        wohnpost_and_geno_missing = []
        for email in ml_members["bewohnende"]:
            if (
                email not in ml_members["wohnpost"]
                and email not in wohnpost_missing
                and email not in ignore_emails
            ):
                wohnpost_and_geno_missing.append(email)
        for email in ml_members["gewerbemietende"]:
            if (
                email not in ml_members["wohnpost"]
                and email not in wohnpost_missing
                and email not in wohnpost_and_geno_missing
                and email not in ignore_emails
            ):
                wohnpost_and_geno_missing.append(email)

        ## Get newsletter subscribers from CreateSend
        if use_demo:
            newsletter_missing_cs = []
            newsletter_extra_cs_unsubscribed = []
            newsletter_extra_cs_bounced = []
            newsletter_extra_cs_deleted = []
            for email in members_email:
                if email in cs_unsubscribed:
                    newsletter_extra_cs_unsubscribed.append(email)
                elif email in cs_bounced:
                    newsletter_extra_cs_bounced.append(email)
                elif email in cs_deleted:
                    newsletter_extra_cs_deleted.append(email)
                elif email not in cs_active:
                    newsletter_missing_cs.append(email)
        else:
            cs_newsletter = createsend.List(list_id=settings.CREATESEND_LIST_ID_NEWSLETTER)
            cs_newsletter.auth({"api_key": settings.CREATESEND_API_KEY})

            cs_newsletter_unsubscribed = []
            page = 1
            num_pages = 1
            while page <= num_pages:
                res = cs_newsletter.unsubscribed(page=page, page_size=1000)
                num_pages = res.NumberOfPages
                page += 1
                for sub in res.Results:
                    cs_newsletter_unsubscribed.append(sub.EmailAddress.lower())

            cs_newsletter_bounced = []
            page = 1
            num_pages = 1
            while page <= num_pages:
                res = cs_newsletter.bounced(page=page, page_size=1000)
                num_pages = res.NumberOfPages
                page += 1
                for sub in res.Results:
                    cs_newsletter_bounced.append(sub.EmailAddress.lower())

            cs_newsletter_deleted = []
            page = 1
            num_pages = 1
            while page <= num_pages:
                res = cs_newsletter.deleted(page=page, page_size=1000)
                num_pages = res.NumberOfPages
                page += 1
                for sub in res.Results:
                    cs_newsletter_deleted.append(sub.EmailAddress.lower())

            cs_newsletter_active = []
            page = 1
            num_pages = 1
            while page <= num_pages:
                res = cs_newsletter.active(page=page, page_size=1000)
                num_pages = res.NumberOfPages
                page += 1
                for sub in res.Results:
                    cs_newsletter_active.append(sub.EmailAddress.lower())

            newsletter_missing_cs = []
            newsletter_extra_cs_unsubscribed = []
            newsletter_extra_cs_bounced = []
            newsletter_extra_cs_deleted = []
            for email in members_email:
                if email in cs_newsletter_unsubscribed:
                    newsletter_extra_cs_unsubscribed.append(email)
                elif email in cs_newsletter_bounced:
                    newsletter_extra_cs_bounced.append(email)
                elif email in cs_newsletter_deleted:
                    newsletter_extra_cs_deleted.append(email)
                elif email not in cs_newsletter_active:
                    newsletter_missing_cs.append(email)

        ## Build response with sub-groups for filtered items
        if use_demo:
            ret.append(
                {
                    "info": str(
                        _(
                            "<i>INFO: Mailman-API ist nicht konfiguriert. Dies ist ein Demo-Output.</i>"
                        )
                    ),
                    "variant": ResponseVariant.INFO.value,
                }
            )

        # Genossenschaft missing
        genossenschaft_objects = list(genossenschaft_missing)
        if genossenschaft_no_email:
            genossenschaft_objects.append(
                {
                    "label": str(_("Mitglieder ohne Email-Adresse (ignoriert)")),
                    "items": genossenschaft_no_email,
                    "variant": "info",
                }
            )
        if genossenschaft_duplicate:
            genossenschaft_objects.append(
                {
                    "label": str(_("Mitglieder mit doppelter Email-Adresse (ignoriert)")),
                    "items": genossenschaft_duplicate,
                    "variant": "info",
                }
            )
        ret.append(
            {"info": _("Mitglied nicht in genossenschaft-ML"), "objects": genossenschaft_objects}
        )

        ret.append(
            {
                "info": _("In genossenschaft-ML aber nicht Mitglied"),
                "objects": ml_members["genossenschaft"],
            }
        )

        # Bewohnende missing
        bewohnende_objects = list(bewohnende_missing)
        if bewohnende_no_email:
            bewohnende_objects.append(
                {
                    "label": str(_("Bewohnende ohne Email-Adresse (ignoriert)")),
                    "items": bewohnende_no_email,
                    "variant": "info",
                }
            )
        if bewohnende_duplicate:
            bewohnende_objects.append(
                {
                    "label": str(_("Bewohnende mit doppelter Email-Adresse (ignoriert)")),
                    "items": bewohnende_duplicate,
                    "variant": "info",
                }
            )
        ret.append(
            {"info": _("Bewohnende aber nicht in bewohnende-ML"), "objects": bewohnende_objects}
        )

        ret.append(
            {
                "info": _("In bewohnende-ML aber nicht Bewohnende"),
                "objects": ml_members["bewohnende"],
            }
        )
        ret.append(
            {"info": _("Mitglied aber NICHT in Newsletter(CS)"), "objects": newsletter_missing_cs}
        )
        ret.append(
            {
                "info": _("Mitglied aber UNSUBSCRIBED in Newsletter(CS)"),
                "objects": newsletter_extra_cs_unsubscribed,
            }
        )
        ret.append(
            {
                "info": _("Mitglied aber BOUNCED in Newsletter(CS)"),
                "objects": newsletter_extra_cs_bounced,
            }
        )
        ret.append(
            {
                "info": _("Mitglied aber DELETED in Newsletter(CS)"),
                "objects": newsletter_extra_cs_deleted,
            }
        )
        ret.append(
            {
                "info": _("In genossenschaft-ML aber nicht in Wohnpost (%s)")
                % len(wohnpost_missing),
                "objects": wohnpost_missing,
            }
        )
        ret.append(
            {
                "info": _(
                    "Bewohnende/Gewerbemietende weder in genossenschaft-ML noch in Wohnpost (%s)"
                )
                % len(wohnpost_and_geno_missing),
                "objects": wohnpost_and_geno_missing,
            }
        )
        ret.append({"info": _("Mailman warnings"), "objects": ml_warnings})

        return [
            item
            for item in ret
            if item.get("objects") or item.get("variant") == ResponseVariant.INFO.value
        ]


## TODO: Refactor to ClassBased view
@login_required
def run_maintenance_tasks(request):
    if not request.user.has_perm("geno.admin_import"):
        return unauthorized(request)

    ret = []
    # ret.append({'info': 'Creating new users:', 'objects': create_users()})

    return render(
        request, "geno/default.html", {"response": ret, "title": "Run maintenance tasks"}
    )


def send_member_mail_filter_rental(form, member_list):
    adr_list = []
    contracts = get_active_contracts(
        include_subcontracts=form.cleaned_data["include_subcontracts"]
        if form.cleaned_data["include_subcontracts"]
        else False
    )
    if form.cleaned_data["filter_building"]:
        contracts = contracts.filter(
            rental_units__building__in=form.cleaned_data["filter_building"]
        ).distinct()
    for contract in contracts:
        rental = []
        for ru in contract.rental_units.all():
            if (
                form.cleaned_data["select_rentaltype"] == "all"
                or form.cleaned_data["select_rentaltype"] == ru.rental_type
                or (
                    form.cleaned_data["select_rentaltype"] == "all_nobusiness"
                    and ru.rental_type not in ["Gewerbe", "Lager", "Hobby", "Parkplatz"]
                )
            ):
                rental.append(ru.name)
        if rental:
            for adr in contract.contractors.all():
                if adr.pk not in adr_list:
                    ## Filter by generic attribute: TODO: move this filter to be additive...
                    include = True
                    if form.cleaned_data["filter_genattribute"] != "none":
                        genatt = (
                            GenericAttribute.objects.filter(
                                name=form.cleaned_data["filter_genattribute"]
                            )
                            .filter(addresses=adr)
                            .first()
                        )
                        if genatt:
                            genatt_val = genatt.value
                        else:
                            genatt_val = "--OHNE--"
                        if form.cleaned_data["filter_genattribute_value"] != genatt_val:
                            include = False
                    if include:
                        adr_list.append(adr.pk)
                        member_list.append(
                            {
                                "id": adr.pk,
                                "address_id": adr.pk,
                                "contract_id": contract.id,
                                "member": str(adr),
                                "extra_info": "/".join(rental),
                                "member_type": "address",
                            }
                        )
                    # print('%s: %s' % (adr,'/'.join(rental)))
    return []  # No error


def filter_select_document_nostatement(form, adr_list):
    try:
        doctype_statement = DocumentType.objects.get(name="statement")
    except DocumentType.DoesNotExist:
        return adr_list
    date_6months_ago = timezone.now() - relativedelta(months=6)
    result = []
    for adr in adr_list:
        exclude = False
        if (
            "select_document" in form.cleaned_data
            and form.cleaned_data["select_document"] == "nostatement"
        ):
            docs = (
                Document.objects.filter(doctype=doctype_statement)
                .filter(addresses=adr)
                .filter(ts_created__gt=date_6months_ago)
            )
            if len(docs):
                exclude = True
        if not exclude:
            result.append(adr)
    return result


def send_member_mail_filter_addresses(form, member_list):
    for adr in filter_select_document_nostatement(form, Address.objects.filter(active=True)):
        member_list.append(
            {
                "id": adr.pk,
                "address_id": adr.pk,
                "member": str(adr),
                "extra_info": None,
                "member_type": "address",
            }
        )
    return []  # No error


def send_member_mail_filter_shares(form, member_list):
    addresses = []
    stype_filter = None
    stype_exclude = None

    if form.cleaned_data["select_sharetype"] == "shares":
        stype_filter = list(ShareType.objects.filter(name__startswith="Anteilschein"))
        if not stype_filter:
            return ["Beteiligungstypen nicht gefunden."]
    elif form.cleaned_data["select_sharetype"] == "loan_deposit":
        stype_filter = list(ShareType.objects.filter(name__startswith="Darlehen"))
        stype_filter.extend(list(ShareType.objects.filter(name="Depositenkasse")))
        if not stype_filter:
            return ["Beteiligungstypen nicht gefunden."]
    elif form.cleaned_data["select_sharetype"] == "with_interest":
        stype_filter = list(ShareType.objects.filter(name__startswith="Darlehen verzinst"))
        stype_filter.extend(list(ShareType.objects.filter(name="Depositenkasse")))
        if not stype_filter:
            return ["Beteiligungstypen nicht gefunden."]
    else:
        stype_exclude = list(ShareType.objects.filter(name="Darlehen spezial"))
        stype_exclude.extend(list(ShareType.objects.filter(name="Hypothek")))

    # print(stype_filter)

    shares = get_active_shares()
    if stype_filter:
        shares = shares.filter(share_type__in=stype_filter)
    if stype_exclude:
        shares = shares.exclude(share_type__in=stype_exclude)
    for share in shares:
        if share.name not in addresses:
            addresses.append(share.name)
    for adr in filter_select_document_nostatement(form, addresses):
        member_list.append(
            {
                "id": adr.pk,
                "address_id": adr.pk,
                "member": str(adr),
                "extra_info": None,
                "member_type": "address",
            }
        )
    return []


def send_member_mail_filter_members(form, member_list):
    errors = []
    members = Member.objects.all()
    if "ignore_join_date" in form.cleaned_data and form.cleaned_data["ignore_join_date"]:
        members = members.exclude(date_join__gt=form.cleaned_data["ignore_join_date"])
    if form.cleaned_data["select_flag_01"] == "true":
        members = members.filter(flag_01=True)
    elif form.cleaned_data["select_flag_01"] == "false":
        members = members.filter(flag_01=False)
    if form.cleaned_data["select_flag_02"] == "true":
        members = members.filter(flag_02=True)
    elif form.cleaned_data["select_flag_02"] == "false":
        members = members.filter(flag_02=False)
    if form.cleaned_data["select_flag_03"] == "true":
        members = members.filter(flag_03=True)
    elif form.cleaned_data["select_flag_03"] == "false":
        members = members.filter(flag_03=False)
    if form.cleaned_data["select_flag_04"] == "true":
        members = members.filter(flag_04=True)
    elif form.cleaned_data["select_flag_04"] == "false":
        members = members.filter(flag_04=False)
    if form.cleaned_data["select_flag_05"] == "true":
        members = members.filter(flag_05=True)
    elif form.cleaned_data["select_flag_05"] == "false":
        members = members.filter(flag_05=False)
    try:
        stype01 = ShareType.objects.get(name="Anteilschein Einzelmitglied")
    except ShareType.DoesNotExist:
        stype01 = None
    try:
        stype02 = ShareType.objects.get(name="Anteilschein Gründungsmitglied")
    except ShareType.DoesNotExist:
        stype02 = None
    for member in members:
        ## Filter active membership
        if not is_member(member.name):
            continue
        ## Filter out members without shares
        if "share_paid_01" in form.cleaned_data and form.cleaned_data["share_paid_01"]:
            share = get_active_shares().filter(share_type=stype01).filter(name=member.name).first()
            if not share:
                continue
        ## Filter out members with shares
        if "share_unpaid" in form.cleaned_data and form.cleaned_data["share_unpaid"]:
            share = (
                get_active_shares()
                .filter(Q(share_type=stype01) | Q(share_type=stype02))
                .filter(name=member.name)
                .first()
            )
            if share:
                continue
        ## Filter attributes (if set in form)
        att_info = []
        if form.cleaned_data["select_attributeA"]:
            attA = MemberAttribute.objects.filter(member=member).filter(
                attribute_type=form.cleaned_data["select_attributeA"]
            )
            if len(attA) > 1:
                errors.append(
                    "%s - ERROR: MORE THAN ONE ATTRIBUTE FOUND! %s -> %s, %s"
                    % (
                        member,
                        form.cleaned_data["select_attributeA"],
                        attA[0].value,
                        attA[1].value,
                    )
                )
                continue
            if (
                len(attA) == 1 and attA[0].value != form.cleaned_data["select_attributeA_value"]
            ) or (len(attA) == 0 and form.cleaned_data["select_attributeA_value"] != "--OHNE--"):
                ## has attribute but wrong value, or, has no attribute but one is required
                continue
            if len(attA) == 0:
                att_info.append("A: (kein Attribut)")
            else:
                att_info.append("A: %s [%s]" % (attA[0].value, attA[0].date))
        if form.cleaned_data["select_attributeB"]:
            attB = MemberAttribute.objects.filter(member=member).filter(
                attribute_type=form.cleaned_data["select_attributeB"]
            )
            if len(attB) > 1:
                errors.append(
                    "%s - ERROR: MORE THAN ONE ATTRIBUTE FOUND! %s -> %s, %s"
                    % (
                        member,
                        form.cleaned_data["select_attributeB"],
                        attB[0].value,
                        attB[1].value,
                    )
                )
                continue
            if (
                len(attB) == 1 and attB[0].value != form.cleaned_data["select_attributeB_value"]
            ) or (len(attB) == 0 and form.cleaned_data["select_attributeB_value"] != "--OHNE--"):
                ## has attribute but wrong value, or, has no attribute but one is required
                continue
            if len(attB) == 0:
                att_info.append("B: (kein Attribut)")
            else:
                att_info.append("B: %s [%s]" % (attB[0].value, attB[0].date))
        ## Add member
        member_list.append(
            {
                "id": member.pk,
                "address_id": member.name.pk,
                "member": str(member),
                "extra_info": "/".join(att_info),
                "member_type": "member",
            }
        )
    return errors


def send_member_mail_filter_by_invoice(form, member_list):
    errors = []
    if "filter_invoice" not in form.cleaned_data or form.cleaned_data["filter_invoice"] == "none":
        return errors
    count = 0
    for member in member_list:
        query = Invoice.objects.filter(person__id=member["address_id"], invoice_type="Invoice")
        if (
            "filter_invoice_category" in form.cleaned_data
            and form.cleaned_data["filter_invoice_category"]
        ):
            query = query.filter(invoice_category=form.cleaned_data["filter_invoice_category"])
        if "filter_invoice_consolidated" in form.cleaned_data:
            if form.cleaned_data["filter_invoice_consolidated"] == "true":
                query = query.filter(consolidated=True)
            elif form.cleaned_data["filter_invoice_consolidated"] == "false":
                query = query.filter(consolidated=False)
        if (
            "filter_invoice_daterange_min" in form.cleaned_data
            and form.cleaned_data["filter_invoice_daterange_min"]
        ):
            query = query.filter(date__gt=form.cleaned_data["filter_invoice_daterange_min"])
        if (
            "filter_invoice_daterange_max" in form.cleaned_data
            and form.cleaned_data["filter_invoice_daterange_max"]
        ):
            query = query.filter(date__lt=form.cleaned_data["filter_invoice_daterange_max"])
        if query.count():
            if form.cleaned_data["filter_invoice"] == "exclude":
                member["id"] = None  # exclude
        else:
            if form.cleaned_data["filter_invoice"] == "include":
                member["id"] = None  # exclude
        if member["id"]:
            count += 1
    if count == 0:
        errors.append(
            _("Keine Empfänger:innen gefunden, welche diesen Filterkriterien entsprechen.")
        )
    return errors


class MailWizardView(CohivaAdminViewMixin, FormView):
    title = _("Dokumente erstellen/versenden")
    step_title = _("Schritt 1 - Empfänger:innen filtern")
    form_action = reverse_lazy("geno:mail-wizard-start")
    back_url = None  # No back button on step 1
    permission_required = "geno.send_mail"
    template_name = "geno/member_send_mail.html"
    form_class = MemberMailForm
    last_step = False

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.result = {"errors": [], "show_results": False}

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(
            {
                "response": self.result,
                "attrs_button": {"form": "mail-wizard-form"},
                "back_url": self.back_url,
                "last_step": self.last_step,
            }
        )
        # Override title to show step
        context["title"] = self.step_title
        return context

    def form_valid(self, form):
        errors = []
        ## Filter members
        self.request.session["members"] = []
        if form.cleaned_data["base_dataset"] == "renters":
            errors = send_member_mail_filter_rental(form, self.request.session["members"])
            ## Filter for rental_type
        elif form.cleaned_data["base_dataset"] == "addresses":
            ## Filter for documents
            errors = send_member_mail_filter_addresses(form, self.request.session["members"])
        elif form.cleaned_data["base_dataset"] == "shares":
            errors = send_member_mail_filter_shares(form, self.request.session["members"])
        elif form.cleaned_data["base_dataset"] == "active_members":
            ## Filter members
            errors = send_member_mail_filter_members(form, self.request.session["members"])
        else:
            errors.append(_("Ungültiger Basis-Datensatz"))
        ## Filter by invoice existence
        errors.extend(send_member_mail_filter_by_invoice(form, self.request.session["members"]))
        if not errors:
            return HttpResponseRedirect(reverse("geno:mail-wizard-select"))
        self.result = {"errors": [{"info": error} for error in errors], "show_results": True}
        return self.get(self.request)


class MailWizardSelectView(MailWizardView):
    step_title = _("Schritt 2 - Empfänger:innen auswählen")
    form_action = reverse_lazy("geno:mail-wizard-select")
    back_url = reverse_lazy("geno:mail-wizard-start")

    def get(self, request, *args, **kwargs):
        if "members" not in request.session:
            return HttpResponseRedirect(reverse("geno:mail-wizard-start"))
        return super().get(request, *args, **kwargs)

    def get_form(self, form_class=None):
        return MemberMailSelectForm(
            self.request.POST or None,
            initial=self.get_initial(),
            members=self.request.session["members"],
        )

    def form_valid(self, form):
        self.request.session["select_members"] = form.cleaned_data["select_members"]
        return HttpResponseRedirect(reverse("geno:mail-wizard-action"))


class MailWizardActionView(MailWizardView):
    step_title = _("Schritt 3 - Aktionen ausführen")
    form_class = MemberMailActionForm
    form_action = reverse_lazy("geno:mail-wizard-action")
    back_url = reverse_lazy("geno:mail-wizard-select")
    last_step = True

    def get_initial(self):
        return {"email_copy": settings.GENO_DEFAULT_EMAIL}

    def form_valid(self, form):
        form.cleaned_data["members"] = []
        for m in self.request.session["members"]:
            if str(m["id"]) in self.request.session["select_members"]:
                form.cleaned_data["members"].append(m)
        self.result = send_member_mail_process(form.cleaned_data)
        if isinstance(self.result, HttpResponse):
            return self.result
        self.result["show_results"] = True
        return self.get(self.request)


class TransactionUploadView(CohivaAdminViewMixin, FormView):
    title = "Zahlungen erfassen"
    permission_required = ("geno.transaction", "geno.transaction_invoice", "geno.add_invoice")
    template_name = "geno/transaction_upload.html"
    form_class = TransactionUploadForm
    import_message = ""
    import_items = []
    actions = []  ## title, path, items (for dropdown), method_name (for dropdown?), icon, variant
    if settings.DEMO:
        actions.append(
            {
                "title": "DEMO-Zahlungsdatei erzeugen",
                "path": reverse_lazy("geno:transaction-testdata"),
            }
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(
            {
                "import_message": self.import_message,
                "import_items": self.import_items,
                "item_name": "Buchungen",
                # "title": "Bankauszug verarbeiten",
            }
        )
        return context

    def form_valid(self, form):
        if "file" in self.request.FILES:
            uploaded_file = self.request.FILES["file"]
            transaction_data = process_transaction_file(self.request.FILES["file"])

            if transaction_data["error"]:
                messages.error(
                    self.request, "Konnte Datei nicht verarbeiten: %s" % transaction_data["error"]
                )
                logger.error(
                    f"Transaction upload: Error while processing {uploaded_file}: "
                    f"{transaction_data['error']}"
                )
            elif transaction_data["type"].startswith("camt.053") or transaction_data[
                "type"
            ].startswith("camt.054"):
                self.import_message = "Import von Buchungen aus %s:" % uploaded_file
                self.import_items = process_sepa_transactions(transaction_data["data"])
                if len(self.import_items["success"]):
                    logger.info(
                        f"Transaction upload: Imported {len(self.import_items['success'])} records "
                        f"from {uploaded_file}."
                    )
            else:
                messages.error(
                    self.request,
                    "Konnte Datei nicht verarbeiten: Unbekannter typ %s"
                    % transaction_data["type"],
                )
                logger.error(
                    f"Transaction upload: Error while processing {uploaded_file}: "
                    f"Invalid type {transaction_data['type']}"
                )

        else:
            messages.error(self.request, "Konnte Datei nicht hochladen.")
        return self.get(self.request)


class TransactionInvoiceView(CohivaAdminViewMixin, FormView):
    title = "Zahlung einer Rechnung manuell erfassen"
    permission_required = ("geno.transaction", "geno.transaction_invoice", "geno.add_invoice")
    template_name = "geno/transaction.html"
    form_class = TransactionFormInvoice

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(
            {
                "form_action": "/geno/transaction_invoice/",
            }
        )
        return context

    def form_valid(self, form):
        invoice_to_pay = form.cleaned_data["invoice"]
        if form.cleaned_data["amount"]:
            amount = form.cleaned_data["amount"]
        else:
            amount = invoice_to_pay.amount
        ret = pay_invoice(invoice_to_pay, form.cleaned_data["date"], amount)
        if ret:
            messages.error(
                self.request,
                "Zahlung von %s konnte nicht gebucht werden: %s" % (invoice_to_pay, ret),
            )
        else:
            messages.success(self.request, "Zahlung gebucht: %s [%.2f]" % (invoice_to_pay, amount))
            self.initial = {"date": form.cleaned_data["date"]}
        return self.get(self.request)


class TransactionManualView(CohivaAdminViewMixin, FormView):
    title = "Zahlung ohne Rechnung manuell erfassen"
    permission_required = ("geno.transaction", "geno.transaction_invoice", "geno.add_invoice")
    template_name = "geno/transaction.html"
    form_class = TransactionForm
    error_flag = False

    def get_initial(self):
        now = datetime.datetime.now()
        if now.month < 6:
            default_transaction = "fee%s" % (now.year - 1)
        else:
            default_transaction = "fee%s" % now.year
        return {"transaction": default_transaction}

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(
            {
                "form_action": "/geno/transaction/",
                # "title": "Bankauszug verarbeiten",
            }
        )
        return context

    def form_valid(self, form):
        if form.cleaned_data["transaction"][0:3] == "fee":
            self.error_flag = self.process_fee(form)
        elif form.cleaned_data["transaction"] in (
            "as_single",
            "as_extra",
            "as_founder",
            "development",
        ):
            self.error_flag = self.process_share(form)
        else:
            self.error_flag = self.process_default(form)
        return self.get(self.request)

    def process_fee(self, form):
        fee_year = form.cleaned_data["transaction"][3:]
        try:
            att_type = MemberAttributeType.objects.get(name="Mitgliederbeitrag %s" % fee_year)
        except MemberAttributeType.DoesNotExist:
            messages.error(
                self.request,
                "Mitglieder Attribut 'Mitgliederbeitrag %s' existiert nicht." % fee_year,
            )
            return True

        member = Member.objects.filter(name=form.cleaned_data["name"])
        if len(member) != 1:
            messages.error(
                self.request,
                "Mitglied nicht gefunden oder nicht eindeutig: %s" % form.cleaned_data["name"],
            )
            return True
        att = MemberAttribute.objects.filter(member=member[0], attribute_type=att_type)
        for a in att:
            messages.info(
                self.request,
                "Mitglieder Attribut gefunden: %s - %s" % (a.date, a.value),
            )
        if len(att) == 0:
            ## Create new attribute
            att = MemberAttribute(member=member[0], attribute_type=att_type)
        elif len(att) == 1:
            att = att[0]
            if att.value.startswith("Bezahlt"):
                messages.error(self.request, "Schon als bezahlt markiert")
                return True
            elif (
                att.value != "Mail-Rechnung geschickt"
                and att.value != "Mail-Reminder geschickt"
                and att.value != "Rechnung geschickt"
                and att.value != "Gefordert"
                and att.value != "Brief-Rechnung geschickt"
                and att.value != "Brief-Reminder geschickt"
                and att.value != "Brief-Mahnung geschickt"
                and att.value != "Brief-Mahnung2 geschickt"
            ):
                messages.error(self.request, f"Unbekannter Attribut-Wert: {att.value}")
                return True
        else:
            messages.error(self.request, "Mehr als ein Attribut gefunden")
            return True

        ## Add transaction to financial accounting
        try:
            with AccountingManager() as book:
                to_account = Account.from_settings(AccountKey.DEFAULT_DEBTOR_MANUAL)
                from_account = Account.from_settings(AccountKey.MEMBER_FEE)
                description = "Mitgliederbeitrag %s %s" % (fee_year, member[0])
                amount = 80.00
                book.add_transaction(
                    amount, from_account, to_account, form.cleaned_data["date"], description
                )
                messages.success(self.request, f"Buchung erstellt: CHF {amount}, {description}")
        except Exception as e:
            messages.error(self.request, "Konnte Buchung nicht erstellen: %s" % e)
            return True

        ## Update/add attribute
        att.value = "Bezahlt"
        att.date = form.cleaned_data["date"]
        att.save()
        messages.success(
            self.request,
            "Mitglieder Attribut hinzugefügt/aktualisiert: %s - %s [%s]"
            % (att.date, att.value, att.member),
        )
        return False

    def process_share(self, form):
        if form.cleaned_data["transaction"] == "development":
            count = 1
            value = form.cleaned_data["amount"]
            share_type = "Entwicklungsbeitrag"
        elif form.cleaned_data["amount"] and float(form.cleaned_data["amount"]) % 200.00 == 0.0:
            value = 200
            count = int(form.cleaned_data["amount"] / value)
            if form.cleaned_data["transaction"] == "as_single":
                share_type = "Anteilschein Einzelmitglied"
            elif form.cleaned_data["transaction"] == "as_founder":
                share_type = "Anteilschein Gründungsmitglied"
            else:
                share_type = "Anteilschein freiwillig"
        else:
            messages.error(self.request, "Betrag ist kein Vielfaches von 200.-!")
            return True

        share = Share(
            name=form.cleaned_data["name"],
            share_type=ShareType.objects.get(name=share_type),
            state="bezahlt",
            date=form.cleaned_data["date"],
            quantity=count,
            value=value,
        )
        share.save()
        messages.info(
            self.request,
            "%sx CHF %s %s hinzugefügt - %s [%s]"
            % (
                count,
                value,
                share_type,
                form.cleaned_data["date"],
                form.cleaned_data["name"],
            ),
        )
        return False

    def process_default(self, form):
        if not form.cleaned_data["amount"]:
            messages.error(self.request, "Kein Betrag angegeben!")
            return True

        if len(form.cleaned_data["note"]):
            note = form.cleaned_data["note"]
        else:
            note = None
        ret_error = process_transaction(
            form.cleaned_data["transaction"],
            form.cleaned_data["date"],
            form.cleaned_data["name"],
            form.cleaned_data["amount"],
            None,
            note,
        )
        info = "%s: %s CHF %s [%s]" % (
            form.cleaned_data["transaction"],
            form.cleaned_data["date"],
            form.cleaned_data["amount"],
            form.cleaned_data["name"],
        )
        if ret_error:
            messages.error(self.request, "FEHLER bei der Buchung: %s -- %s" % (info, ret_error))
            return True
        messages.success(self.request, "Buchung ausgeführt: %s" % (info))
        return False


@login_required
def transaction_upload_process(request):
    if not request.user.has_perm("geno.transaction"):
        return unauthorized(request)

    initial = {}
    ret = []
    info_text = ""
    get_next = False
    transaction = None

    if request.method == "POST":
        transaction = request.session["last_transaction"]
        form = TransactionUploadProcessForm(request.POST, initial=initial, transaction=transaction)
        if form.is_valid():
            if form.cleaned_data["transaction"] == "ignore":
                get_next = True
            else:
                note = form.cleaned_data["note"]
                if form.cleaned_data["extra_info"]:
                    if note:
                        note = "%s / %s" % (note, form.cleaned_data["extra_info"])
                    else:
                        note = form.cleaned_data["extra_info"]
                error = process_transaction(
                    form.cleaned_data["transaction"],
                    form.cleaned_data["date"],
                    form.cleaned_data["name"],
                    form.cleaned_data["amount"],
                    form.cleaned_data["save_sender"],
                    note,
                )
                if error:
                    messages.error(request, error)
                else:
                    messages.info(
                        request,
                        "Gebucht: %s/%s/%s/Fr. %s"
                        % (
                            transaction["date"],
                            transaction["person"],
                            transaction["note"],
                            transaction["amount"],
                        ),
                    )
                    get_next = True
        else:
            messages.error(request, "Form is invalid")
    else:
        get_next = True

    while get_next and "transaction_data" in request.session:
        if request.session["transaction_data"]:
            transaction = request.session["transaction_data"].pop()
            request.session["last_transaction"] = transaction
            initial = guess_transaction(transaction)
            info_text = "<ul><li>%s</li><li>%s</li><li>Fr. %s</li></ul>" % (
                transaction["person"],
                transaction["note"],
                transaction["amount"],
            )
            if initial["process"]:
                get_next = False
        else:
            info_text = "<ul><li>KEINE WEITEREN TRANSAKTIONEN MEHR GEFUNDEN</li></ul>"
            transaction = None
            get_next = False

    form = TransactionUploadProcessForm(None, initial=initial, transaction=transaction)
    return render(
        request,
        "geno/transaction_upload.html",
        {
            "response": ret,
            "title": "Bankauszug verarbeiten",
            "info": "Schritt 2: Buchungsaktionen auswählen",
            "form": form,
            "form_action": "/geno/transaction_upload/process/",
            "info_text": info_text,
        },
    )


@login_required
def transaction_testdata(request):
    if not settings.DEMO:
        return unauthorized(request)
    data = generate_demo_camt053_file()
    filename = f"camt053_demo_data_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xml"
    response = HttpResponse(data, content_type="application/xml")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


class InvoiceManualView(CohivaAdminViewMixin, TemplateView):
    title = "Rechnung erstellen"
    permission_required = (
        "geno.canview_billing",
        "geno.transaction",
        "geno.transaction_invoice",
        "geno.add_invoice",
    )
    template_name = "geno/invoice_manual.html"
    error_flag = False
    #    actions = [
    #        {
    #            "title": "Mietzins-Rechnungen erstellen",
    #            "variant": ActionVariant.PRIMARY,
    #            "items": [
    #                {
    #                    "title": "Bis nächster Monat",
    #                    "path": (reverse_lazy("geno:invoice-batch-generate"), "?date=next_month"),
    #                    "permission_required": ("geno.view_rentalunit",),
    #                },
    #                {
    #                    "title": "Bis aktueller Monat",
    #                    "path": reverse_lazy("geno:invoice-batch-gernerate"),
    #                    "permission_required": ("geno.view_rentalunit",),
    #                },
    #                {
    #                    "title": "Bis letzten Monat",
    #                    "path": (reverse_lazy("geno:invoice-batch-generate"), "?date=last_month"),
    #                    "permission_required": ("geno.view_rentalunit",),
    #                },
    #            ],
    #        }
    #    ]

    def post(self, request, *args, **kwargs):
        form = self.get_form()
        formset = self.get_formset()
        if form.is_valid() and formset.is_valid():
            ret = self.process(form, formset)
            if isinstance(ret, FileResponse):
                return ret
        return self.get(request, *args, **kwargs)

    def get_invoice_templates(self):
        """
        Define invoice category templates with pre-filled data.
        Returns a dict mapping category name to template data.

        NOTE: Keys must match the exact InvoiceCategory.name values in the database.
        If category names change in the admin, update these keys accordingly.
        """
        return {
            "Miete Gästezimmer": {
                "extra_text": _(
                    "Die ersten 3 Nächte sind gemäss Reglement Jokerzimmer kostenlos."
                ),
                "lines": [
                    {
                        "text": _("X Nächte à Fr. 22.00, Gästezimmer 003"),
                        "amount": None,
                        "date": datetime.date.today(),
                    },
                    {
                        "text": _("X Nächte à Fr. 17.00, Gästezimmer 410"),
                        "amount": None,
                        "date": datetime.date.today(),
                    },
                    {
                        "text": _("X Nächte à Fr. 17.00, Gästezimmer 509"),
                        "amount": None,
                        "date": datetime.date.today(),
                    },
                ],
            },
            "Mitgliedschaft": {
                "extra_text": _(
                    "Vielen Dank für die Beitrittserklärung. Die Mitgliedschaft in der "
                    "Genossenschaft ist verbunden mit dem Zeichnen von Anteilscheinen und einer "
                    "einmaligen Beitrittsgeb ühr. Dies stellen wir hiermit in Rechnung. Nach "
                    "Zahlungseingang folgt dann die Bestätigung der Mitgliedschaft."
                ),
                "lines": [
                    {
                        "text": _("Beitrittsgeb ühr Genossenschaft"),
                        "amount": 200,
                        "date": datetime.date.today(),
                    },
                    {
                        "text": _("Zeichnung von 1 Anteilschein der Genossenschaft"),
                        "amount": 200,
                        "date": datetime.date.today(),
                    },
                ],
            },
        }

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = self.get_form()
        context["formset"] = self.get_formset()
        context["error_flag"] = self.error_flag
        return context

    def get_form(self):
        if self.request.method == "POST":
            form = ManualInvoiceForm(self.request.POST)
        else:
            initial_data = {"date": datetime.date.today()}

            # If category is selected via GET parameter, preserve it and apply template if available
            category_id = self.request.GET.get("category")
            if category_id:
                try:
                    category_id = int(category_id)
                    # Fetch the actual InvoiceCategory instance
                    category = InvoiceCategory.objects.filter(
                        id=category_id, active=True, manual_allowed=True
                    ).first()
                    if category:
                        # Always preserve the selected category (as model instance)
                        initial_data["category"] = category

                        # Apply template data if available for this category (by name)
                        templates = self.get_invoice_templates()
                        if category.name in templates:
                            template = templates[category.name]
                            initial_data["extra_text"] = template["extra_text"]
                except (ValueError, TypeError):
                    pass

            form = ManualInvoiceForm(initial=initial_data)
        return form

    def get_formset(self):
        MAX_FORMS = 5

        invoice_formset_class = formset_factory(
            ManualInvoiceLineForm,
            extra=0,
            max_num=MAX_FORMS,
            validate_max=True,
            min_num=1,
            validate_min=True,
        )

        if self.request.method == "POST":
            formset = invoice_formset_class(self.request.POST, self.request.FILES)
        else:
            # Check if category is selected via GET parameter
            category_id = self.request.GET.get("category")
            extra_param = self.request.GET.get("extra")
            templates = self.get_invoice_templates()

            # Determine initial data based on parameters
            if extra_param:
                # User explicitly requested a specific number of forms (via add/remove buttons)
                # This takes precedence over template defaults
                try:
                    extra_forms = int(extra_param)
                    num_forms = min(1 + extra_forms, MAX_FORMS)
                    initial_data = [{"date": datetime.date.today()} for _ in range(num_forms)]
                except (ValueError, TypeError):
                    initial_data = [{"date": datetime.date.today()}]
            elif category_id:
                # Category selected but no explicit extra param - use template if available
                try:
                    category_id = int(category_id)
                    # Fetch the category to get its name for template lookup
                    category = InvoiceCategory.objects.filter(
                        id=category_id, active=True, manual_allowed=True
                    ).first()

                    if category and category.name in templates:
                        # Use template data to create initial forms
                        template = templates[category.name]
                        initial_data = template["lines"]
                    else:
                        # Category selected but no template, start with 1 empty form
                        initial_data = [{"date": datetime.date.today()}]
                except (ValueError, TypeError):
                    # Invalid category ID, start with 1 empty form
                    initial_data = [{"date": datetime.date.today()}]
            else:
                # No category or extra param - default to 1 empty form
                initial_data = [{"date": datetime.date.today()}]

            formset = invoice_formset_class(initial=initial_data)
        return formset

    def process(self, form, formset):
        email_template = None

        ## TODO: Use InvoiceCreator to do this
        invoice_category = form.cleaned_data["category"]
        address = form.cleaned_data["address"]
        invoice_date = form.cleaned_data["date"]

        lines_count = 0
        total_amount = 0
        invoice_lines = []
        comment = []
        for line in formset.cleaned_data:
            if line["amount"]:
                if not len(line["text"]) or not line["date"]:
                    messages.error(
                        self.request,
                        "Zeile mit Betrag aber ohne Datum/Beschreibung! "
                        "Bitte eingeben oder Betrag löschen.",
                    )
                    self.error_flag = True
                    break
                line["date"] = line["date"].strftime("%d.%m.%Y")
                line["total"] = nformat(line["amount"])
                invoice_lines.append(line)
                lines_count += 1
                total_amount += line["amount"]
                comment.append("%s CHF %s" % (line["text"], line["total"]))

        if not lines_count:
            messages.error(
                self.request,
                "Keine Rechnungspositionen eingegeben! Bitte mindestens eine Zeile ausfüllen.",
            )
            self.error_flag = True

        if not self.error_flag and "submit_action_save" in self.request.POST:
            dry_run = False
            # Check if email checkbox is checked
            if self.request.POST.get("send_email"):
                ## Send email
                email_template = "email_invoice.html"

            try:
                invoice = add_invoice(
                    None,
                    invoice_category,
                    invoice_category.name,
                    invoice_date,
                    total_amount,
                    address=address,
                    comment="/".join(comment),
                )
                invoice_id = invoice.id
            except Exception as e:
                messages.error(self.request, f"Konnte Rechnungs-Objekt nicht erzeugen: {e}")
                self.error_flag = True
                invoice = None
                invoice_id = None
        else:
            ## Test/Preview
            dry_run = True
            invoice = None
            invoice_id = 9999999999

        if not self.error_flag:
            ref_number = get_reference_nr(invoice_category, address.id, invoice_id)
            output_filename = "Rechnung_%s_%s_%s.pdf" % (
                invoice_category.name,
                invoice_date.strftime("%Y%m%d"),
                esr.compact(ref_number),
            )
            context = address.get_context()
            if invoice_category.name == "Geschäftsstelle":
                context["betreff"] = "Rechnung"
            else:
                context["betreff"] = "Rechnung %s" % invoice_category.name
            context["extra_text"] = form.cleaned_data["extra_text"]
            context["invoice_date"] = invoice_date.strftime("%d.%m.%Y")
            context["invoice_duedate"] = (invoice_date + relativedelta(months=1)).strftime(
                "%d.%m.%Y"
            )
            context["invoice_nr"] = invoice_id
            context["show_liegenschaft"] = False
            context["contract_info"] = None
            context["sect_rent"] = False
            context["sect_generic"] = True
            context["generic_info"] = invoice_lines
            context["s_generic_total"] = nformat(total_amount)
            context["qr_amount"] = total_amount
            context["qr_extra_info"] = "Rechnung %s" % context["invoice_nr"]
            context["preview"] = dry_run

            email_subject = "%s Nr. %s/%s" % (
                context["betreff"],
                context["invoice_nr"],
                context["invoice_date"],
            )

            (ret, mails_sent, mail_recipient) = create_qrbill(
                ref_number,
                address,
                context,
                output_filename,
                render,
                email_template,
                email_subject,
                dry_run,
            )

            info = "%s CHF %s Nr. %s/%s, %s" % (
                address,
                total_amount,
                context["invoice_nr"],
                context["invoice_date"],
                context["betreff"],
            )
            if ret:
                messages.error(
                    self.request, "Fehler beim Erzeugen der Rechnung für %s: %s" % (info, ret)
                )
                self.error_flag = True
            elif email_template:
                if mails_sent == 1:
                    messages.success(
                        self.request,
                        "Email '%s' mit QR-Rechnung an %s geschickt. %s"
                        % (email_subject, mail_recipient, output_filename),
                    )
                else:
                    messages.error(
                        self.request,
                        "FEHLER beim Versenden des Emails '%s' mit QR-Rechnung an %s! %s."
                        % (email_subject, mail_recipient, output_filename),
                    )
                    self.error_flag = True
            else:
                pdf_file = open("/tmp/%s" % output_filename, "rb")
                resp = FileResponse(pdf_file, content_type="application/pdf")
                resp["Content-Disposition"] = "attachment; filename=%s" % output_filename
                return resp
            if self.error_flag and invoice:
                # Rollback transaction
                invoice.delete()
        return self.get(self.request)


class InvoiceBatchView(CohivaAdminViewMixin, FormView):
    title = "Mietzinsrechnungen erstellen"
    permission_required = (
        "geno.canview_billing",
        "geno.transaction",
        "geno.transaction_invoice",
        "geno.add_invoice",
        "geno.view_rentalunit",
    )
    form_class = SendInvoicesForm
    template_name = "geno/invoice_batch.html"

    def get_context_data(self, **kwargs):
        return super().get_context_data(
            **kwargs,
            submit_title="Mietzinsrechnungen erstellen (Probelauf)",
        )

    def get_initial(self):
        return {
            "date": "next_month",
            "buildings": [b.id for b in Building.objects.filter(active=True)],
        }

    def form_valid(self, form):
        buildings = form.cleaned_data["buildings"]
        date = form.cleaned_data["date"]
        return HttpResponseRedirect(
            f"{reverse('geno:invoice-batch-generate')}?date={date}&buildings[]="
            + ",".join([str(b) for b in buildings])
        )


class InvoiceBatchGenerateView(DryRunActionView):
    title = _("Mietzinsrechnung erstellen")
    permission_required = (
        "geno.canview_billing",
        "geno.transaction",
        "geno.transaction_invoice",
        "geno.add_invoice",
        "geno.view_rentalunit",
    )
    # Uses generic dry_run_action.html template
    action = "create"
    navigation_view_name = (
        "geno.views.InvoiceBatchView"  # Use this for navigation rendering (active tabs etc.)
    )

    def get_action_params(self):
        """Return URL parameters for building the execute URL."""
        return {
            "date": self.request.GET.get("date", ""),
            "buildings[]": self.request.GET.get("buildings[]", ""),
        }

    def get_item_count(self):
        """Extract invoice count from the results."""
        if self.result and len(self.result) > 0:
            last_section = self.result[-1]
            if isinstance(last_section, dict) and last_section.get("objects"):
                objects = last_section["objects"]
                if objects and isinstance(objects[-1], str) and "Rechnungen" in objects[-1]:
                    invoice_count_text = objects[-1]
                    last_section["objects"] = objects[:-1]
                    return int(invoice_count_text.split()[0])
        return None

    def get_item_label(self):
        """Return 'Rechnung' as the item label."""
        return _("Rechnung")

    def process_action(self, dry_run):
        """Process invoice generation with given dry_run mode."""
        action = self.kwargs.get("action", self.action)
        key = self.kwargs.get("key", None)
        key_type = self.kwargs.get("key_type", None)

        if action not in ("create", "download"):
            return [{"info": _("Ungültige Aktion: %s") % action}]

        ret = []
        if action == "download":
            if key_type != "contract":
                raise RuntimeError("invoice(): Key type %s not implemented yet!" % key_type)
            ## Just download PDF of invoices for this contract
            download_only = key
            dry_run = True  # Always dry run for downloads
        else:
            download_only = None

        today = datetime.date.today()
        reference_date = datetime.date(today.year, today.month, 1)
        if self.request.GET.get("date", "") == "this_month":
            pass  # NOOP this is the default
        elif self.request.GET.get("date", "") == "last_month":
            if today.month == 1:
                reference_date = datetime.date(today.year - 1, 12, 1)
            else:
                reference_date = datetime.date(today.year, today.month - 1, 1)
        elif self.request.GET.get("date", "") == "next_month":
            if today.month == 12:
                reference_date = datetime.date(today.year + 1, 1, 1)
            else:
                reference_date = datetime.date(today.year, today.month + 1, 1)
        elif len(self.request.GET.get("date", "")) == 10:
            reference_date = datetime.datetime.strptime(
                self.request.GET.get("date"), "%Y-%m-%d"
            ).date()

        ret.append(
            {
                "info": _("Optionen"),
                "objects": [
                    _("Probelauf: %s") % (_("Ja") if dry_run else _("Nein")),
                    _("Rechungen bis: %s") % reference_date,
                ],
            }
        )

        building_ids = [
            int(x) for x in self.request.GET.get("buildings[]", "").split(",") if x.strip()
        ]
        invoices = create_invoices(
            dry_run,
            reference_date,
            self.request.GET.get("single_contract", None),
            building_ids,
            download_only,
        )
        if isinstance(invoices, str):
            pdf_file = open("/tmp/%s" % invoices, "rb")
            resp = FileResponse(pdf_file, content_type="application/pdf")
            resp["Content-Disposition"] = "attachment; filename=%s" % invoices
            return resp
        ret.extend(invoices)
        return ret


class ResidentUnitListView(CohivaAdminViewMixin, TemplateView):
    title = "Mietobjektespiegel"
    permission_required = "geno.rental_objects"

    def get(self, *args, **kwargs):
        return self.generate_resident_unit_list()

    def generate_resident_unit_list(self):
        data = []
        data_fields = [
            ("building", "Liegenschaft"),
            ("ru_name", "Bezeichnung"),
            ("ru_type", "Typ"),
            ("ru_floor", "Stockwerk"),
            ("ru_area", "Fläche (m2)"),
            ("ru_area_add", "Zusatzfläche (m2)"),
            ("ru_height", "Raumhöhe (m)"),
            ("ru_rooms", "Anzahl Zimmer"),
            ("ru_occ", "Min. Belegung"),
            ("n_adults", "Erwachsene"),
            ("n_children", "Anzahl Kinder"),
            ("ru_rent_netto", "Nettomiete"),
            ("ru_rent_total", "Bruttomiete"),
            ("contract_rent_reduction", "Mietzinsreduktion auf Nettomiete"),
            ("contract_rent_reservation", "Mietzinsvorbehalt auf Nettomiete"),
            ("ru_nk", "NK akonto"),
            ("ru_nk_flat", "NK pauschal"),
            ("ru_nk_electricity", "NK Strom"),
            ("name", "Name Mieter:in 1"),
            ("first_name", "Vorname Mieter:in 1"),
            ("email", "Email Mieter:in 1"),
            ("name2", "Name Mieter:in 2"),
            ("first_name2", "Vorname Mieter:in 2"),
            ("other_names", "Weitere Mieter:innen"),
            ("children", "Kinder"),
            ("contract_date", "Vertragsbeginn"),
            ("comment", "Bemerkungen"),
        ]

        rentalUnits = RentalUnit.objects.filter(active=True)

        for ru in rentalUnits:
            obj = lambda: None
            obj._fields = list(map(lambda x: x[0], data_fields))
            obj.building = ru.building.name
            if ru.label:
                obj.ru_name = "%s %s" % (ru.name, ru.label)
            else:
                obj.ru_name = "%s %s" % (ru.name, ru.rental_type)
            obj.ru_type = ru.rental_type
            obj.ru_floor = ru.floor

            # obj.ru_label = ru.label
            obj.ru_area = ru.area
            obj.ru_area_add = ru.area_add
            obj.ru_height = ru.height
            obj.ru_rooms = ru.rooms
            obj.ru_occ = ru.min_occupancy
            obj.n_adults = 0
            obj.n_children = 0
            obj.ru_nk = ru.nk
            obj.ru_nk_flat = ru.nk_flat
            obj.ru_nk_electricity = ru.nk_electricity
            obj.ru_rent_netto = ru.rent_netto
            obj.ru_rent_total = ru.rent_total

            ## Default values
            obj.name = "(Leerstand)"
            obj.first_name = ""
            obj.email = ""
            obj.name2 = ""
            obj.first_name2 = ""
            obj.contract_date = ""
            obj.comment = ""
            other_names = []
            children = []
            comments = []

            contracts = get_active_contracts().filter(rental_units__id__exact=ru.id)
            n_contracts = contracts.count()
            if n_contracts > 0:
                if n_contracts > 1:
                    comments.append("ACHTUNG: MEHR ALS 1 VERTRAG!")
                    logger.warning("Rental unit %s has %d contracts!" % (ru, n_contracts))
                count = 0
                contract = contracts.first()
                if contract.main_contact:
                    if contract.main_contact.organization:
                        obj.name = contract.main_contact.organization
                        obj.first_name = "%s %s" % (
                            contract.main_contact.first_name,
                            contract.main_contact.name,
                        )
                    else:
                        obj.name = contract.main_contact.name
                        obj.first_name = contract.main_contact.first_name
                    obj.email = contract.main_contact.email
                    count += 1
                for adr in contract.contractors.all():
                    if adr == contract.main_contact:
                        continue
                    if adr.organization:
                        str_name = adr.organization
                        str_first_name = "%s %s" % (adr.first_name, adr.name)
                        str_full_name = "%s/%s %s" % (adr.organization, adr.first_name, adr.name)
                    else:
                        str_name = adr.name
                        str_first_name = adr.first_name
                        str_full_name = "%s %s" % (adr.first_name, adr.name)
                    if count == 0:
                        obj.name = str_name
                        obj.first_name = str_first_name
                        obj.email = adr.email
                    elif count == 1:
                        obj.name2 = str_name
                        obj.first_name2 = str_first_name
                    else:
                        other_names.append(str_full_name)
                    count += 1
                obj.n_adults = count
                obj.contract_date = contract.date.strftime("%d.%m.%Y")
                if contract.rent_reduction:
                    obj.contract_rent_reduction = contract.rent_reduction
                if contract.rent_reservation:
                    obj.contract_rent_reservation = contract.rent_reservation

                for child in contracts.first().children.all():
                    children.append(
                        "%s %s (%s)" % (child.name.first_name, child.name.name, child.age())
                    )

            obj.other_names = ", ".join(other_names)
            obj.children = ", ".join(children)
            obj.n_children = len(children)
            obj.comment = ", ".join(comments)

            data.append(obj)

        return export_data_to_xls(
            data,
            title="Mietobjektespiegel",
            header=dict(data_fields),
            filename_suffix="mietobjektespiegel",
        )


@login_required
def rental_unit_list_create_documents(request, doc="mailbox"):
    if request.GET.get("makezip", "") == "yes":
        makezip = True
    else:
        makezip = False

    ret = []
    zipcount = 0
    zipfile_content = []
    if doc == "mailbox":
        options = {"beschreibung": "Briefkasten", "link_url": "/geno/rental/units/mailbox"}
    elif doc == "protocol":
        options = {"beschreibung": "Protokoll", "link_url": "/geno/rental/units/protocol"}
    else:
        raise RuntimeError("Unknown doc in rental_unit_list_create_documents '%s'." % doc)
    try:
        doctype = DocumentType.objects.get(name=doc)
        template = doctype.template.file.path
    except DocumentType.DoesNotExist:
        ret.append({"info": f"FEHLER: Dokumenttyp {doc} nicht gefunden."})
        return render(
            request,
            "geno/default.html",
            {
                "response": ret,
                "title": "Mietobjekt-Dokumente erzeugen - %s" % options["beschreibung"],
            },
        )

    for ru in RentalUnit.objects.filter(active=True):
        context = {"title": ru.name}

        if ru.floor == "Keller":
            continue

        first_names = {}
        count_names = 0
        count_first_names = 0
        mietpartei = []

        contracts = get_active_contracts().filter(rental_units__id__exact=ru.id)
        n_contracts = contracts.count()
        if n_contracts > 0:
            if n_contracts > 1:
                logger.warning("Rental unit %s has %d contracts!" % (ru, n_contracts))
            contract = contracts.first()
            for adr in contract.contractors.all():
                if adr.organization:
                    str_name = adr.organization
                    str_first_name = None
                    mietpartei.insert(0, str_name)
                else:
                    str_name = adr.name
                    str_first_name = adr.first_name
                    mietpartei.append("%s %s" % (str_first_name, str_name))
                if str_name not in first_names:
                    first_names[str_name] = []
                    count_names += 1
                if str_first_name:
                    first_names[str_name].append(str_first_name)
                    count_first_names += 1

            children_first_names = []
            for child in contracts.first().children.all():
                if child.name.name not in first_names:
                    first_names[child.name.name] = []
                    count_names += 1
                if child.name.first_name:
                    first_names[child.name.name].append(child.name.first_name)
                    children_first_names.append(child.name.first_name)
                    count_first_names += 1
            if children_first_names:
                if len(mietpartei) >= 10:
                    mietpartei = [mietpartei[0]] + [
                        ", ".join(x) for x in zip(mietpartei[1::2], mietpartei[2::2])
                    ]
                mietpartei.append("Kinder: %s" % ", ".join(children_first_names))

        if not first_names:
            continue

        lines = []
        for name in sorted(first_names):
            if first_names[name]:
                txt = ", %s" % (" & ".join(first_names[name]))
            else:
                txt = ""
            lines.append({"bold": name, "normal": txt})
        if len(mietpartei) > 10:
            raise RuntimeError("Mietpartei too long (%s lines)." % len(mietpartei))
        for _i in range(len(mietpartei), 10):
            mietpartei.append("")
        context["lines"] = lines
        context["mietpartei"] = mietpartei
        context["mietbeginn"] = contract.date.strftime("%d.%m.%Y")
        context["wohnungsnr"] = ru.name
        context["zimmer"] = ru.rooms

        zipcount += 1
        ret.append({"info": context["title"], "objects": context["mietpartei"]})
        if makezip:
            filename = fill_template_pod(template, context, output_format="odt")
            if not filename:
                raise RuntimeError("Could not fill template")
            zipfile_content.append(
                {
                    "file": filename,
                    "filename": "%s_%s.odt" % (options["beschreibung"], context["title"]),
                }
            )

    if makezip:
        ## Build ZIP-file from list of files
        file_like_object = io.BytesIO()
        zipfile_ob = zipfile.ZipFile(file_like_object, "w")
        for f in zipfile_content:
            # print "%s -> %s_%s" % (f['file'], f['member'].name.name,f['member'].name.first_name)
            zipfile_ob.write(f["file"], f["filename"])
        zipfile_ob.close()
        resp = HttpResponse(
            file_like_object.getvalue(), content_type="application/x-zip-compressed"
        )
        resp["Content-Disposition"] = "attachment; filename=%s" % "output.zip"
        return resp

    if zipcount > 0:
        link_text = "%s erstellen und herunterladen" % options["beschreibung"]
        ret.append(
            {
                "info": "Dokumente:",
                "objects": [
                    '<a href="%s?makezip=yes">%s (%d LibreOffice Dokumente in ZIP)</a>'
                    % (options["link_url"], link_text, zipcount)
                ],
            }
        )
    else:
        ret.append({"info": "Keine zu erstellenden Dokumente gefunden."})

    return render(
        request,
        "geno/default.html",
        {"response": ret, "title": "Mietobjekt-Dokumente erzeugen - %s" % options["beschreibung"]},
    )


@login_required
def check_portal_users(request):
    if "portal" not in settings.INSTALLED_APPS:
        raise ImproperlyConfigured(
            "check_portal_users() requires the portal app, with is not installed/enabled."
        )
    users = portal_auth.check_address_user_auth()
    ret = [
        {"info": "Benutzer ohne Login-Erlaubnis", "objects": users},
    ]
    return render(request, "geno/default.html", {"response": ret, "title": "Benutzer überprüfen"})


@login_required
def check_duplicate_invoices(request):
    ret = [
        {"info": "Rechnungs-Duplikate", "objects": get_duplicate_invoices()},
    ]
    return render(
        request, "geno/default.html", {"response": ret, "title": "Rechnungen überprüfen"}
    )


class Odt2PdfView(CohivaAdminViewMixin, FormView):
    title = "ODT in PDF umwandeln"
    form_class = Odt2PdfForm
    template_name = "geno/odt2pdf.html"
    permission_required = ("geno.tools_odt2pdf",)
    tmpdir = "/tmp/odt2pdf"

    def form_valid(self, form):
        if not os.path.isdir(self.tmpdir):
            os.mkdir(self.tmpdir)
        tmp_file = tempfile.NamedTemporaryFile(
            suffix=".odt", prefix="django_odt2pdf_input_", dir=self.tmpdir, delete=False
        )
        with open(tmp_file.name, "wb+") as destination:
            for chunk in self.request.FILES["file"].chunks():
                destination.write(chunk)
        # resp = HttpResponse(odt2pdf(tmp_file.name), content_type = "application/pdf")
        pdf_file_path = odt2pdf(tmp_file.name, self.request.user.get_username())
        pdf_file = open(pdf_file_path, "rb")
        pdf_file_name = self.request.FILES["file"].name[0:-4]
        resp = FileResponse(pdf_file, content_type="application/pdf")
        resp["Content-Disposition"] = "attachment; filename=%s.pdf" % pdf_file_name
        if os.path.isfile(tmp_file.name):
            os.remove(tmp_file.name)
        if os.path.isfile(pdf_file_path):
            os.remove(pdf_file_path)
        return resp


class WebstampView(CohivaAdminViewMixin, FormView):
    title = "PDFs frankieren"
    form_class = WebstampForm
    permission_required = ("geno.tools_webstmap",)
    tmpdir = "/tmp/webstamp"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.result = []
        self.download_file_url = None

    def get_initial(self):
        return {"stamp_type": "A-STANDARD-ENV"}

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["stamps_available"] = self.get_available_webstamps()
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(
            {
                "response": self.result,
                "form_title": (
                    "PDF Dateien hochladen. Die erste Seite wird frankiert (Fenster-Couvert links)"
                ),
                "download_file_url": self.download_file_url,
            }
        )
        return context

    def get(self, request, *args, **kwargs):
        if request.GET.get("download", None):
            return self.send_file(request.GET.get("download"))
        return super().get(request, *args, **kwargs)

    def form_valid(self, form):
        self.result = self.add_webstamps(
            form.cleaned_data["files"], form.cleaned_data["stamp_type"]
        )
        if isinstance(self.result, str):
            self.download_file_url = (
                f"{reverse('geno:webstamp')}?download={os.path.basename(self.result)}"
            )
        return self.get(self.request)

    def send_file(self, tmp_file_name):
        tmp_file_path = os.path.normpath(os.path.join(self.tmpdir, tmp_file_name))
        if not tmp_file_path.startswith(self.tmpdir):
            raise PermissionDenied()
        if os.path.isfile(tmp_file_path):
            pdf_file_name = "PDF_frankiert"
            resp = FileResponse(open(tmp_file_path, "rb"), content_type="application/pdf")
            resp["Content-Disposition"] = "attachment; filename=%s.pdf" % pdf_file_name
            os.remove(tmp_file_path)
            return resp
        else:
            raise Http404(f"File {tmp_file_name} not found.")

    # TODO: Refactor: Move this to a utility class
    def get_available_webstamps(self):
        stamp_names = {
            "A-GROSS-ENV": "A-Post Grossbrief (C4, bis 500g)",
            "A-GROSS-SCHWER-ENV": "A-Post Grossbrief schwer (C4, 500-1000g)",
            "A-STANDARD-ENV": "A-Post Standardbrief (bis C5, 100g)",
            "B-STANDARD-ENV": "B-Post Standardbrief (bis C5, 100g)",
            "ECONOMY-EURO1-50g-ENV": "Europa Economy (bis C5, 50g)",
        }
        cmd_out = subprocess.run(["/usr/local/bin/webstamp"], stdout=subprocess.PIPE)
        type_list = False
        stamps = {}
        pattern = re.compile(r"^\s+- (?P<type>\S+)\s+\((?P<num>\d+) available")
        for line in cmd_out.stdout.decode("utf-8").splitlines():
            # print(line)
            if line == "   stamp types:":
                type_list = True
            elif type_list:
                m = pattern.search(line)
                if m.group("type") and m.group("num"):
                    stamp_type = m.group("type")
                    stamp_name = stamp_names.get(stamp_type, stamp_type)
                    stamps[stamp_type] = "%s: %d verfügbar" % (stamp_name, int(m.group("num")))
        return stamps

    # TODO: Refactor: Move this to a utitlity class
    def add_webstamps(self, files, stamp_type):
        ret = []
        if not os.path.isdir(self.tmpdir):
            os.mkdir(self.tmpdir)
        tmp_files = []
        for f in files:
            tmp_file = tempfile.NamedTemporaryFile(
                suffix=".pdf", prefix="django_webstamp_input_", dir=self.tmpdir, delete=False
            )
            with open(tmp_file.name, "wb+") as destination:
                for chunk in f.chunks():
                    destination.write(chunk)
            tmp_files.append(tmp_file.name)
        cmd_out = subprocess.run(
            ["/usr/local/bin/webstamp", "-t", stamp_type] + tmp_files, stdout=subprocess.PIPE
        )
        ret.append(
            {
                "info": "Webstamp output:",
                "objects": ["<pre>%s</pre>" % cmd_out.stdout.decode("utf-8")],
            }
        )
        ## Check if output files are there
        for f in tmp_files:
            outfile = f[0:-4] + "_stamp.pdf"
            if os.path.isfile(outfile):
                os.rename(outfile, f)
            else:
                return ret

        ## Concatenate stamped PDFs and return one PDF with all pages.
        tmp_file = tempfile.NamedTemporaryFile(
            suffix=".pdf", prefix="django_webstamp_output_", dir=self.tmpdir, delete=False
        )
        pdfcat_out = subprocess.run(
            ["pdftk"] + tmp_files + ["cat", "output", tmp_file.name],
            stdout=subprocess.PIPE,
        )
        # print(pdfcat_out.stdout.decode('utf-8'))
        ret.append(
            {
                "info": "PDFtk output:",
                "objects": ["<pre>%s</pre>" % pdfcat_out.stdout.decode("utf-8")],
            }
        )
        for f in tmp_files:
            os.remove(f)
        if os.path.isfile(tmp_file.name):
            return tmp_file.name
        return ret


class SysadminView(CohivaAdminViewMixin, TemplateView):
    template_name = "geno/sysadmin_overview.html"
    permission_required = ("geno.sysadmin",)
    title = "Übersicht"


def oauth_client_test(request, action="start"):
    ret = []

    ## Secondary portal
    # base_url = settings.PORTAL_SECONDARY_HOST
    ## Main portal
    base_url = settings.BASE_URL

    client_id = ""
    client_secret = ""
    redirect_uri = f"{base_url}/geno/oauth_client/callback/"
    authorize_url = base_url + "/o/authorize/"
    access_token_url = base_url + "/o/token/"
    resource_url = base_url + "/portal/me/"

    ret.append({"info": "current django user: %s" % request.user})

    if action == "login":
        ## Start Authorization
        oauth = OAuth2Session(client_id, redirect_uri=redirect_uri)
        authorization_url, state = oauth.authorization_url(authorize_url)
        request.session["test_oauth_state"] = state
        request.session["test_oauth_origin_url"] = request.get_full_path()
        print("Redirecting to authorization URL: %s" % authorization_url)
        return redirect(authorization_url)
    elif action == "callback":
        oauth = OAuth2Session(
            client_id, state=request.session["test_oauth_state"], redirect_uri=redirect_uri
        )
        token = oauth.fetch_token(
            access_token_url,
            client_secret=client_secret,
            authorization_response=request.build_absolute_uri(
                settings.BASE_URL + request.get_full_path()
            ),
        )
        print("Got token: %s" % token)
        ## Save token to session
        request.session["test_oauth_token"] = token
        return redirect("/geno/oauth_client/test/")
    elif action == "access":
        if "test_oauth_token" not in request.session:
            ## Login first
            print("No token in session -> redirect to login")
            return oauth_client_test(request, action="login")

        access_token = request.session["test_oauth_token"]
        oauth = OAuth2Session(client_id, token=access_token)
        try:
            response = oauth.get(resource_url)
        except TokenExpiredError:
            try:
                ret.append({"info": "Refreshing token"})
                extra = {"client_id": client_id, "client_secret": client_secret}
                new_token = oauth.refresh_token(access_token_url, **extra)
                response = oauth.get(resource_url)
                ## Save token to session
                request.session["test_oauth_token"] = new_token
            except Exception:
                ## Could not refresh token
                print("Could not refresh token -> redirect to login")
                return oauth_client_test(request, action="login")

        if response.status_code == 200:
            ret.append({"info": "Response: %s" % response.json()})
            # for item in response.json():
            #    ret.append({'info': '%s' % (item)})
        else:
            ret.append(
                {
                    "info": "HTTP response status_code ist not 200!",
                    "objects": [
                        "status_code: %s" % response.status_code,
                    ],
                }
            )
    else:  ## Start
        if "test_oauth_token" in request.session:
            print("Forgeting token.")
            del request.session["test_oauth_token"]
        return oauth_client_test(request, action="login")
    return render(request, "geno/default.html", {"response": ret, "title": "Oauth Client Test"})


def preview_template(request):
    invoice_category_id = request.GET.get("invoice_category", None)
    contract_id = request.GET.get("contract", None)
    if invoice_category_id and contract_id:
        invoice_category = InvoiceCategory.objects.get(id=invoice_category_id)
        mail_template = invoice_category.email_template.get_template()
        contract = Contract.objects.get(id=contract_id)
        address = contract.get_contact_address()
        context = address.get_context()
        mail_text_html = mail_template.render(Context(context))
        return HttpResponse(mail_text_html)
    else:
        return HttpResponse("Vorschau wird noch nicht unterstützt.")


def registration(request, registration_id=None):
    template_data = {}
    template_data["forms"] = process_registration_forms(request, selector=registration_id)
    return render(request, "website/registration_form.html", template_data)
